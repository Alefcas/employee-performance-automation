# -*- coding: utf-8 -*-
# ============================================================
#  GERADOR DE DADOS FICTICIOS  -  TOP AMBASSADORS PROJECT
#  ------------------------------------------------------------
#  Este script cria TODOS os dados ficticios usados pela versao
#  de portfolio (demo) do TOP AMBASSADORS. NENHUM dado real da
#  Amazon e utilizado aqui - nomes, logins, IDs e metricas sao
#  totalmente inventados.
#
#  Ele gera:
#    - associados.xlsx            (planilha de Onboarding ficticia)
#    - dados_funcionarios.xlsx    (planilha de Produtividade ficticia)
#    - employee_roster.csv        (Employee Roster ficticio)
#    - atlas_totals_combined.csv  (ATLAS - Totals Report)
#    - atlas_pack_single.csv / atlas_pack_multi.csv
#    - atlas_raw_<processo>.csv   (ATLAS - Raw Reports por processo)
#    - ExtracaoATLAS.xlsx         (consolidado, para o modo "Reaproveitar ATLAS")
#    - totais_processo.json       (targets de qualidade reaproveitaveis)
#
#  Os arquivos de "download" ficam em  SITE OF ALEFCAS/downloads/
#  e sao servidos pelos sites HTML falsos. Copias das planilhas
#  principais tambem ficam na raiz do projeto para uso imediato.
#
#  COMPOSICAO DA DEMO (para maior visualizacao dos resultados):
#    - 5 TOP em qualidade      + 2 TOP em produtividade   =  7 TOP
#    - 1 BOTTOM em qualidade   + 3 BOTTOM em produtividade =  4 BOTTOM
#    - 1 INVALIDO (TOP prod + BOTTOM qual, sem reconhecimento)
#    - 4 MEIO em qualidade     + 3 MEIO em produtividade
#
#  Cada embaixador pertence a um dos dois turnos simulados:
#    "Turno Dia"  ou  "Turno Noite".  O identificador do galpao
#    (antigo warehouse) e simulado como "FC locale".
# ============================================================
import os
import csv
import json
from pathlib import Path

from openpyxl import Workbook

AQUI        = Path(__file__).resolve().parent          # SITE OF ALEFCAS
PROJETO     = AQUI.parent                              # TOP AMBASSADORS PROJECT
RECURSOS    = PROJETO / "2. RECURSOS"
DOWNLOADS   = AQUI / "downloads"
DOWNLOADS.mkdir(parents=True, exist_ok=True)

# ------------------------------------------------------------
# PARAMETROS GERAIS DA SIMULACAO
# ------------------------------------------------------------
FC_ID       = "FC locale"      # substitui o antigo "GRU5" / warehouse id
TURNO_DIA   = "Turno Dia"
TURNO_NOITE = "Turno Noite"
LC_LEVEL    = "LC1-LC3"
OPPS        = 5000             # opportunities por associado (qualidade)

# Targets de DPMO por processo (Totals Report).
# Quality Rate = (DPMO do embaixador / target) * 100.
#   TOP <= 30%   |   MEIO 30%..100%   |   BOTTOM > 100%
TARGETS = {
    "Pick":    1200,
    "Sort":    1000,
    "Stow":    1500,
    "Ship":     900,
    "Receive": 1100,
}
TARGET_PACK_SINGLE = 1800
TARGET_PACK_MULTI  = 6000

# Defects por associado conforme o "destaque" de qualidade desejado.
# (opps fixo em 5000; a razao defects/opps define o DPMO e, logo, o %)
DEFECTS_POR_TIPO = {
    "QUAL_TOP":  1,    # DPMO baixo  -> % bem abaixo de 30
    "QUAL_MEIO": 4,    # DPMO medio  -> % entre 30 e 100
    "QUAL_BOT": 10,    # DPMO alto   -> % acima de 100
    "INVALIDO":  8,    # qualidade ruim (bottom) apesar da otima produtividade
}

# Rate na LC por associado conforme o destaque de produtividade.
#   TOP >= 90%   |   MEIO 60%..90%   |   BOTTOM <= 60%
RATE_POR_TIPO = {
    "PROD_TOP":  1.12,   # ~112%
    "PROD_MEIO": 0.75,   # ~75%
    "PROD_BOT":  0.55,   # ~55%
    "INVALIDO":  1.06,   # ~106% (otima produtividade)
}

TIPOS_QUALIDADE     = ("QUAL_TOP", "QUAL_MEIO", "QUAL_BOT", "INVALIDO")
TIPOS_PRODUTIVIDADE = ("PROD_TOP", "PROD_MEIO", "PROD_BOT", "INVALIDO")

# ------------------------------------------------------------
# EMBAIXADORES (tudo ficticio)
# Cada um: nome, tipo (destaque), processo, turno, base de login.
# 3 associados por embaixador (logins = base + "1"/"2"/"3").
# Qualidade usa apenas Pick/Sort/Stow/Receive (processos com target
# direto). Produtividade pode usar Pack Single/Multi tambem.
# ------------------------------------------------------------
EMBAIXADORES = [
    # ===== 5 TOP em QUALIDADE =====
    {"nome": "Ana Beatriz Souza",   "tipo": "QUAL_TOP", "processo": "Pick",    "turno": TURNO_DIA,   "base": "anasouza"},
    {"nome": "Bruno Carvalho Nunes","tipo": "QUAL_TOP", "processo": "Sort",    "turno": TURNO_NOITE, "base": "brunonun"},
    {"nome": "Camila Ferreira Rocha","tipo": "QUAL_TOP","processo": "Stow",    "turno": TURNO_DIA,   "base": "camilarc"},
    {"nome": "Daniel Oliveira Costa","tipo": "QUAL_TOP","processo": "Receive", "turno": TURNO_NOITE, "base": "danieloc"},
    {"nome": "Elaine Martins Alves","tipo": "QUAL_TOP", "processo": "Pick",    "turno": TURNO_DIA,   "base": "elainema"},

    # ===== 4 MEIO em QUALIDADE =====
    {"nome": "Gabriela Santos Pinto","tipo": "QUAL_MEIO","processo": "Sort",   "turno": TURNO_NOITE, "base": "gabrisp"},
    {"nome": "Henrique Alves Moreira","tipo":"QUAL_MEIO","processo": "Stow",   "turno": TURNO_DIA,   "base": "henriqam"},
    {"nome": "Isabela Costa Ramos", "tipo": "QUAL_MEIO","processo": "Receive", "turno": TURNO_NOITE, "base": "isabelcr"},
    {"nome": "Joao Pedro Almeida",  "tipo": "QUAL_MEIO","processo": "Pick",    "turno": TURNO_DIA,   "base": "joaopa"},

    # ===== 1 BOTTOM em QUALIDADE =====
    {"nome": "Carlos Mendes Lima",  "tipo": "QUAL_BOT", "processo": "Stow",    "turno": TURNO_NOITE, "base": "carloslm"},

    # ===== 2 TOP em PRODUTIVIDADE =====
    {"nome": "Fernanda Rocha Dias", "tipo": "PROD_TOP", "processo": "Pack Singles", "turno": TURNO_DIA,   "base": "fernrd"},
    {"nome": "Rafael Souza Barros", "tipo": "PROD_TOP", "processo": "Pick",         "turno": TURNO_NOITE, "base": "rafaelsb"},

    # ===== 3 MEIO em PRODUTIVIDADE =====
    {"nome": "Eduardo Santos Melo", "tipo": "PROD_MEIO","processo": "Sort",         "turno": TURNO_NOITE, "base": "edusm"},
    {"nome": "Patricia Lima Souza", "tipo": "PROD_MEIO","processo": "Pack Multi",   "turno": TURNO_DIA,   "base": "patrils"},
    {"nome": "Thiago Ferreira Dias","tipo": "PROD_MEIO","processo": "Pick",         "turno": TURNO_NOITE, "base": "thiagofd"},

    # ===== 3 BOTTOM em PRODUTIVIDADE =====
    {"nome": "Diego Alves Pinto",   "tipo": "PROD_BOT", "processo": "Pick",         "turno": TURNO_DIA,   "base": "diegoap"},
    {"nome": "Larissa Gomes Faria", "tipo": "PROD_BOT", "processo": "Sort",         "turno": TURNO_NOITE, "base": "larissgf"},
    {"nome": "Marcos Vinicius Rocha","tipo":"PROD_BOT", "processo": "Stow",         "turno": TURNO_DIA,   "base": "marcosvr"},

    # ===== 1 INVALIDO (TOP prod + BOTTOM qual) =====
    {"nome": "Vanessa Ribeiro Cruz","tipo": "INVALIDO", "processo": "Pick",         "turno": TURNO_DIA,   "base": "vanessrc"},
]


def logins_do(emb):
    return [emb["base"] + str(i) for i in (1, 2, 3)]

def eh_qualidade(emb):
    return emb["tipo"] in TIPOS_QUALIDADE

def eh_produtividade(emb):
    return emb["tipo"] in TIPOS_PRODUTIVIDADE


# ------------------------------------------------------------
# associados.xlsx  (formato "Modelo": header na linha 0, 21 colunas)
#   Col 5 = Shift pattern | Col 7 = Login | Col 11 = Processo
#   Col 14 = TWI Embaixador
#   TODOS os embaixadores entram aqui (mapa login -> embaixador).
# ------------------------------------------------------------
def gerar_associados(destinos):
    header = ["RH YES/NO", "Start Date", "First Name", "Middle Name", "Last Name",
              "Shift pattern", "Social Name", "Login", "EmplID", "Agencia", "",
              "Processo", "Cracha", "Smile Maker", "TWI Embaixador",
              "Presenca Day 1", "Presenca Day 2", "Presenca Day 3",
              "Registro Trein", "Entrega certi", "Barreiras Obse"]
    linhas = [header]
    empid = 700100
    for emb in EMBAIXADORES:
        for login in logins_do(emb):
            empid += 7
            row = [""] * 21
            row[0]  = "YES"
            row[1]  = "30/06/2026"
            row[2]  = login.capitalize()
            row[4]  = "Demo"
            row[5]  = emb["turno"]
            row[7]  = login
            row[8]  = str(empid)
            row[9]  = "Alefcas Staffing"
            row[11] = emb["processo"]
            row[14] = emb["nome"]
            row[15] = "YES"
            row[16] = "YES"
            row[17] = "YES"
            linhas.append(row)
    for destino in destinos:
        wb = Workbook()
        ws = wb.active
        ws.title = "BD 30.06"
        for r in linhas:
            ws.append(r)
        wb.save(str(destino))
        print("   OK ->", destino)


# ------------------------------------------------------------
# dados_funcionarios.xlsx  (3 linhas de titulo + header na linha 3)
#   Col 0 = FC | Col 3 = Login | Col 4 = Shift Pattern
#   Col 5 = Process Name | Col 12 = Rate na LC
#   Entram: PROD_TOP, PROD_MEIO, PROD_BOT e INVALIDO.
# ------------------------------------------------------------
def gerar_dados_funcionarios(destinos):
    header = ["FC", "Week", "Employee ID", "Login", "Shift Pattern", "Process Name",
              "LC Level", "Expected LC", "size_category", "Units", "Hours",
              "Rate Real", "Rate na LC", "UPH PPR", "TARGET_PPR",
              "UPH_RATE %", "PPR/LC %"]
    linhas = [
        ["Rate dos Associados por Semana"] + [""] * 16,
        ["Considerar a Coluna Rate na LC"] + [""] * 16,
        [""] * 17,
        header,
    ]
    empid = 800200
    for emb in EMBAIXADORES:
        if not eh_produtividade(emb):
            continue
        rate_lc = RATE_POR_TIPO[emb["tipo"]]
        for login in logins_do(emb):
            empid += 5
            units = 2600
            hours = 40.0
            rate_real = round(rate_lc * 0.90, 4)
            uph_ppr = round(units / hours, 4)
            row = [FC_ID, 29, str(empid), login, emb["turno"], emb["processo"],
                   LC_LEVEL, 0.90, "", units, round(hours, 4),
                   rate_real, round(rate_lc, 4), uph_ppr, 32.076, "", ""]
            linhas.append(row)
    for destino in destinos:
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        for r in linhas:
            ws.append(r)
        wb.save(str(destino))
        print("   OK ->", destino)


# ------------------------------------------------------------
# Employee Roster ficticio (Col A = Employee ID, B = User ID, C = Name)
#   Inclui os embaixadores (para preencher o historico) e todos
#   os associados. A coluna Department guarda o turno.
# ------------------------------------------------------------
EMPL_TIPOS = ["EMPL1", "EMPL2", "EMPL3"]

def gerar_roster(destino):
    rows = [["Employee ID", "User ID", "Employee Name",
             "Department ID", "Employment Type", "Employee Status", "Manager Name"]]
    empid = 900300
    idx = 0
    for emb in EMBAIXADORES:
        empid += 2
        rows.append([str(empid), emb["nome"], emb["nome"],
                     emb["turno"], EMPL_TIPOS[idx % 3], "Active", emb["nome"]])
        idx += 1
    for emb in EMBAIXADORES:
        for login in logins_do(emb):
            empid += 3
            nome = login.capitalize() + " Demo"
            rows.append([str(empid), login, nome, emb["turno"], EMPL_TIPOS[idx % 3],
                         "Active", emb["nome"]])
            idx += 1
    with open(destino, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerows(rows)
    print("   OK ->", destino)


# ------------------------------------------------------------
# ATLAS - Totals Report (combined)
#   Col A = Defect Type ("Total Pick"...) | Col E (idx4) = DPMO/target
# ------------------------------------------------------------
def gerar_atlas_totals(destino):
    header = ["Defect Type", "Process Path", "Defect Count",
              "Opportunities", "DPMO/Percentage", "Threshold"]
    rows = [header]
    for proc, tgt in TARGETS.items():
        rows.append(["Total " + proc, proc, 25000, 12000000, tgt, "-"])
    rows.append(["Total Pack", "Pack", 26000, 1900000, 15807, "-"])
    with open(destino, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)
    print("   OK ->", destino)

def gerar_atlas_pack_total(destino, valor):
    header = ["Defect Type", "Process Path", "Defect Count",
              "Opportunities", "DPMO/Percentage", "Threshold"]
    rows = [header, ["Total", "Pack", 12000, 1000000, valor, "-"]]
    with open(destino, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)
    print("   OK ->", destino)


# ------------------------------------------------------------
# ATLAS - Raw Reports por processo
#   Colunas: User, Manager ID, Warehouse ID, Subprocess,
#            totalDefects, Opportunities, DPMO/Percentage
#   Entram: QUAL_TOP, QUAL_MEIO, QUAL_BOT e INVALIDO.
# ------------------------------------------------------------
def _assoc_qualidade_por_processo():
    """Agrupa os associados com dados de QUALIDADE por processo do ATLAS."""
    por_proc = {"Pack": [], "Pick": [], "Sort": [], "Stow": [], "Ship": [], "Receive": []}
    for emb in EMBAIXADORES:
        if not eh_qualidade(emb):
            continue
        proc    = emb["processo"]
        defects = DEFECTS_POR_TIPO[emb["tipo"]]
        for login in logins_do(emb):
            por_proc.setdefault(proc, []).append((login, OPPS, defects, "Single"))
    return por_proc

def gerar_atlas_raw(pasta):
    header = ["User", "Manager ID", "Warehouse ID", "Subprocess",
              "totalDefects", "Opportunities", "DPMO/Percentage"]
    por_proc = _assoc_qualidade_por_processo()
    caminhos = {}
    for proc in ["Pack", "Pick", "Sort", "Stow", "Ship", "Receive"]:
        rows = [header]
        for login, opps, defects, sub in por_proc.get(proc, []):
            dpmo = round((defects / opps) * 1000000, 2) if opps else 0
            rows.append([login, "mgralef", FC_ID, sub, defects, opps, dpmo])
        destino = pasta / ("atlas_raw_" + proc.lower() + ".csv")
        with open(destino, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerows(rows)
        caminhos[proc] = destino
        print("   OK ->", destino)
    return caminhos


# ------------------------------------------------------------
# ExtracaoATLAS.xlsx consolidado (para o modo "Reaproveitar ATLAS")
#   Formato em blocos: [processo] / cabecalho / dados / linha vazia
# ------------------------------------------------------------
def gerar_extracao_atlas(destino):
    OUTPUT_COLS = ["User", "Subprocess", "Total Defects", "Opportunities", "DPMO"]
    por_proc = _assoc_qualidade_por_processo()
    wb = Workbook()
    ws = wb.active
    ws.title = "ATLAS"
    for proc in ["Pack", "Pick", "Sort", "Stow", "Ship", "Receive"]:
        assocs = por_proc.get(proc, [])
        if not assocs:
            continue
        ws.append([proc, "", "", "", ""])
        ws.append(OUTPUT_COLS)
        for login, opps, defects, sub in assocs:
            dpmo = round((defects / opps) * 1000000, 2) if opps else 0
            ws.append([login, sub, defects, opps, dpmo])
        ws.append(["", "", "", "", ""])
    wb.save(str(destino))
    print("   OK ->", destino)

def gerar_totais_json(destino):
    dados = {proc.upper(): tgt for proc, tgt in TARGETS.items()}
    dados["PACK_SINGLE"] = TARGET_PACK_SINGLE
    dados["PACK_MULTI"]  = TARGET_PACK_MULTI
    payload = {"start_date": "2026-06-16", "end_date": "2026-06-30",
               "totais": dados}
    with open(destino, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    print("   OK ->", destino)


# ------------------------------------------------------------
def _remover_antigos():
    """Remove arquivos de versoes anteriores (nomes com GRU5) para
    evitar confusao apos a mudanca para 'FC locale'."""
    antigos = [DOWNLOADS / "employee_roster_GRU5.csv"]
    for a in antigos:
        try:
            if a.exists():
                a.unlink()
                print("   removido antigo ->", a)
        except Exception:
            pass


def _resumo_composicao():
    from collections import Counter
    c = Counter(e["tipo"] for e in EMBAIXADORES)
    print("   Composicao dos embaixadores (ficticios):")
    print("      TOP qualidade .......: " + str(c.get("QUAL_TOP", 0)))
    print("      TOP produtividade ...: " + str(c.get("PROD_TOP", 0)))
    print("      BOTTOM qualidade ....: " + str(c.get("QUAL_BOT", 0)))
    print("      BOTTOM produtividade : " + str(c.get("PROD_BOT", 0)))
    print("      INVALIDO ............: " + str(c.get("INVALIDO", 0)))
    print("      MEIO qualidade ......: " + str(c.get("QUAL_MEIO", 0)))
    print("      MEIO produtividade ..: " + str(c.get("PROD_MEIO", 0)))
    print("      Turnos ..............: " + TURNO_DIA + " / " + TURNO_NOITE)
    print("      FC (warehouse) ......: " + FC_ID)


def main():
    print("\n== Gerando dados ficticios do TOP AMBASSADORS PROJECT ==\n")

    _resumo_composicao()
    print()
    _remover_antigos()

    print(" associados.xlsx:")
    gerar_associados([DOWNLOADS / "associados.xlsx", PROJETO / "associados.xlsx"])

    print(" dados_funcionarios.xlsx:")
    gerar_dados_funcionarios([DOWNLOADS / "dados_funcionarios.xlsx",
                              PROJETO / "dados_funcionarios.xlsx"])

    print(" Employee Roster:")
    gerar_roster(DOWNLOADS / "employee_roster.csv")

    print(" ATLAS Totals / Pack:")
    gerar_atlas_totals(DOWNLOADS / "atlas_totals_combined.csv")
    gerar_atlas_pack_total(DOWNLOADS / "atlas_pack_single.csv", TARGET_PACK_SINGLE)
    gerar_atlas_pack_total(DOWNLOADS / "atlas_pack_multi.csv", TARGET_PACK_MULTI)

    print(" ATLAS Raw Reports:")
    gerar_atlas_raw(DOWNLOADS)

    print(" Consolidado / targets (modo reaproveitar ATLAS):")
    gerar_extracao_atlas(RECURSOS / "ExtracaoATLAS.xlsx")
    gerar_totais_json(RECURSOS / "totais_processo.json")

    print("\n== Concluido. Todos os dados sao ficticios. ==\n")

if __name__ == "__main__":
    main()
