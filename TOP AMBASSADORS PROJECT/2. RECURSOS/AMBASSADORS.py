# -*- coding: utf-8 -*-
import os
import subprocess
import sys


# ============================================
# VERIFICAR E INSTALAR DEPENDENCIAS
# ============================================
def verificar_dependencias():
    # Dependencias multiplataforma (funcionam em qualquer SO)
    dependencias = {
        "pandas": "pandas",
        "openpyxl": "openpyxl",
        "selenium": "selenium",
        "webdriver_manager": "webdriver-manager",
        "PIL": "Pillow",
    }
    # pywin32 e usado apenas no Windows (clipboard HTML / automacao do Excel).
    # Em Linux/Mac o programa usa alternativas, entao nao instalamos la.
    if os.name == "nt":
        dependencias["win32clipboard"] = "pywin32"
    faltando = []
    for modulo, pacote in dependencias.items():
        try:
            __import__(modulo)
        except ImportError:
            faltando.append(pacote)

    if not faltando:
        return
    print()
    print("=" * 55)
    print("   INSTALANDO DEPENDENCIAS FALTANTES")
    print("=" * 55)
    print("   Pacotes: " + ", ".join(faltando))
    print()

    for pacote in faltando:
        print("   Instalando " + pacote + "...", end="", flush=True)
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pacote],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            print(" OK")
        except Exception as e:
            print(" ERRO: " + str(e))
    print()
    print("   Dependencias verificadas!")
    print("=" * 55)


verificar_dependencias()


import math
import os
import re
import time
import json
import threading
import configparser
import shutil
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta


# NOTA: selenium e webdriver_manager sao importados sob demanda (lazy)
# dentro das funcoes que usam navegador. Isso acelera muito a abertura
# dos modos que NAO usam o Firefox (email pronto, mala direta).


# ============================================
# CORES NO TERMINAL (ANSI) + UI HELPERS
# Padrao Amazon: laranja, azul, verde, vermelho
# ============================================
def _ativar_cores_windows():
    if os.name == "nt":
        try:
            import ctypes
            k = ctypes.windll.kernel32
            k.SetConsoleMode(k.GetStdHandle(-11), 7)  # ENABLE_VIRTUAL_TERMINAL_PROCESSING
        except Exception:
            pass

_ativar_cores_windows()

class C:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    ITALIC  = "\033[3m"
    ORANGE  = "\033[38;5;208m"   # laranja Amazon
    BLUE    = "\033[38;5;39m"    # azul claro
    SKY     = "\033[38;5;117m"   # azul ceu (comentarios)
    CYAN    = "\033[38;5;87m"    # ciano
    GREEN   = "\033[38;5;42m"    # verde sucesso
    RED     = "\033[38;5;203m"   # vermelho
    GOLD    = "\033[38;5;220m"   # dourado
    GRAY    = "\033[38;5;245m"   # cinza
    WHITE   = "\033[97m"
    # Fundos
    BG_ORANGE = "\033[48;5;208m"
    BG_GREEN  = "\033[48;5;28m"
    BG_BLUE   = "\033[48;5;24m"


def titulo(texto):
    """Cabecalho de secao limpo, em laranja Amazon."""
    print()
    print(C.ORANGE + C.BOLD + "  " + texto + C.RESET)
    print(C.GRAY + "  " + ("\u2500" * 48) + C.RESET)


def ok(texto):
    print(C.GREEN + "  \u2714 " + C.RESET + texto)

def info(texto):
    print(C.BLUE + "  \u2022 " + C.RESET + texto)

def dica(texto):
    """Comentario/observacao em azul claro (sky)."""
    print(C.SKY + C.ITALIC + "  \u203a " + texto + C.RESET)

def aviso(texto):
    print(C.GOLD + "  ! " + C.RESET + texto)

def erro(texto):
    print(C.RED + "  \u2717 " + C.RESET + texto)

def passo(texto):
    print(C.GRAY + "    " + texto + C.RESET)


def perguntar(texto):
    """Input com prompt em laranja (toda interacao com o usuario)."""
    try:
        return input(C.ORANGE + C.BOLD + "  " + texto + C.RESET)
    except EOFError:
        return ""


def arquivo_esta_aberto(caminho):
    """Detecta se um .xlsx esta aberto no Excel.
    O Excel cria um arquivo de lock '~$nome.xlsx' na mesma pasta.
    Tambem tenta abrir em modo exclusivo como verificacao extra."""
    p = Path(caminho)
    lock = p.parent / ("~$" + p.name)
    if lock.exists():
        return True
    # Tentativa de abrir para escrita (se travado, levanta excecao)
    try:
        with open(str(p), "a"):
            pass
        return False
    except (PermissionError, OSError):
        return True


def garantir_arquivo_fechado(caminho, nome_amigavel=None):
    """Se o arquivo estiver aberto, avisa o usuario de forma clara e
    espera ele fechar antes de continuar."""
    nome_amigavel = nome_amigavel or os.path.basename(str(caminho))
    primeira = True
    while arquivo_esta_aberto(caminho):
        if primeira:
            print()
            aviso("O arquivo '" + nome_amigavel + "' parece estar ABERTO no Excel.")
            dica("Feche o arquivo para o programa ler os dados mais recentes.")
            primeira = False
        resp = perguntar("   Ja fechou? (ENTER = verificar de novo / P = pular): ").strip().lower()
        if resp == "p":
            aviso("Continuando com o arquivo possivelmente aberto...")
            return
    ok("'" + nome_amigavel + "' fechado. Prosseguindo.")


def fechar_excel_arquivo(caminho):
    """Fecha SOMENTE a pasta de trabalho indicada (se estiver aberta no Excel),
    sem fechar o Excel inteiro nem outros arquivos. Usa automacao COM (Windows).
    Falha em silencio caso o Excel/COM nao esteja disponivel."""
    if os.name != "nt":
        return
    try:
        import win32com.client  # parte do pywin32
    except Exception:
        return
    alvo = os.path.normcase(os.path.abspath(str(caminho)))
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return  # Excel nao esta aberto
    try:
        for wb in list(excel.Workbooks):
            try:
                if os.path.normcase(os.path.abspath(wb.FullName)) == alvo:
                    wb.Close(SaveChanges=False)
                    break
            except Exception:
                continue
        # Se nao sobrou nenhuma pasta aberta, encerra o Excel
        try:
            if excel.Workbooks.Count == 0:
                excel.Quit()
        except Exception:
            pass
    except Exception:
        pass


# ============================================
# CONFIGURACOES - GERAL
# ============================================
# RAIZ do projeto: este arquivo fica em "2. RECURSOS",
# entao a pasta do projeto (4. TOP AMBASSADORS) e a pasta-pai.
# Se o arquivo estiver na propria raiz (modo antigo), detecta tambem.
DOWNLOAD_DIR      = str(Path.home() / "Downloads")
_AQUI             = Path(__file__).resolve().parent
if _AQUI.name.upper().startswith("2. RECURSOS") or _AQUI.name.upper() == "2. RECURSOS":
    SCRIPT_DIR    = _AQUI.parent          # arquivo dentro de 2. RECURSOS
elif (_AQUI / "2. RECURSOS").exists():
    SCRIPT_DIR    = _AQUI                 # arquivo na raiz do projeto
else:
    SCRIPT_DIR    = _AQUI.parent if (_AQUI.parent / "2. RECURSOS").exists() else _AQUI
RECURSOS_DIR       = SCRIPT_DIR / "2. RECURSOS"    # arquivos gerados pelo programa
HISTORICO_DIR      = SCRIPT_DIR / "1. HISTÓRICO"   # historico acumulado
# DEMO / PORTFOLIO: pasta com os sites HTML falsos ("Site of Alefcas")
SITE_ALEFCAS_DIR   = SCRIPT_DIR / "SITE OF ALEFCAS"
PLANILHA_HISTORICO = "HistoricoAmbassadors.xlsx"
WAREHOUSE_ID      = "FC locale"
DOWNLOAD_TIMEOUT  = 120
URL_ATLAS         = "https://atlas.qubit.amazon.dev/reporting"
# URLs de download das planilhas principais (lidas do config.ini)
URL_ASSOCIADOS    = "https://amazon-my.sharepoint.com/:x:/p/terto/IQB22IYz-6yRT6ql5lVe-YPHAepLSz2-0qBK6ZStac9b3iM?e=kXGUyY"
URL_DADOS         = "https://us-east-1.quicksight.aws.amazon.com/sn/account/amazonbi/dashboards/70d2ff95-0852-4002-8afb-8d9c0cff2218/sheets/70d2ff95-0852-4002-8afb-8d9c0cff2218_8f36a6db-5acc-4ba5-8a8d-aa7cecb62b8b"


COMBINED_PROCESSES  = ["Pack", "Pick", "Sort", "Stow", "Ship", "Receive"]
PLANILHA_ASSOCIADOS = "associados.xlsx"
PLANILHA_ATLAS      = "ExtracaoATLAS.xlsx"
IDX_ASSOCIADO       = 7
IDX_EMBAIXADOR      = 14
IDX_TURNO_ASSOC     = 5    # Coluna F - Shift Pattern (associados.xlsx)
MIN_OPPORTUNITIES   = 100
TOTAL_PREFIX        = "Total "


# ============================================
# CONFIGURACOES - REGRAS DE NEGOCIO
# ============================================
MIN_ASSOC_POR_EMBAIXADOR = 3
MAX_ASSOC_POR_EMBAIXADOR = 5


# RC SORT fica separado do fluxo normal (so vai para rc_cret.xlsx).
# C-RET NAO entra mais aqui: e uma area que tambem recebe reconhecimento,
# entao segue o fluxo normal de produtividade (ranking, email, cartas).
PROCESSOS_RC_CRET = ["RC SORT", "RC-SORT", "RCSORT"]


# Mapeamento de normalizacao de nomes de processos
# Pack e a unica excecao: SEMPRE Pack Single ou Pack Multi, nunca so "Pack"
PROCESSO_NORMALIZADO = {
    "pick":           "Pick",
    "pack single":    "Pack Single",
    "pack singles":   "Pack Single",
    "pack multi":     "Pack Multi",
    "pack multis":    "Pack Multi",
    "sort":           "Sort",
    "sort bath":      "Sort",
    "stow":           "Stow",
    "stow to prime":  "Stow",
    "receive":        "Receive",
    "each - receive": "Receive",
    "each receive":   "Receive",
    "each-receive":   "Receive",
    "ship":           "Doca",
    "transferout":    "Doca",
    "transfer out":   "Doca",
    "transfer-out":   "Doca",
    "rc sort":        "RC Sort",
    "rc-sort":        "RC Sort",
    "rcsort":         "RC Sort",
    "c-ret":          "C-Ret",
    "cret":           "C-Ret",
    "c ret":          "C-Ret",
}


QUALITY_TOP_MAX    = 60
QUALITY_BOTTOM_MIN = 100


# Mapeamento de processo normalizado -> chave no totais_processo (Totals Report)
PROCESSO_TO_TOTAIS_KEY = {
    "Pick":        "PICK",
    "Pack Single": "PACK_SINGLE",
    "Pack Multi":  "PACK_MULTI",
    "Sort":        "SORT",
    "Stow":        "STOW",
    "Receive":     "RECEIVE",
    "Doca":        "SHIP",
    "RC Sort":     "RC SORT",
    "C-Ret":       "C-RET",
}


PROD_TOP_MINIMO    = 90
PROD_BOTTOM_MAXIMO = 60


HORAS_MINIMAS = 20


# ============================================
# CONFIGURACOES - PRODUTIVIDADE (indices)
# ============================================
PLANILHA_DADOS     = "dados_funcionarios.xlsx"
IDX_LOGIN_FUNC     = -1
IDX_NOME_FUNC      = 3
IDX_TURNO_FUNC     = 4
IDX_PROCESSO_FUNC  = 5
IDX_LC_LEVEL_FUNC  = 6
IDX_UPH_PPR_FUNC   = 13
IDX_HORAS_FUNC     = 10
IDX_RATE_REAL_FUNC = 11
IDX_RATE_LC_FUNC   = 12
IDX_TARGET_PPR_FUNC = 14   # Coluna O - TARGET_PPR (dados_funcionarios.xlsx)


# ============================================
# CONFIGURACOES - ARQUIVOS DE SAIDA
# ============================================
PLANILHA_SAIDA_COMPLETO = "resultado_completo.xlsx"
PLANILHA_SAIDA_RC_CRET  = "rc_cret.xlsx"
PLANILHA_SAIDA_FINAL    = "ranking_final.xlsx"


# ============================================
# CRIAR ESTRUTURA DE PASTAS
# Recursos/  -> tudo que o programa gera
# Resultados/ -> backups do ranking final (DD.MM.xlsx)
# ============================================
def criar_estrutura_pastas():
    for pasta in [RECURSOS_DIR, HISTORICO_DIR]:
        pasta.mkdir(parents=True, exist_ok=True)
    print("   Pastas verificadas:")
    print("      2. RECURSOS  -> " + str(RECURSOS_DIR))
    print("      1. HISTORICO -> " + str(HISTORICO_DIR))


# ============================================
# CARREGAR CONFIG.INI
# Le o arquivo config.ini da mesma pasta do script.
# Se nao encontrado, usa os valores padrao acima.
# Cada parametro usa fallback= para nunca quebrar.
# ============================================
def _resolver_url_local(valor):
    """DEMO / PORTFOLIO: converte URLs do tipo 'local:arquivo.html' para um
    caminho file:/// absoluto apontando para a pasta 'SITE OF ALEFCAS'.
    Assim os sites falsos funcionam em qualquer PC apos clonar do GitHub.
    URLs normais (http/https) sao devolvidas sem alteracao."""
    if not valor:
        return valor
    valor = str(valor).strip()
    if valor.lower().startswith("local:"):
        nome_arquivo = valor.split(":", 1)[1].strip()
        caminho = (SITE_ALEFCAS_DIR / nome_arquivo).resolve()
        return caminho.as_uri()   # ex: file:///D:/.../SITE%20OF%20ALEFCAS/roster.html
    return valor


def carregar_config():
    global WAREHOUSE_ID, DOWNLOAD_TIMEOUT, MIN_OPPORTUNITIES
    global PLANILHA_ASSOCIADOS, PLANILHA_ATLAS, PLANILHA_DADOS
    global PLANILHA_SAIDA_COMPLETO, PLANILHA_SAIDA_RC_CRET, PLANILHA_SAIDA_FINAL
    global QUALITY_TOP_MAX, QUALITY_BOTTOM_MIN
    global PROD_TOP_MINIMO, PROD_BOTTOM_MAXIMO
    global MIN_ASSOC_POR_EMBAIXADOR, MAX_ASSOC_POR_EMBAIXADOR, HORAS_MINIMAS
    global URL_ROSTER, URL_ASSOCIADOS, URL_DADOS, URL_ATLAS
    global IDX_ASSOCIADO, IDX_EMBAIXADOR, IDX_TURNO_ASSOC
    global IDX_LOGIN_FUNC, IDX_NOME_FUNC, IDX_TURNO_FUNC, IDX_PROCESSO_FUNC
    global IDX_LC_LEVEL_FUNC, IDX_UPH_PPR_FUNC, IDX_HORAS_FUNC
    global IDX_RATE_REAL_FUNC, IDX_RATE_LC_FUNC, IDX_TARGET_PPR_FUNC


    config_path = RECURSOS_DIR / "config.ini"
    if not config_path.exists():
        print("   AVISO: config.ini nao encontrado. Usando valores padrao do script.")
        # Garante caminho absoluto das planilhas mesmo sem config
        PLANILHA_ASSOCIADOS = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_ASSOCIADOS)))
        PLANILHA_DADOS      = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_DADOS)))
        return


    cfg = configparser.ConfigParser()
    try:
        with open(str(config_path), "r", encoding="utf-8") as f:
            content = f.read()
        # Remove escapes do Markdown que o Pippin adiciona ao exportar
        # \[ vira [   e   \_ vira _
        content = content.replace("\\[", "[").replace("\\_", "_")
        cfg.read_string(content)
    except Exception as e:
        print("   AVISO: Erro ao ler config.ini: " + str(e))
        PLANILHA_ASSOCIADOS = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_ASSOCIADOS)))
        PLANILHA_DADOS      = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_DADOS)))
        return
    print("   config.ini encontrado: " + str(config_path))


    # [GERAL]
    WAREHOUSE_ID      = cfg.get    ("GERAL",              "warehouse_id",           fallback=WAREHOUSE_ID)
    DOWNLOAD_TIMEOUT  = cfg.getint ("GERAL",              "download_timeout",        fallback=DOWNLOAD_TIMEOUT)
    MIN_OPPORTUNITIES = cfg.getint ("GERAL",              "min_opportunities",       fallback=MIN_OPPORTUNITIES)
    URL_ROSTER        = _resolver_url_local(cfg.get("GERAL", "url_roster", fallback=URL_ROSTER))
    URL_ASSOCIADOS    = _resolver_url_local(cfg.get("GERAL", "url_associados",  fallback=URL_ASSOCIADOS))
    URL_DADOS         = _resolver_url_local(cfg.get("GERAL", "url_dados",        fallback=URL_DADOS))
    URL_ATLAS         = _resolver_url_local(cfg.get("GERAL", "url_atlas",        fallback=URL_ATLAS))


    # [ARQUIVOS]
    PLANILHA_ASSOCIADOS     = cfg.get("ARQUIVOS", "planilha_associados",     fallback=PLANILHA_ASSOCIADOS)
    PLANILHA_DADOS          = cfg.get("ARQUIVOS", "planilha_dados",          fallback=PLANILHA_DADOS)
    PLANILHA_ATLAS          = cfg.get("ARQUIVOS", "planilha_atlas",          fallback=PLANILHA_ATLAS)
    PLANILHA_SAIDA_COMPLETO = cfg.get("ARQUIVOS", "planilha_saida_completo", fallback=PLANILHA_SAIDA_COMPLETO)
    PLANILHA_SAIDA_RC_CRET  = cfg.get("ARQUIVOS", "planilha_saida_rc_cret",  fallback=PLANILHA_SAIDA_RC_CRET)
    PLANILHA_SAIDA_FINAL    = cfg.get("ARQUIVOS", "planilha_saida_final",    fallback=PLANILHA_SAIDA_FINAL)


    # [QUALIDADE]
    QUALITY_TOP_MAX    = cfg.getfloat("QUALIDADE",     "quality_top_max",    fallback=QUALITY_TOP_MAX)
    QUALITY_BOTTOM_MIN = cfg.getfloat("QUALIDADE",     "quality_bottom_min", fallback=QUALITY_BOTTOM_MIN)


    # [PRODUTIVIDADE]
    PROD_TOP_MINIMO    = cfg.getfloat("PRODUTIVIDADE", "prod_top_minimo",    fallback=PROD_TOP_MINIMO)
    PROD_BOTTOM_MAXIMO = cfg.getfloat("PRODUTIVIDADE", "prod_bottom_maximo", fallback=PROD_BOTTOM_MAXIMO)


    # [REGRAS_NEGOCIO]
    MIN_ASSOC_POR_EMBAIXADOR = cfg.getint("REGRAS_NEGOCIO", "min_assoc_por_embaixador", fallback=MIN_ASSOC_POR_EMBAIXADOR)
    MAX_ASSOC_POR_EMBAIXADOR = cfg.getint("REGRAS_NEGOCIO", "max_assoc_por_embaixador", fallback=MAX_ASSOC_POR_EMBAIXADOR)
    HORAS_MINIMAS            = cfg.getint("REGRAS_NEGOCIO", "horas_minimas",            fallback=HORAS_MINIMAS)


    # [INDICES_ASSOCIADOS]
    IDX_ASSOCIADO   = cfg.getint("INDICES_ASSOCIADOS",  "idx_associado",   fallback=IDX_ASSOCIADO)
    IDX_EMBAIXADOR  = cfg.getint("INDICES_ASSOCIADOS",  "idx_embaixador",  fallback=IDX_EMBAIXADOR)
    IDX_TURNO_ASSOC = cfg.getint("INDICES_ASSOCIADOS",  "idx_turno_assoc", fallback=IDX_TURNO_ASSOC)


    # [INDICES_FUNCIONARIOS]
    IDX_LOGIN_FUNC     = cfg.getint("INDICES_FUNCIONARIOS", "idx_login_func",     fallback=IDX_LOGIN_FUNC)
    IDX_NOME_FUNC      = cfg.getint("INDICES_FUNCIONARIOS", "idx_nome_func",      fallback=IDX_NOME_FUNC)
    IDX_TURNO_FUNC     = cfg.getint("INDICES_FUNCIONARIOS", "idx_turno_func",     fallback=IDX_TURNO_FUNC)
    IDX_PROCESSO_FUNC  = cfg.getint("INDICES_FUNCIONARIOS", "idx_processo_func",  fallback=IDX_PROCESSO_FUNC)
    IDX_LC_LEVEL_FUNC  = cfg.getint("INDICES_FUNCIONARIOS", "idx_lc_level_func",  fallback=IDX_LC_LEVEL_FUNC)
    IDX_UPH_PPR_FUNC   = cfg.getint("INDICES_FUNCIONARIOS", "idx_uph_ppr_func",   fallback=IDX_UPH_PPR_FUNC)
    IDX_HORAS_FUNC     = cfg.getint("INDICES_FUNCIONARIOS", "idx_horas_func",     fallback=IDX_HORAS_FUNC)
    IDX_RATE_REAL_FUNC  = cfg.getint("INDICES_FUNCIONARIOS", "idx_rate_real_func", fallback=IDX_RATE_REAL_FUNC)
    IDX_RATE_LC_FUNC    = cfg.getint("INDICES_FUNCIONARIOS", "idx_rate_lc_func",   fallback=IDX_RATE_LC_FUNC)
    IDX_TARGET_PPR_FUNC = cfg.getint("INDICES_FUNCIONARIOS", "idx_target_ppr_func",fallback=IDX_TARGET_PPR_FUNC)


    # Resolve as planilhas de ENTRADA para caminho absoluto na RAIZ do projeto
    # (evita depender da "pasta atual" de execucao).
    if not os.path.isabs(str(PLANILHA_ASSOCIADOS)):
        PLANILHA_ASSOCIADOS = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_ASSOCIADOS)))
    if not os.path.isabs(str(PLANILHA_DADOS)):
        PLANILHA_DADOS = str(SCRIPT_DIR / os.path.basename(str(PLANILHA_DADOS)))


    print("   config.ini carregado com sucesso!")


# ============================================
# DETECTAR PERFIL PADRÃO DO FIREFOX
# Lê profiles.ini e retorna o caminho do
# perfil padrão (onde o Midway já está ativo)
# Windows: %APPDATA%\Mozilla\Firefox\profiles.ini
# ============================================
def obter_perfil_firefox():
    candidates = [
        os.path.join(os.environ.get("APPDATA", ""),
                     "Mozilla", "Firefox", "profiles.ini"),
        os.path.expanduser(
            "~/Library/Application Support/Firefox/profiles.ini"),
        os.path.expanduser("~/.mozilla/firefox/profiles.ini"),
    ]
    for profiles_ini in candidates:
        if not os.path.exists(profiles_ini):
            continue
        cfg = configparser.ConfigParser()
        try:
            cfg.read(profiles_ini, encoding="utf-8")
        except Exception:
            continue
        base_dir    = os.path.dirname(profiles_ini)
        perfil_path = None


        # Método 1: seção Install (Firefox moderno — mais confiável)
        for section in cfg.sections():
            if section.lower().startswith("install"):
                default = cfg.get(section, "Default", fallback=None)
                if default:
                    perfil_path = os.path.join(
                        base_dir, default.replace("/", os.sep))
                    break


        # Método 2: Profile com Default=1
        if not perfil_path:
            for section in cfg.sections():
                if section.lower().startswith("profile"):
                    if cfg.get(section, "Default", fallback="") == "1":
                        is_rel = cfg.getint(section, "IsRelative", fallback=1)
                        path   = cfg.get(section, "Path", fallback=None)
                        if path:
                            perfil_path = (
                                os.path.join(base_dir, path.replace("/", os.sep))
                                if is_rel else path
                            )
                            break


        if perfil_path and os.path.isdir(perfil_path):
            return perfil_path


    return None


# ============================================
# DETECTAR CAMINHO DO FIREFOX
# ============================================
def find_firefox_binary():
    candidates = [
        r"C:\Program Files\Mozilla Firefox\firefox.exe",
        r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
        "/Applications/Firefox.app/Contents/MacOS/firefox",
        "/usr/bin/firefox",
        "/usr/lib/firefox/firefox",
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


# ============================================
# CACHE DO GECKODRIVER
# GeckoDriverManager().install() faz checagem de rede toda
# vez (lento). Cacheia o caminho na 1a chamada e reusa.
# ============================================
_GECKO_PATH_CACHE = None

def obter_gecko_path():
    global _GECKO_PATH_CACHE
    if _GECKO_PATH_CACHE and os.path.isfile(_GECKO_PATH_CACHE):
        return _GECKO_PATH_CACHE
    from webdriver_manager.firefox import GeckoDriverManager
    _GECKO_PATH_CACHE = GeckoDriverManager().install()
    return _GECKO_PATH_CACHE


# ============================================
# FECHAR FIREFOX (libera o perfil para o -profile)
# ============================================
def kill_firefox():
    try:
        if os.name == "nt":
            subprocess.run(["taskkill", "/F", "/IM", "firefox.exe"],
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        else:
            subprocess.run(["pkill", "-f", "firefox"],
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.2)
    except Exception:
        pass


# ============================================
# SETUP FIREFOX
# ============================================
def setup_firefox(usar_perfil=False):
    print()
    print("    Iniciando navegador Firefox...")
    from selenium import webdriver
    from selenium.webdriver.firefox.service import Service as FirefoxService
    from selenium.webdriver.firefox.options import Options as FirefoxOptions
    # Com perfil no lugar (-profile), o Firefox precisa estar fechado
    if usar_perfil:
        kill_firefox()
    options = FirefoxOptions()
    firefox_bin = find_firefox_binary()
    if firefox_bin:
        options.binary_location = firefox_bin
        print("    Firefox encontrado: " + firefox_bin)
    else:
        print("    Firefox: usando deteccao automatica do sistema")


    # Carrega perfil autenticado (Midway ativo) quando solicitado
    if usar_perfil:
        perfil_path = obter_perfil_firefox()
        if perfil_path:
            # Usa o perfil NO LUGAR (-profile) — rapido, sem copiar
            options.add_argument("-profile")
            options.add_argument(perfil_path)
            print("    Perfil Firefox: " + perfil_path)
        else:
            print("    AVISO: Perfil Firefox nao encontrado. Abrindo sem perfil.")
            print("    Abra o Firefox, faca login na Amazon e tente novamente.")


    options.set_preference("browser.download.folderList", 2)
    options.set_preference("browser.download.dir", DOWNLOAD_DIR)
    options.set_preference("browser.download.useDownloadDir", True)
    options.set_preference("browser.download.manager.showWhenStarting", False)
    options.set_preference(
        "browser.helperApps.neverAsk.saveToDisk",
        "text/csv,application/csv,application/octet-stream,"
        "text/plain,application/vnd.ms-excel,"
        "application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet,text/tab-separated-values,"
        "application/json,application/zip"
    )
    options.set_preference("browser.download.manager.showAlertOnComplete", False)
    options.set_preference("browser.download.panel.shown", False)
    spinner = iniciar_spinner("Preparando driver do Firefox")
    gecko_path = obter_gecko_path()
    service    = FirefoxService(executable_path=gecko_path)
    try:
        driver = webdriver.Firefox(service=service, options=options)
    except Exception as e:
        parar_spinner(spinner)
        print("ERRO: Falha ao abrir o Firefox. Detalhe: " + str(e))
        raise
    parar_spinner(spinner)
    driver.set_page_load_timeout(60)
    print("    Firefox pronto! Downloads: " + DOWNLOAD_DIR)
    return driver


# ============================================
# GERAR URL
# ============================================
def generate_url(start_date, end_date, target_process="combined"):
    # DEMO / PORTFOLIO: se o ATLAS aponta para um site local (file://),
    # abre o HTML falso direto, sem anexar parametros de query.
    if str(URL_ATLAS).lower().startswith("file:"):
        return URL_ATLAS
    return (
        URL_ATLAS
        + "?aggregateType=WAREHOUSE_ID"
        + "&queryType=NORMAL"
        + "&targetProcess=" + target_process
        + "&warehouseId=" + WAREHOUSE_ID
        + "&startDate=" + start_date
        + "&startTime=00%3A00%3A00"
        + "&endDate=" + end_date
        + "&endTime=23%3A45%3A00"
    )


# ============================================
# LOGIN
# ============================================
def esperar_pagina(driver, timeout=15):
    """Espera a pagina terminar de carregar (readyState=complete).
    Retorna assim que pronta — muito mais rapido que time.sleep fixo."""
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            if driver.execute_script("return document.readyState") == "complete":
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


def handle_login(driver, url):
    esperar_pagina(driver, timeout=8)
    current_url = driver.current_url.lower()
    keywords    = ["midway", "signin", "login", "sso", "auth"]
    if any(k in current_url for k in keywords):
        print()
        print("LOGIN NECESSARIO")
        print("Faca login com sua YubiKey no navegador e depois volte aqui.")
        input(C.ORANGE + C.BOLD + "  Pressione ENTER apos fazer login... " + C.RESET)
        driver.get(url)
        esperar_pagina(driver, timeout=15)


# ============================================
# SNAPSHOT / AGUARDAR ARQUIVO
# ============================================
def snapshot_downloads():
    files = {}
    for f in os.listdir(DOWNLOAD_DIR):
        full_path = os.path.join(DOWNLOAD_DIR, f)
        if os.path.isfile(full_path):
            files[f] = os.path.getmtime(full_path)
    return files


def wait_new_file(before_snapshot, timeout=DOWNLOAD_TIMEOUT):
    start_time = time.time()
    temp_exts  = ('.part', '.crdownload', '.tmp')
    while time.time() - start_time < timeout:
        time.sleep(0.25)
        try:
            current_files = os.listdir(DOWNLOAD_DIR)
        except Exception:
            continue
        if any(f.endswith(temp_exts) for f in current_files):
            continue
        for filename in current_files:
            if filename.endswith(temp_exts):
                continue
            full_path = os.path.join(DOWNLOAD_DIR, filename)
            try:
                if not os.path.isfile(full_path):
                    continue
                is_new      = filename not in before_snapshot
                is_modified = os.path.getmtime(full_path) > start_time
                if is_new or is_modified:
                    size1 = os.path.getsize(full_path)
                    if size1 > 0:
                        time.sleep(0.25)
                        size2 = os.path.getsize(full_path)
                        if size1 == size2:
                            return full_path
            except Exception:
                continue
    return None


# ============================================
# BARRA DE PROGRESSO
# ============================================
def progress_bar(current, total, largura=24):
    if total <= 0:
        total = 1
    ratio   = max(0.0, min(1.0, current / total))
    filled  = int(largura * ratio)
    percent = int(100 * ratio)
    # Cor por progresso: vermelho < 40% < dourado < 80% < verde
    if percent < 40:
        cor = C.RED
    elif percent < 80:
        cor = C.GOLD
    else:
        cor = C.GREEN
    barra = (cor + "\u2588" * filled + C.RESET
             + C.GRAY + "\u2591" * (largura - filled) + C.RESET)
    return barra + "  " + cor + C.BOLD + str(percent).rjust(3) + "%" + C.RESET


def barra_progresso_linha(current, total, rotulo=""):
    """Imprime uma barra de progresso que se atualiza na mesma linha."""
    bar = progress_bar(current, total)
    rot = (C.SKY + rotulo + C.RESET + "  ") if rotulo else ""
    print("\r  " + rot + bar + "   ", end="", flush=True)
    if current >= total:
        print()


# ============================================
# BAIXAR TOTALS REPORT
# ============================================
def baixar_totals_report(driver, start_date, end_date, is_first_open):
    url = generate_url(start_date, end_date, target_process="combined")
    print()
    print("=" * 55)
    print("TOTALS REPORT")
    print("=" * 55)
    driver.get(url)
    if is_first_open:
        handle_login(driver, url)
    esperar_pagina(driver, timeout=12)
    print("No navegador:")
    print("  O relatorio ja foi aberto automaticamente.")
    print("  1. Clique em 'Download' para baixar o CSV")
    before     = snapshot_downloads()
    perguntar("Pressione ENTER apos iniciar o download do Totals Report... ")
    print("Aguardando download...", end="", flush=True)
    downloaded = wait_new_file(before, timeout=DOWNLOAD_TIMEOUT)
    if downloaded:
        print()
        print("Totals Report detectado: " + os.path.basename(downloaded))
        return downloaded
    else:
        print()
        caminho_manual = perguntar(
            "Cole o caminho completo do Totals Report (ou ENTER para pular): "
        ).strip().strip('"')
        if caminho_manual and os.path.isfile(caminho_manual):
            return caminho_manual
        print("Totals Report ignorado. Quality Rate nao sera calculado.")
        return None


# ============================================
# LER TOTALS REPORT
# ============================================
def ler_totals_report(caminho):
    print()
    print("--------------------------------------------------")
    print("Lendo Totals Report: " + os.path.basename(caminho))
    print("--------------------------------------------------")
    totais = {}
    try:
        df = None
        for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            try:
                df = pd.read_csv(caminho, dtype=str, encoding=enc)
                break
            except Exception:
                continue
        if df is None:
            print("  AVISO: Nao foi possivel ler o Totals Report.")
            return totais
        df.columns = [str(c).strip().lstrip('\ufeff') for c in df.columns]
        if len(df.columns) < 5:
            print("  AVISO: CSV com menos de 5 colunas.")
            return totais
        col_a = df.columns[0]
        col_e = df.columns[4]
        print("  Col A (defectType): '" + col_a + "' | Col E (metricValue): '" + col_e + "'")
        print()
        for _, row in df.iterrows():
            tipo = str(row[col_a]).strip()
            if not tipo.startswith(TOTAL_PREFIX):
                continue
            processo = tipo[len(TOTAL_PREFIX):].strip().upper()
            if "PACK" in processo:
                continue   # PACK tratado separadamente por subprocess
            try:
                valor = float(str(row[col_e]).replace(",", ".").strip())
            except (ValueError, TypeError):
                continue
            totais[processo] = valor
            print("  " + tipo.ljust(22) + " -> " + processo.ljust(12) + " = " + str(valor))
    except Exception as e:
        print("  ERRO: " + str(e))
    print()
    print("  Processos extraidos: " + str(len(totais)))
    print("--------------------------------------------------")
    return totais


# ============================================
# BAIXAR PACK SUBPROCESSOS (Single e Multi)
# Chamado apos o Totals Report geral
# ============================================
def baixar_pack_subprocessos(driver, start_date, end_date):
    subprocessos = [
        ("PACK_SINGLE", "Single", "  1. Mude Target Process para PACK\n  2. Mude Subprocess para Single\n  3. Clique em Totals Report e baixe"),
        ("PACK_MULTI",  "Multi",  "  1. Mude Subprocess para Multi\n  2. Clique em Totals Report e baixe"),
    ]
    total     = len(subprocessos)
    resultados = {}

    print()
    print("=" * 55)
    print("PACK SUBPROCESSOS - TOTALS REPORT")
    print("=" * 55)


    for i, (chave, nome, instrucoes) in enumerate(subprocessos):
        idx = i + 1
        print()
        print("  " + C.ORANGE + C.BOLD + "PACK " + nome + C.RESET
              + C.GRAY + "  (" + str(idx) + "/" + str(total) + ")" + C.RESET)
        print(C.GRAY + instrucoes + C.RESET)


        before     = snapshot_downloads()
        perguntar("ENTER apos baixar PACK " + nome + "... ")
        print("Detectando...", end="", flush=True)
        downloaded = wait_new_file(before, timeout=10)
        if not downloaded:
            downloaded = wait_new_file(before, timeout=5)


        if downloaded:
            resultados[chave] = downloaded
            print(" OK -> " + os.path.basename(downloaded))
        else:
            print()
            retry = perguntar("Nenhum arquivo. Tentar novamente? (ENTER = sim / N = nao): ").strip().lower()
            if retry != "n":
                before2    = snapshot_downloads()
                perguntar("ENTER apos baixar PACK " + nome + "... ")
                print("Detectando...", end="", flush=True)
                downloaded2 = wait_new_file(before2, timeout=15)
                if downloaded2:
                    resultados[chave] = downloaded2
                    print(" OK -> " + os.path.basename(downloaded2))
                else:
                    print()
                    print("Pulando PACK " + nome + ". Quality Rate do PACK " + nome + " nao sera calculado.")
            else:
                print("Pulando PACK " + nome + ". Quality Rate do PACK " + nome + " nao sera calculado.")

        # Barra de conclusao (mostra a % concluida, incluindo 100% no ultimo)
        print("  " + progress_bar(idx, total)
              + C.GREEN + "  concluido" + C.RESET)

    print()
    encontrados = [k for k in resultados]
    print("RESUMO Pack Subprocessos: " + (", ".join(encontrados) if encontrados else "nenhum"))
    return resultados


# ============================================
# LER PACK TOTALS (Single e Multi)
# Extrai o valor total de cada CSV de subprocess
# ============================================
def ler_pack_totals(caminhos_pack):
    totais = {}
    for chave, caminho in caminhos_pack.items():
        print()
        print("  Lendo " + chave + ": " + os.path.basename(caminho))
        try:
            df = None
            for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(caminho, dtype=str, encoding=enc)
                    break
                except Exception:
                    continue
            if df is None:
                print("  AVISO: Nao foi possivel ler " + chave)
                continue


            df.columns = [str(c).strip().lstrip('\ufeff') for c in df.columns]
            if len(df.columns) < 5:
                print("  AVISO: CSV com menos de 5 colunas em " + chave)
                continue


            col_a = df.columns[0]
            col_e = df.columns[4]
            valor_total = None


            for _, row in df.iterrows():
                tipo = str(row[col_a]).strip()
                if not tipo.startswith(TOTAL_PREFIX):
                    continue
                try:
                    v = float(str(row[col_e]).replace(",", ".").strip())
                    valor_total = v
                    break
                except (ValueError, TypeError):
                    continue


            if valor_total is not None:
                totais[chave] = valor_total
                print("  " + chave + " = " + str(valor_total))
            else:
                print("  AVISO: Nenhuma linha 'Total' encontrada em " + chave)


        except Exception as e:
            print("  ERRO lendo " + chave + ": " + str(e))

    return totais


# ============================================
# ABRIR A ABA "RAW REPORTS" (MANUAL)
# O usuario clica na aba "Raw Reports" no navegador e da ENTER.
# ============================================
def clicar_raw_reports(driver, tentativas=3):
    print()
    print(C.GRAY + "    No navegador, clique na aba [ Raw Reports ]." + C.RESET)
    perguntar("    Pressione ENTER apos clicar em 'Raw Reports'... ")
    return True


# ============================================
# PROCESSAR GRUPO (raw reports)
# ============================================
def process_group(driver, processes, group_name, start_date, end_date,
                  target_process, is_first_group, counter_start, total):
    success_list     = []
    failed_list      = []
    downloaded_paths = []
    url = generate_url(start_date, end_date, target_process)
    print()
    print("--------------------------------------------------")
    print("GRUPO: " + group_name + " | Processos: " + ", ".join(processes))
    print("--------------------------------------------------")
    driver.get(url)
    if is_first_group:
        handle_login(driver, url)
    esperar_pagina(driver, timeout=12)
    clicar_raw_reports(driver)


    for i, name in enumerate(processes):
        idx = counter_start + i
        print()
        print("  " + C.ORANGE + C.BOLD + name + C.RESET
              + C.GRAY + "  (" + str(idx) + "/" + str(total) + ")" + C.RESET)
        print(C.GRAY + "  1. Clique no botao [ " + name + " ]   2. Clique em download" + C.RESET)
        before     = snapshot_downloads()
        perguntar("ENTER apos baixar " + name + "... ")
        print("Detectando...", end="", flush=True)
        downloaded = wait_new_file(before, timeout=10)
        if not downloaded:
            downloaded = wait_new_file(before, timeout=5)
        if downloaded:
            downloaded_paths.append((name.upper(), downloaded))
            success_list.append(name)
            print(" OK -> " + os.path.basename(downloaded))
        else:
            print()
            retry = perguntar("Nenhum arquivo. Tentar novamente? (ENTER = sim / N = nao): ").strip().lower()
            if retry != "n":
                before2     = snapshot_downloads()
                perguntar("ENTER apos baixar... ")
                downloaded2 = wait_new_file(before2, timeout=15)
                if downloaded2:
                    downloaded_paths.append((name.upper(), downloaded2))
                    success_list.append(name)
                    print("OK -> " + os.path.basename(downloaded2))
                else:
                    print("Pulando " + name + ".")
                    failed_list.append(name)
            else:
                failed_list.append(name)

        # Barra de conclusao (mostra a % concluida, incluindo 100% no ultimo)
        print("  " + progress_bar(idx, total)
              + C.GREEN + "  concluido" + C.RESET)

    print()
    print("RESUMO " + group_name + ": OK=" + ", ".join(success_list)
          + (" | FALHA=" + ", ".join(failed_list) if failed_list else ""))
    return downloaded_paths


# ============================================
# CALCULAR DPMO POR PROCESSO
# PACK : real = totalDefects - soma(todas cols "wrong box")
# PICK : real = totalDefects - Short - Damage
# STOW : real = totalDefects - bin collision
# Outros: real = totalDefects (sem subtracao)
# DPMO = (real_defects / opportunities) * 1.000.000
# ============================================
def calcular_dpmo_por_processo(df, processo_nome, total_defects, opps):
    processo = str(processo_nome).strip().upper()


    def to_num(series):
        return pd.to_numeric(series, errors='coerce').fillna(0.0)


    def col_exata(nome_lower):
        cols = [c for c in df.columns if str(c).strip().lower() == nome_lower]
        return cols[0] if cols else None


    def cols_contendo(fragmento):
        return [c for c in df.columns if fragmento in str(c).strip().lower()]


    if processo == "PACK":
        # Subtrai Wrong box (todas as variações)
        wrong_box_cols = cols_contendo("wrong box")
        if wrong_box_cols:
            wrong_box_sum = df[wrong_box_cols].apply(to_num).sum(axis=1)
            print("      [PACK] Subtraindo " + str(len(wrong_box_cols))
                  + " coluna(s) Wrong box: " + str(wrong_box_cols))
        else:
            wrong_box_sum = pd.Series(0.0, index=df.index)
            print("      [PACK] Colunas 'Wrong box' nao encontradas. Assumindo 0.")
        # Subtrai Missing (todas as variações)
        missing_cols = cols_contendo("missing")
        if missing_cols:
            missing_sum = df[missing_cols].apply(to_num).sum(axis=1)
            print("      [PACK] Subtraindo " + str(len(missing_cols))
                  + " coluna(s) Missing: " + str(missing_cols))
        else:
            missing_sum = pd.Series(0.0, index=df.index)
            print("      [PACK] Colunas 'Missing' nao encontradas. Assumindo 0.")
        real_defects = (total_defects - wrong_box_sum - missing_sum).clip(lower=0)


    elif processo == "PICK":
        short_col  = col_exata("short")
        damage_col = col_exata("damage")
        short_val  = to_num(df[short_col])  if short_col  else pd.Series(0.0, index=df.index)
        damage_val = to_num(df[damage_col]) if damage_col else pd.Series(0.0, index=df.index)
        if not short_col:
            print("      [PICK] Coluna 'Short' nao encontrada. Assumindo 0.")
        if not damage_col:
            print("      [PICK] Coluna 'Damage' nao encontrada. Assumindo 0.")
        if short_col or damage_col:
            achou = []
            if short_col:  achou.append("Short")
            if damage_col: achou.append("Damage")
            print("      [PICK] Subtraindo: " + " + ".join(achou))
        real_defects = (total_defects - short_val - damage_val).clip(lower=0)


    elif processo == "STOW":
        bin_cols = cols_contendo("bin collision")
        bin_val  = to_num(df[bin_cols[0]]) if bin_cols else pd.Series(0.0, index=df.index)
        if not bin_cols:
            print("      [STOW] Coluna 'bin collision' nao encontrada. Assumindo 0.")
        else:
            print("      [STOW] Subtraindo bin collision.")
        # Subtrai sips over and short
        sips_cols = cols_contendo("sips over and short")
        sips_val  = to_num(df[sips_cols[0]]) if sips_cols else pd.Series(0.0, index=df.index)
        if not sips_cols:
            print("      [STOW] Coluna 'sips over and short' nao encontrada. Assumindo 0.")
        else:
            print("      [STOW] Subtraindo sips over and short.")
        real_defects = (total_defects - bin_val - sips_val).clip(lower=0)


    else:
        # Sort, Ship, Receive, RC SORT, C-RET, etc.
        real_defects = total_defects.copy()


    # DPMO = (real_defects / opps) * 1.000.000
    dpmo       = pd.Series(0.0, index=df.index)
    mask       = opps > 0
    dpmo[mask] = ((real_defects[mask] / opps[mask]) * 1000000).round(2)


    return real_defects.reset_index(drop=True), dpmo.reset_index(drop=True)


# ============================================
# CONSOLIDAR -> ExtracaoATLAS.xlsx
# ============================================
def consolidate_to_excel(downloaded_files):
    print()
    print("--------------------------------------------------")
    print("CONSOLIDANDO PLANILHAS")
    print("--------------------------------------------------")
    if not downloaded_files:
        print("Nenhum arquivo para consolidar.")
        return None
    OUTPUT_COLS = ["User", "Subprocess", "Total Defects", "Opportunities", "DPMO"]
    all_rows    = []
    skipped     = []
    for process_name, path in downloaded_files:
        filename = os.path.basename(path)
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in (".xlsx", ".xls"):
                df = pd.read_excel(path, dtype=str)
            else:
                df = pd.read_csv(path, dtype=str, sep=None, engine="python")


            # Normalizar nomes de colunas (preserva original para busca de defeitos)
            df.columns = [str(c).strip() for c in df.columns]


            # Coluna User (primeira coluna do CSV)
            user_col = df.iloc[:, 0].rename("User").reset_index(drop=True)


            # Coluna Subprocess
            sub_match      = [c for c in df.columns if str(c).lower() == "subprocess"]
            subprocess_col = (df[sub_match[0]].reset_index(drop=True)
                              if sub_match else pd.Series("", index=range(len(df))))


            # totalDefects e Opportunities como numerico
            td_match = [c for c in df.columns if str(c).lower() == "totaldefects"]
            op_match = [c for c in df.columns if str(c).lower() == "opportunities"]


            total_defects_raw = (pd.to_numeric(df[td_match[0]], errors='coerce').fillna(0.0)
                                 if td_match else pd.Series(0.0, index=df.index))
            opps_raw          = (pd.to_numeric(df[op_match[0]], errors='coerce').fillna(0.0)
                                 if op_match else pd.Series(0.0, index=df.index))


            # Calcular real_defects e DPMO baseado nas regras do processo
            print("  " + process_name + ": calculando DPMO real...")
            real_defects, dpmo_calc = calcular_dpmo_por_processo(
                df, process_name, total_defects_raw, opps_raw)


            # Montar DataFrame de saida (Total Defects = real defects calculados)
            df_filtered = pd.DataFrame({
                "User":          user_col,
                "Subprocess":    subprocess_col,
                "Total Defects": real_defects,
                "Opportunities": opps_raw.reset_index(drop=True),
                "DPMO":          dpmo_calc,
            })


            label_row  = pd.DataFrame([[process_name, "", "", "", ""]], columns=OUTPUT_COLS)
            header_row = pd.DataFrame([OUTPUT_COLS],                    columns=OUTPUT_COLS)
            empty_row  = pd.DataFrame([["", "", "", "", ""]],           columns=OUTPUT_COLS)
            all_rows.extend([label_row, header_row, df_filtered, empty_row])
            print("  " + process_name + ": " + filename + " (" + str(len(df_filtered)) + " linhas)")
        except Exception as e:
            print("  AVISO: nao foi possivel ler '" + filename + "': " + str(e))
            skipped.append(filename)
    if not all_rows:
        print("Nenhum arquivo valido para consolidar.")
        return None
    final_df    = pd.concat(all_rows, ignore_index=True)
    output_path = str(RECURSOS_DIR / PLANILHA_ATLAS)
    try:
        final_df.to_excel(output_path, index=False, header=False, engine="openpyxl")
        print("Consolidado: " + output_path)
        return output_path
    except Exception as e:
        print("ERRO ao salvar: " + str(e))
        return None


# ============================================
# EXTRAIR DATA DO NOME DA ABA
# Ex: "BD 21.04" -> dia=21, mes=04
# ============================================
def extrair_data_de_aba(nome_aba):
    match = re.search(r'(\d{1,2})\.(\d{2})', str(nome_aba))
    if match:
        try:
            dia = int(match.group(1))
            mes = int(match.group(2))
            return dia, mes
        except ValueError:
            pass
    return None, None


# ============================================
# EXTRAIR DATA PARA NOME DO BACKUP
# Usa a data mais antiga das abas de associados.xlsx
# Ex: abas ["BD 21.04", "BD 28.04"] -> "21.04"
# Fallback: data do sistema se nenhuma aba tiver data
# ============================================
def extrair_data_backup(abas_nomes):
    ano_atual = datetime.now().year
    datas     = []
    for nome in (abas_nomes or []):
        dia, mes = extrair_data_de_aba(nome)
        if dia and mes:
            try:
                datas.append(datetime(ano_atual, mes, dia))
            except ValueError:
                pass
    if datas:
        return min(datas).strftime("%d.%m")   # data mais antiga das abas
    return datetime.now().strftime("%d.%m")   # fallback: hoje


# ============================================
# CONSTRUIR STRING DE DATAS (para col A do resultado)
# Ex: abas ["BD 21.04","BD 28.04"] -> "21.04 | 28.04"
# ============================================
def construir_data_str(abas_nomes):
    datas = []
    for nome in (abas_nomes or []):
        dia, mes = extrair_data_de_aba(nome)
        if dia and mes:
            d = str(dia).zfill(2) + "." + str(mes).zfill(2)
            if d not in datas:
                datas.append(d)
    if datas:
        return " | ".join(sorted(datas))
    return datetime.now().strftime("%d.%m")   # fallback: hoje


# ============================================
# CALCULAR DATAS DO ATLAS AUTOMATICAMENTE
# Start = data mais antiga das abas + 2 dias
# End   = Start + 13 dias
# ============================================
def calcular_datas_atlas(abas_nomes):
    ano_atual        = datetime.now().year
    datas_encontradas = []

    print()
    print("=" * 55)
    print("DATAS CALCULADAS AUTOMATICAMENTE")
    print("=" * 55)
    print("   Abas reconhecidas:")


    for nome in abas_nomes:
        dia, mes = extrair_data_de_aba(nome)
        if dia and mes:
            try:
                dt = datetime(ano_atual, mes, dia)
                datas_encontradas.append((dt, nome))
                print("      [" + str(nome) + "] -> " + dt.strftime("%d/%m/%Y"))
            except ValueError:
                print("      [" + str(nome) + "] -> data invalida, ignorada")
        else:
            print("      [" + str(nome) + "] -> sem data reconhecida")


    if not datas_encontradas:
        print()
        print("   Nenhuma data reconhecida nas abas.")
        print("   Alternando para entrada manual...")
        return ask_dates()


    data_base = min(dt for dt, _ in datas_encontradas)
    start_dt  = data_base + timedelta(days=2)
    end_dt    = start_dt  + timedelta(days=13)


    start_str = start_dt.strftime("%Y-%m-%d")
    end_str   = end_dt.strftime("%Y-%m-%d")

    print()
    print("   Data base (mais antiga) : " + data_base.strftime("%d/%m/%Y"))
    print("   Start Date              : " + start_dt.strftime("%d/%m/%Y")
          + "  (data base + 2 dias)")
    print("   End Date                : " + end_dt.strftime("%d/%m/%Y")
          + "  (Start + 13 dias)")
    print()
    confirmar = perguntar("Confirmar estas datas? (ENTER para aceitar / N para digitar manualmente): ").strip().lower()


    if confirmar == "n":
        return ask_dates()


    print("   Datas confirmadas: " + start_str + " -> " + end_str)
    return start_str, end_str


# ============================================
# PEDIR DATAS
# ============================================
def ask_dates():
    print()
    print("--------------------------------------------------")
    print("CONFIGURACAO DE DATAS  (formato: YYYY-MM-DD)")
    print("--------------------------------------------------")
    while True:
        start = perguntar("Data inicial (YYYY-MM-DD): ").strip()
        if re.match(r'^\d{4}-\d{2}-\d{2}$', start):
            break
        print("Formato invalido. Use YYYY-MM-DD")
    while True:
        end = perguntar("Data final   (YYYY-MM-DD): ").strip()
        if re.match(r'^\d{4}-\d{2}-\d{2}$', end):
            break
        print("Formato invalido. Use YYYY-MM-DD")
    return start, end


# ============================================
# OBTER APENAS ABAS VISIVEIS DE UM EXCEL
# Abas ocultas (hidden / veryHidden) sao ignoradas.
# Garante que abas de controle interno nao sejam
# processadas acidentalmente.
# ============================================
def obter_abas_visiveis(caminho):
    try:
        from openpyxl import load_workbook
        wb       = load_workbook(caminho, data_only=True)
        visiveis = [ws.title for ws in wb.worksheets
                    if ws.sheet_state == "visible"]
        ocultas  = [ws.title for ws in wb.worksheets
                    if ws.sheet_state != "visible"]
        wb.close()
        if ocultas:
            print("   Ignorando abas ocultas (" + str(len(ocultas)) + ")")
        return visiveis
    except Exception as e:
        print("   AVISO: Nao foi possivel verificar abas ocultas: " + str(e))
        try:
            return pd.ExcelFile(caminho, engine="openpyxl").sheet_names
        except Exception:
            return []


# ============================================
# HELPERS - LEITURA DE PLANILHAS
# ============================================
def ler_excel_com_header(caminho, sheet_name=0):
    # Otimizado: le o arquivo UMA vez (header=None) e detecta a
    # melhor linha de cabecalho em memoria, evitando 15 leituras de disco.
    try:
        raw = pd.read_excel(caminho, engine="openpyxl",
                            sheet_name=sheet_name, header=None)
    except Exception:
        df = pd.read_excel(caminho, engine="openpyxl", sheet_name=sheet_name)
        return df, 0

    melhor_header = 0
    menor_unnamed = 999
    limite = min(15, len(raw))
    for header_row in range(0, limite):
        linha = raw.iloc[header_row]
        nao_nulos = linha.notna().sum()
        if nao_nulos <= 1:
            continue
        # "unnamed" = celulas vazias na linha de cabecalho candidata
        unnamed = int(linha.isna().sum())
        if unnamed < menor_unnamed:
            menor_unnamed = unnamed
            melhor_header = header_row
            if unnamed == 0:
                break

    # Constroi o DataFrame a partir dos dados ja carregados
    colunas = raw.iloc[melhor_header].tolist()
    dados   = raw.iloc[melhor_header + 1:].reset_index(drop=True)
    dados.columns = colunas
    dados.columns = [
        str(c).replace("\n", " ").replace("\r", " ").strip()
        for c in dados.columns
    ]
    return dados, melhor_header


def ler_todas_abas(caminho):
    _, ext = os.path.splitext(caminho)
    ext    = ext.lower()
    if ext == ".csv":
        for enc in ["utf-8", "latin-1", "cp1252", "iso-8859-1"]:
            for sep in [",", ";", "\t", "|"]:
                try:
                    df = pd.read_csv(caminho, sep=sep, encoding=enc, on_bad_lines="skip")
                    if len(df.columns) > 1:
                        df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
                        return [("CSV", df)]
                except Exception:
                    continue
        raise Exception("Nao foi possivel ler: " + caminho)
    xl         = pd.ExcelFile(caminho, engine="openpyxl")
    nomes_abas = obter_abas_visiveis(caminho)   # ignora abas ocultas
    resultado  = []
    for nome_aba in nomes_abas:
        try:
            df, header = ler_excel_com_header(caminho, nome_aba)
            if len(df) > 0:
                resultado.append((nome_aba, df))
        except Exception:
            pass
    if not resultado:
        raise Exception("Nenhuma aba com dados em: " + caminho)
    return resultado


# Modo detalhado: True mostra dump de colunas (debug). False = terminal limpo.
VERBOSE = False

def mostrar_colunas(df, limite=20):
    if not VERBOSE:
        return
    print("\n      Colunas (" + str(len(df.columns)) + "):")
    for i, col in enumerate(df.columns):
        if i >= limite:
            print("         ... e mais " + str(len(df.columns) - limite))
            break
        letra   = chr(65 + i) if i < 26 else "Col" + str(i)
        vals    = df.iloc[:, i].dropna()
        exemplo = "  ex: " + str(vals.iloc[0]) if len(vals) > 0 else ""
        print("         " + letra + " (idx " + str(i).rjust(2) + "): "
              + str(col).ljust(30) + exemplo)


# ============================================
# CARREGAR ASSOCIACOES
# ============================================
def carregar_associacoes(caminho):
    print("   Lendo associados de: " + caminho)
    abas  = ler_todas_abas(caminho)
    todos = []
    for nome_aba, df_full in abas:
        print()
        print("=" * 55)
        print("   Segue os resultados da data [" + nome_aba + "]")
        print("=" * 55)
        mostrar_colunas(df_full)
        max_idx = len(df_full.columns) - 1
        idx_a   = min(IDX_ASSOCIADO,   max_idx)
        idx_e   = min(IDX_EMBAIXADOR,  max_idx)
        idx_t   = min(IDX_TURNO_ASSOC, max_idx)
        print("      Associado='"  + str(df_full.columns[idx_a]) + "' | "
              + "Embaixador='" + str(df_full.columns[idx_e]) + "' | "
              + "Turno='"      + str(df_full.columns[idx_t]) + "'")
        df = pd.DataFrame({
            "Associado":  df_full.iloc[:, idx_a],
            "Embaixador": df_full.iloc[:, idx_e],
            "Turno":      df_full.iloc[:, idx_t],
        })
        df = df.dropna(subset=["Associado", "Embaixador"])
        df["Associado"]  = (df["Associado"].astype(str).str.strip()
                            .str.replace(r'\.0$', '', regex=True))
        df["Embaixador"] = df["Embaixador"].astype(str).str.strip()
        df["Turno"]      = df["Turno"].astype(str).str.strip().replace("nan", "")
        df = df[(df["Associado"] != "") & (df["Embaixador"] != "")]
        df = df[(df["Associado"].str.lower() != "nan") &
                (df["Embaixador"].str.lower() != "nan")]
        print("      " + str(len(df)) + " associacoes carregadas em [" + nome_aba + "]")
        todos.append(df)
    df_final = pd.concat(todos, ignore_index=True)
    antes    = len(df_final)
    df_final = df_final.drop_duplicates(subset=["Associado"], keep="first")
    if len(df_final) < antes:
        print("\n   Removidas " + str(antes - len(df_final)) + " duplicatas")
    print("\n   TOTAL: " + str(len(df_final)) + " associacoes carregadas")
    return df_final


# ============================================
# LER EXTRACAO ATLAS
# ============================================
def ler_extracao_atlas(caminho):
    print("\n   Lendo Atlas: " + caminho)
    df_raw          = pd.read_excel(caminho, header=None, dtype=str, engine="openpyxl")
    HEADER_KEYWORDS = {"user", "subprocess", "total defects", "opportunities", "dpmo"}
    registros       = []
    processo_atual  = ""
    for _, row in df_raw.iterrows():
        valores = []
        for v in row:
            if pd.notna(v) and str(v).strip() != "nan":
                valores.append(str(v).strip())
            else:
                valores.append("")
        if all(v == "" for v in valores):
            continue
        nao_vazios_lower = set(v.lower() for v in valores if v)
        if nao_vazios_lower & HEADER_KEYWORDS:
            continue
        nao_vazios = [v for v in valores if v]
        if len(nao_vazios) == 1:
            processo_atual = nao_vazios[0].upper()
            continue
        if processo_atual and len(valores) >= 5:
            registros.append({
                "User":          valores[0],
                "Processo":      processo_atual,
                "Subprocess":    valores[1],
                "Total Defects": valores[2],
                "Opportunities": valores[3],
                "DPMO":          valores[4],
            })
    if registros:
        df = pd.DataFrame(registros)
    else:
        df = pd.DataFrame(columns=["User", "Processo", "Subprocess",
                                    "Total Defects", "Opportunities", "DPMO"])
    if len(df) > 0:
        df["User_busca"] = df["User"].str.lower().str.strip()
    n_proc = df["Processo"].nunique() if len(df) > 0 else 0
    print("   " + str(len(df)) + " registros lidos (" + str(n_proc) + " processos)")
    return df


# ============================================
# HELPER: VERIFICAR RC/CRET
# ============================================
def eh_rc_cret(processo):
    p = str(processo).strip().upper().replace(" ", "").replace("-", "")
    for rc in PROCESSOS_RC_CRET:
        rc_norm = rc.upper().replace(" ", "").replace("-", "")
        if p == rc_norm or rc_norm in p:
            return True
    return False


# ============================================
# CRUZAR ATLAS COM ASSOCIADOS
# ============================================
def cruzar_atlas_associados(df_atlas, df_associacoes, totais_processo=None):
    print("\n   Cruzando " + str(len(df_associacoes)) + " associados com o Atlas...")
    todos_resultados = []
    nao_encontrados  = []
    encontrados      = set()
    tem_totais       = totais_processo is not None and len(totais_processo) > 0


    for _, row in df_associacoes.iterrows():
        associado  = row["Associado"]
        embaixador = row["Embaixador"]
        busca_val  = associado.lower().strip()
        busca      = df_atlas[df_atlas["User_busca"] == busca_val] if len(df_atlas) > 0 else pd.DataFrame()


        if not busca.empty:
            encontrados.add(associado)
            for _, dados in busca.iterrows():
                processo_upper = str(dados["Processo"]).strip().upper()
                quality_rate   = None
                if tem_totais:
                    if processo_upper == "PACK":
                        sub = str(dados["Subprocess"]).lower()
                        if "single" in sub:
                            total_proc = totais_processo.get("PACK_SINGLE", None)
                        elif "multi" in sub:
                            total_proc = totais_processo.get("PACK_MULTI", None)
                        else:
                            total_proc = None
                    else:
                        total_proc = totais_processo.get(processo_upper, None)
                    try:
                        dpmo_val = float(str(dados["DPMO"]).replace(",", ".").strip())
                        if total_proc and total_proc > 0:
                            quality_rate = round((dpmo_val / total_proc) * 100, 2)
                    except (ValueError, TypeError):
                        quality_rate = None
                todos_resultados.append({
                    "Embaixador":       embaixador,
                    "Associado":        associado,
                    "Processo":         dados["Processo"],
                    "Subprocess":       dados["Subprocess"],
                    "Total Defects":    dados["Total Defects"],
                    "Opportunities":    dados["Opportunities"],
                    "DPMO":             dados["DPMO"],
                    "Percentual Embaixador": quality_rate if quality_rate is not None else "",
                })
        else:
            nao_encontrados.append(associado)


    df_resultado = pd.DataFrame(todos_resultados)
    if len(df_resultado) > 0:
        df_resultado = df_resultado.sort_values(
            ["Embaixador", "Associado", "Processo"]
        ).reset_index(drop=True)


    print("   Encontrados    : " + str(len(encontrados)))
    print("   Nao encontrados: " + str(len(nao_encontrados)))
    return df_resultado, nao_encontrados


# ============================================
# FILTRAR POR OPPORTUNITIES
# ============================================
def filtrar_por_opportunities(df, minimo=MIN_OPPORTUNITIES):
    antes     = len(df)
    aprovados = []
    for _, row in df.iterrows():
        try:
            opps = float(str(row["Opportunities"]).replace(",", ".").strip())
            if opps >= minimo:
                aprovados.append(row)
        except (ValueError, TypeError):
            pass
    df_aprovado = pd.DataFrame(aprovados).reset_index(drop=True) if aprovados else pd.DataFrame(columns=df.columns)
    print("   FILTRO Opportunities >= " + str(minimo) + ": "
          + str(antes) + " -> " + str(len(df_aprovado)) + " linhas")
    return df_aprovado


# ============================================
# FILTRO EMBAIXADORES MIN-MAX ASSOCIADOS
# ============================================
def filtro_embaixadores_min_max(df, col_emb="Embaixador", col_assoc="Associado"):
    antes    = df[col_emb].nunique()
    contagem = df.groupby(col_emb)[col_assoc].nunique()
    validos  = contagem[(contagem >= MIN_ASSOC_POR_EMBAIXADOR) &
                        (contagem <= MAX_ASSOC_POR_EMBAIXADOR)].index
    rem_min  = contagem[contagem < MIN_ASSOC_POR_EMBAIXADOR]
    rem_max  = contagem[contagem > MAX_ASSOC_POR_EMBAIXADOR]
    df       = df[df[col_emb].isin(validos)].reset_index(drop=True)
    depois   = df[col_emb].nunique()
    print("   FILTRO " + str(MIN_ASSOC_POR_EMBAIXADOR) + "-" + str(MAX_ASSOC_POR_EMBAIXADOR)
          + " assoc/embaixador: " + str(antes) + " -> " + str(depois) + " embaixadores")
    if len(rem_min) > 0:
        print("      Removidos (< " + str(MIN_ASSOC_POR_EMBAIXADOR) + "): "
              + ", ".join([str(e) + "(" + str(q) + ")" for e, q in rem_min.items()]))
    if len(rem_max) > 0:
        print("      Removidos (> " + str(MAX_ASSOC_POR_EMBAIXADOR) + "): "
              + ", ".join([str(e) + "(" + str(q) + ")" for e, q in rem_max.items()]))
    return df


# ============================================
# CLASSIFICAR QUALIDADE POR EMBAIXADOR
# Metrica: DPMO Emb = (Soma Defects / Soma Opps) * 1.000.000
# Quality Rate (%) = DPMO Emb / Target do processo * 100
# TOP: Quality Rate <= QUALITY_TOP_MAX
# BOTTOM: Quality Rate > QUALITY_BOTTOM_MIN
# ============================================
def classificar_qualidade_embaixadores(df_qual, totais_processo):
    if len(df_qual) == 0:
        empty = pd.DataFrame(columns=["Embaixador", "DPMO Embaixador", "Percentual Embaixador"])
        return empty, empty


    ignorados = []
    metricas  = []


    for emb, grupo in df_qual.groupby("Embaixador"):


        # 1. Validar processo unico por embaixador
        # Exceção PACK: subprocessos diferentes (Single, Multi, Untraceable) são permitidos
        processos_unicos = grupo["Processo"].astype(str).str.strip().str.upper().unique()

        # Normaliza: trata todas as variantes de PACK como "PACK" para validação
        processos_normalizados = set()
        for p in processos_unicos:
            if "PACK" in p:
                processos_normalizados.add("PACK")
            else:
                processos_normalizados.add(p)

        if len(processos_normalizados) > 1:
            detalhes = ", ".join([
                str(r["Associado"]) + " -> " + str(r["Processo"])
                for _, r in grupo.iterrows()
            ])
            print()
            print("   AVISO: Embaixador '" + str(emb) + "' tem associados em processos diferentes:")
            print("          " + detalhes)
            print("          Possivel erro de filtragem. Embaixador ignorado no calculo de qualidade.")
            ignorados.append(str(emb))
            continue


        processo_upper = list(processos_normalizados)[0]


        # Para PACK: subprocessos diferentes (Single, Multi, Untraceable) sao permitidos
        # Nao bloqueia mais — soma tudo junto


        # 2. Somar Total Defects e Opportunities de todos os AAs
        try:
            soma_defects = grupo["Total Defects"].apply(
                lambda x: float(str(x).replace(",", ".").strip())
                if str(x).strip() not in ("", "nan", "-") else 0.0
            ).sum()
            soma_opps = grupo["Opportunities"].apply(
                lambda x: float(str(x).replace(",", ".").strip())
                if str(x).strip() not in ("", "nan", "-") else 0.0
            ).sum()
        except Exception as e:
            print("   AVISO: Erro ao somar valores do embaixador '" + str(emb) + "': " + str(e))
            ignorados.append(str(emb))
            continue


        if soma_opps <= 0:
            print("   AVISO: Embaixador '" + str(emb) + "' tem Opportunities = 0. Ignorado.")
            ignorados.append(str(emb))
            continue


        # 3. Calcular DPMO do Embaixador
        dpmo_emb = round((soma_defects / soma_opps) * 1000000, 2)


        # 4. Buscar target do processo nos totais
        if processo_upper == "PACK":
            # PACK pode ter múltiplos subprocessos — calcula target ponderado
            subs_presentes = grupo["Subprocess"].astype(str).str.strip().str.lower().unique()
            targets_pack = {}
            for sub in subs_presentes:
                if "single" in sub:
                    t = totais_processo.get("PACK_SINGLE", None)
                    if t and t > 0:
                        targets_pack["PACK_SINGLE"] = t
                elif "multi" in sub:
                    t = totais_processo.get("PACK_MULTI", None)
                    if t and t > 0:
                        targets_pack["PACK_MULTI"] = t
                # Untraceable e outros: tenta PACK_SINGLE como fallback
                else:
                    t = totais_processo.get("PACK_SINGLE", None)
                    if t and t > 0:
                        targets_pack["PACK_FALLBACK"] = t

            if targets_pack:
                # Usa a media dos targets encontrados
                target = round(sum(targets_pack.values()) / len(targets_pack), 2)
            else:
                target = None
            chave_target = "PACK (media)"
        else:
            chave_target = processo_upper
            target = totais_processo.get(chave_target, None)


        if target is None or target <= 0:
            print("   AVISO: Target nao encontrado para '" + chave_target + "'"
                  + " do embaixador '" + str(emb) + "'. Ignorado.")
            ignorados.append(str(emb))
            continue


        # 5. Calcular Quality Rate
        quality_rate = round((dpmo_emb / target) * 100, 2)


        metricas.append({
            "Embaixador":          emb,
            "Processo":            processo_upper,
            "Soma Defects":        soma_defects,
            "Soma Opps":           soma_opps,
            "DPMO Embaixador":     dpmo_emb,
            "Target DPMO":         target,
            "Percentual Embaixador": quality_rate,
        })


    if not metricas:
        empty = pd.DataFrame(columns=["Embaixador", "DPMO Embaixador", "Percentual Embaixador"])
        return empty, empty


    df_metricas = pd.DataFrame(metricas)


    # 6. Classificar: TOP <= 60% | BOTTOM > 100%
    emb_top    = df_metricas[df_metricas["Percentual Embaixador"] <= QUALITY_TOP_MAX].sort_values(
        "Percentual Embaixador", ascending=True).reset_index(drop=True)
    emb_bottom = df_metricas[df_metricas["Percentual Embaixador"] > QUALITY_BOTTOM_MIN].sort_values(
        "Percentual Embaixador", ascending=False).reset_index(drop=True)


    print("   Qualidade -> TOP  (Percentual <= " + str(QUALITY_TOP_MAX) + "%): "
          + str(len(emb_top)) + " embaixadores")
    print("   Qualidade -> BOTTOM (Percentual > " + str(QUALITY_BOTTOM_MIN) + "%): "
          + str(len(emb_bottom)) + " embaixadores")
    if ignorados:
        print("   Ignorados (erro de processo): " + str(len(ignorados))
              + " -> " + ", ".join(ignorados))


    return emb_top, emb_bottom


# ============================================
# NORMALIZAR NOME DE PROCESSO
# Pack: SEMPRE Single ou Multi (nunca so "Pack")
# Ship/TransferOut -> Doca
# Demais: Title Case padrao
# ============================================
def normalizar_processo(nome, subprocess_nome=None):
    if not nome or str(nome).strip() in ("", "-", "nan", "NAO ENCONTRADO"):
        return str(nome) if nome else ""
    key = str(nome).strip().lower()


    # Pack: SEMPRE distingue Single ou Multi
    if "pack" in key:
        sub = str(subprocess_nome or "").strip().lower()
        # Verifica subprocess externo (Atlas: col Subprocess)
        if "multi" in sub:
            return "Pack Multi"
        if "single" in sub:
            return "Pack Single"
        # Verifica se Single/Multi ja esta no proprio nome (prod: "Pack Singles")
        if "multi" in key:
            return "Pack Multi"
        return "Pack Single"   # fallback padrao


    # Busca match exato
    if key in PROCESSO_NORMALIZADO:
        return PROCESSO_NORMALIZADO[key]


    # Busca match parcial (ex: "Stow To Prime" contem "stow to prime")
    for k, v in PROCESSO_NORMALIZADO.items():
        if k in key:
            return v


    # Fallback: Title Case
    return str(nome).strip().title()


# ============================================
# CONVERTER RATE PARA NUMERO
# ============================================
def converter_rate_para_numero(valor):
    try:
        s   = str(valor).strip().replace('%', '').replace(',', '.')
        num = float(s)
        if 0 <= num <= 3.0:
            return round(num * 100, 2)
        return round(num, 2)
    except (ValueError, TypeError):
        return None


def converter_horas_para_numero(valor):
    try:
        s = str(valor).strip().replace(',', '.').replace('h', '').replace('H', '')
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


# ============================================
# CLASSIFICAR PRODUTIVIDADE POR EMBAIXADOR
# ============================================
def classificar_produtividade_embaixadores(df_prod):
    if len(df_prod) == 0:
        empty = pd.DataFrame(columns=["Embaixador", "Rate LC Medio (%)"])
        return empty, empty
    df_val          = df_prod.copy()
    df_val["_rate"] = df_val["Rate LC"].apply(converter_rate_para_numero)
    df_val          = df_val.dropna(subset=["_rate"])
    media           = (df_val.groupby("Embaixador")["_rate"]
                       .mean().reset_index()
                       .rename(columns={"_rate": "Rate LC Medio (%)"}))
    media["Rate LC Medio (%)"] = media["Rate LC Medio (%)"].round(2)
    emb_top    = media[media["Rate LC Medio (%)"] >= PROD_TOP_MINIMO].sort_values(
        "Rate LC Medio (%)", ascending=False).reset_index(drop=True)
    emb_bottom = media[media["Rate LC Medio (%)"] <= PROD_BOTTOM_MAXIMO].sort_values(
        "Rate LC Medio (%)", ascending=True).reset_index(drop=True)
    print("   Produtividade -> TOP (rate medio >= " + str(PROD_TOP_MINIMO) + "%): "
          + str(len(emb_top)) + " embaixadores")
    print("   Produtividade -> BOTTOM (rate medio <= " + str(PROD_BOTTOM_MAXIMO) + "%): "
          + str(len(emb_bottom)) + " embaixadores")
    return emb_top, emb_bottom


# ============================================
# LOGICA RANKING FINAL
# TOP e BOTTOM: Qualidade tem prioridade
# Overlap (aparece nos dois): fica como Qualidade
# BOTTOM: limitado a ceil(TOP/2)+1 (impar) ou +2 (par)
# ============================================
def selecionar_50_50(emb_top_prod, emb_top_qual, emb_bot_prod, emb_bot_qual):


    l_tp = list(emb_top_prod["Embaixador"]) if len(emb_top_prod) > 0 else []
    l_tq = list(emb_top_qual["Embaixador"]) if len(emb_top_qual) > 0 else []
    l_bp = list(emb_bot_prod["Embaixador"]) if len(emb_bot_prod) > 0 else []
    l_bq = list(emb_bot_qual["Embaixador"]) if len(emb_bot_qual) > 0 else []


    # TOP: Qualidade tem prioridade
    # 1o: todos de qual -> Fonte = Qualidade
    # 2o: de prod apenas quem NAO esta em qual -> Fonte = Produtividade
    set_tq      = set(l_tq)
    overlap_top = [e for e in l_tq if e in set(l_tp)]
    tops  = [{"Embaixador": e, "Classificacao": "TOP", "Fonte": "Qualidade"}
             for e in l_tq]
    tops += [{"Embaixador": e, "Classificacao": "TOP", "Fonte": "Produtividade"}
             for e in l_tp if e not in set_tq]


    # BOTTOM: Qualidade tem prioridade (mesma logica)
    set_bq         = set(l_bq)
    overlap_bottom = [e for e in l_bq if e in set(l_bp)]
    bottoms  = [{"Embaixador": e, "Classificacao": "BOTTOM", "Fonte": "Qualidade"}
                for e in l_bq]
    bottoms += [{"Embaixador": e, "Classificacao": "BOTTOM", "Fonte": "Produtividade"}
                for e in l_bp if e not in set_bq]


    # Limitador dinamico do BOTTOM:
    # TOP = 2 (excecao) -> max 1 BOTTOM
    # TOP par           -> max TOP/2 + 1   (ex: 8 TOP = 5 BOTTOM)
    # TOP impar         -> max ceil(TOP/2) (ex: 5 TOP = 3 BOTTOM)
    if len(tops) > 0:
        n = len(tops)
        if n == 2:
            max_bottom = 1
            formula_str = "2 TOP = 1 BOTTOM (excecao)"
        elif n % 2 == 0:
            max_bottom  = n // 2 + 1
            formula_str = str(n) + "/2 + 1 = " + str(max_bottom)
        else:
            max_bottom  = math.ceil(n / 2)
            formula_str = "ceil(" + str(n) + "/2) = " + str(max_bottom)
        if len(bottoms) > max_bottom:
            bottoms = bottoms[:max_bottom]
            print("      BOTTOM limitado a " + str(max_bottom) + "  (" + formula_str + ")")


    n_tp = len([x for x in tops    if x["Fonte"] == "Produtividade"])
    n_tq = len([x for x in tops    if x["Fonte"] == "Qualidade"])
    n_bp = len([x for x in bottoms if x["Fonte"] == "Produtividade"])
    n_bq = len([x for x in bottoms if x["Fonte"] == "Qualidade"])


    print("\n   SELECAO FINAL:")
    print("      TOP    (qual prioridade): " + str(n_tq) + " qual + " + str(n_tp)
          + " prod = " + str(len(tops)) + " total")
    if overlap_top:
        print("         Overlap TOP: " + ", ".join(overlap_top)
              + " -> priorizados como Qualidade")
    print("      BOTTOM (qual prioridade): " + str(n_bq) + " qual + " + str(n_bp)
          + " prod = " + str(len(bottoms)) + " total")
    if overlap_bottom:
        print("         Overlap BOTTOM: " + ", ".join(overlap_bottom)
              + " -> priorizados como Qualidade")
    return tops + bottoms


# ============================================
# PRODUTIVIDADE - AUTO DETECTAR LOGIN
# ============================================
def auto_detectar_coluna_login(df, logins_amostra):
    logins_set     = set(str(l).lower().strip() for l in logins_amostra)
    melhor_coluna  = -1
    melhor_matches = 0
    for i in range(len(df.columns)):
        try:
            valores = (df.iloc[:, i].astype(str).str.strip()
                       .str.replace(r'\.0$', '', regex=True).str.lower())
            matches = sum(1 for v in valores if v in logins_set)
            if matches > melhor_matches:
                melhor_matches = matches
                melhor_coluna  = i
        except Exception:
            continue
    if melhor_coluna >= 0 and melhor_matches > 0:
        print("      Login: '" + str(df.columns[melhor_coluna])
              + "' (idx " + str(melhor_coluna) + ") -> " + str(melhor_matches) + " matches")
    return melhor_coluna


# ============================================
# PRODUTIVIDADE - EXTRAIR DADOS DE UMA ABA
# ============================================
def extrair_dados_aba(df_aba, idx_login, idx_nome, idx_turno,
                      idx_processo, idx_lc_level, idx_uph_ppr, idx_horas,
                      idx_rate_lc, idx_rate_real, idx_target_ppr=-1):
    max_idx       = len(df_aba.columns) - 1
    idx_login     = min(idx_login,      max_idx)
    idx_nome      = min(idx_nome,       max_idx)
    idx_turno     = min(idx_turno,      max_idx)
    idx_processo  = min(idx_processo,   max_idx)
    idx_lc_level  = min(idx_lc_level,   max_idx)
    idx_uph_ppr   = min(idx_uph_ppr,    max_idx)
    idx_horas     = min(idx_horas,      max_idx)
    idx_rate_lc   = min(idx_rate_lc,    max_idx)
    idx_rate_real = min(idx_rate_real,  max_idx)
    # TARGET_PPR: coluna O (idx 14) — usa -1 se nao disponivel
    usar_target   = (idx_target_ppr >= 0 and idx_target_ppr <= max_idx)
    df = pd.DataFrame({
        "Login":             df_aba.iloc[:, idx_login],
        "Nome":              df_aba.iloc[:, idx_nome],
        "Turno":             df_aba.iloc[:, idx_turno],
        "Processo":          df_aba.iloc[:, idx_processo],
        "LC LEVEL":          df_aba.iloc[:, idx_lc_level],
        "UPH PPR":           df_aba.iloc[:, idx_uph_ppr],
        "Horas de processo": df_aba.iloc[:, idx_horas],
        "Rate LC":           df_aba.iloc[:, idx_rate_lc],
        "Rate Real":         df_aba.iloc[:, idx_rate_real],
        "TARGET_PPR":        (df_aba.iloc[:, idx_target_ppr] if usar_target
                              else pd.Series("", index=df_aba.index)),
    })
    df = df.dropna(subset=["Login"])
    df["Login"]    = (df["Login"].astype(str).str.strip()
                      .str.replace(r'\.0$', '', regex=True))
    df["Nome"]     = df["Nome"].astype(str).str.strip()
    df["Turno"]    = df["Turno"].astype(str).str.strip()
    df["Processo"] = df["Processo"].astype(str).str.strip()
    df["LC LEVEL"] = df["LC LEVEL"].astype(str).str.strip().str.upper()
    df = df[(df["Login"] != "") & (df["Login"].str.lower() != "nan")]
    df["Processo"]    = df["Processo"].apply(normalizar_processo)
    df["Login_busca"] = df["Login"].str.lower().str.strip()
    return df


# ============================================
# PRODUTIVIDADE - PRIORIDADE LC
# ============================================
def prioridade_lc(valor):
    valor = str(valor).strip().upper()
    if valor == "LC3":   return 3
    elif valor == "LC2": return 2
    elif valor == "LC1": return 1
    else:
        m = re.search(r'(\d+)', valor)
        return int(m.group(1)) if m else 0


def deduplicar_por_lc_level(df):
    if len(df) == 0 or "LC LEVEL" not in df.columns:
        return df
    df["_lc_prioridade"] = df["LC LEVEL"].apply(prioridade_lc)
    antes   = len(df)
    mask_nf = df["Rate LC"] == "NAO ENCONTRADO"
    df_nf   = df[mask_nf].copy()
    df_enc  = df[~mask_nf].copy()
    df_enc  = (df_enc.sort_values("_lc_prioridade", ascending=False)
                     .drop_duplicates(subset=["Associado"], keep="first"))
    df_r    = (pd.concat([df_enc, df_nf], ignore_index=True)
                 .drop(columns=["_lc_prioridade"]))
    removidos = antes - len(df_r)
    if removidos > 0:
        print("   DEDUP LC: removidas " + str(removidos) + " linhas (LC inferior)")
    return df_r


# ============================================
# PRODUTIVIDADE - PROCESSAR TODAS AS ABAS
# ============================================
def processar_todas_abas_dados(caminho, df_associacoes):
    xl         = pd.ExcelFile(caminho, engine="openpyxl")
    nomes_abas = xl.sheet_names
    print("\n   Abas: " + ", ".join(nomes_abas))
    associados_list    = list(df_associacoes["Associado"])
    todos_resultados   = []
    encontrados_global = set()


    for nome_aba in nomes_abas:
        print("\n   " + "=" * 48)
        print("   ABA: '" + nome_aba + "'")
        print("   " + "=" * 48)
        try:
            df_aba, header_row = ler_excel_com_header(caminho, nome_aba)
            print("      " + str(len(df_aba)) + " linhas")
        except Exception as e:
            print("      Erro: " + str(e))
            continue
        if len(df_aba) == 0:
            continue
        mostrar_colunas(df_aba)
        max_idx = len(df_aba.columns) - 1


        if IDX_LOGIN_FUNC == -1:
            idx_login = auto_detectar_coluna_login(df_aba, associados_list)
            if idx_login == -1:
                print("      Login nao encontrado, pulando...")
                continue
        else:
            idx_login = min(IDX_LOGIN_FUNC, max_idx)


        df_func = extrair_dados_aba(
            df_aba, idx_login,
            min(IDX_NOME_FUNC,       max_idx),
            min(IDX_TURNO_FUNC,      max_idx),
            min(IDX_PROCESSO_FUNC,   max_idx),
            min(IDX_LC_LEVEL_FUNC,   max_idx),
            min(IDX_UPH_PPR_FUNC,    max_idx),
            min(IDX_HORAS_FUNC,      max_idx),
            min(IDX_RATE_LC_FUNC,    max_idx),
            min(IDX_RATE_REAL_FUNC,  max_idx),
            min(IDX_TARGET_PPR_FUNC, max_idx),   # TARGET_PPR col O
        )
        print("      " + str(len(df_func)) + " funcionarios")
        encontrados_aba = 0


        for _, row in df_associacoes.iterrows():
            associado  = row["Associado"]
            embaixador = row["Embaixador"]
            busca      = df_func[df_func["Login_busca"] == associado.lower().strip()]
            if not busca.empty:
                for _, dados in busca.iterrows():
                    encontrados_aba += 1
                    encontrados_global.add(associado)
                    todos_resultados.append({
                        "Embaixador":        embaixador,
                        "LC LEVEL":          dados["LC LEVEL"],
                        "Associado":         associado,
                        "Nome":              dados["Nome"],
                        "Turno":             dados["Turno"],
                        "Processo":          dados["Processo"],
                        "UPH PPR":           dados["UPH PPR"],
                        "Horas de processo": dados["Horas de processo"],
                        "Rate LC":           dados["Rate LC"],
                        "Rate Real":         dados["Rate Real"],
                        "TARGET_PPR":        dados.get("TARGET_PPR", ""),
                    })
                primeiro       = busca.iloc[0]
                lc_encontrados = list(busca["LC LEVEL"].unique())
                print("      " + str(associado).ljust(15) + " -> "
                      + str(primeiro["Nome"]).ljust(18)
                      + " | LC: " + ", ".join(lc_encontrados)
                      + " | Rate: " + str(primeiro["Rate LC"]))
        print("      " + str(encontrados_aba) + " registros encontrados")


    nao_encontrados = []
    for _, row in df_associacoes.iterrows():
        if row["Associado"] not in encontrados_global:
            nao_encontrados.append(row["Associado"])
            todos_resultados.append({
                "Embaixador":        row["Embaixador"],
                "LC LEVEL":          "-",
                "Associado":         row["Associado"],
                "Nome":              "-",
                "Turno":             "-",
                "Processo":          "-",
                "UPH PPR":           "NAO ENCONTRADO",
                "Horas de processo": "-",
                "Rate LC":           "NAO ENCONTRADO",
                "Rate Real":         "NAO ENCONTRADO",
            })


    df_resultado  = pd.DataFrame(todos_resultados)
    df_resultado  = deduplicar_por_lc_level(df_resultado)
    colunas_ordem = ["Embaixador", "LC LEVEL", "Associado", "Nome", "Turno",
                     "Processo", "UPH PPR", "Horas de processo",
                     "Rate LC", "Rate Real", "TARGET_PPR"]
    df_resultado  = df_resultado[[c for c in colunas_ordem if c in df_resultado.columns]]
    if len(df_resultado) > 0:
        df_resultado = df_resultado.sort_values(
            ["Embaixador", "Associado"]).reset_index(drop=True)
    return df_resultado, nao_encontrados


def filtro_horas_minimas(df, minimo=HORAS_MINIMAS):
    antes            = len(df)
    df["_horas_num"] = df["Horas de processo"].apply(converter_horas_para_numero)
    df = df[df["_horas_num"] >= minimo].copy().drop(columns=["_horas_num"]).reset_index(drop=True)
    print("   FILTRO HORAS >= " + str(minimo) + "h: " + str(antes) + " -> " + str(len(df)))
    return df


# ============================================
# HELPERS ESTILOS EXCEL
# ============================================
def _estilos_excel():
    from openpyxl.styles import Alignment, Border, Side
    borda  = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")
    return borda, center


def _escrever_secao(ws, df, titulo, fill_titulo, fill_header, fill_par, fill_impar,
                    colunas, linha, borda, center, font_titulo, font_header, font_dado):
    num_cols = len(colunas)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=num_cols)
    cell = ws.cell(row=linha, column=1, value=titulo)
    cell.font = font_titulo; cell.fill = fill_titulo
    cell.alignment = center; cell.border = borda
    for c in range(2, num_cols + 1):
        ws.cell(row=linha, column=c).fill   = fill_titulo
        ws.cell(row=linha, column=c).border = borda
    linha += 1
    for j, col in enumerate(colunas, 1):
        cell = ws.cell(row=linha, column=j, value=col)
        cell.font = font_header; cell.fill = fill_header
        cell.alignment = center; cell.border = borda
    linha += 1
    if len(df) > 0:
        for idx_row, (_, data_row) in enumerate(df.iterrows()):
            fill = fill_par if idx_row % 2 == 0 else fill_impar
            for j, col in enumerate(colunas, 1):
                v    = data_row[col] if col in data_row else ""
                cell = ws.cell(row=linha, column=j, value=v)
                cell.font = font_dado; cell.fill = fill
                cell.alignment = center; cell.border = borda
            linha += 1
    else:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=num_cols)
        ws.cell(row=linha, column=1, value="Nenhum registro encontrado")
        linha += 1
    return linha


def _ajustar_largura(ws, dfs, colunas):
    for j, col in enumerate(colunas, 1):
        max_len = len(str(col))
        for df in dfs:
            if len(df) > 0 and col in df.columns:
                max_len = max(max_len, df[col].astype(str).str.len().max())
        letra = chr(64 + j) if j <= 26 else chr(64 + j // 26) + chr(65 + (j - 1) % 26)
        ws.column_dimensions[letra].width = min(max_len + 4, 35)


# ============================================
# ARQUIVO 1: resultado_completo.xlsx
# Mostra TODOS os embaixadores (TOP, MEIO e BOTTOM)
# Col A: Data (datas das abas de associados.xlsx)
# Col C: Classificacao (TOP / MEIO / BOTTOM)
# Cor da linha: Verde=TOP | Amarelo=MEIO | Vermelho=BOTTOM
# ============================================
def salvar_resultado_completo(df_qual_all, df_prod_all, caminho, data_str=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill


    borda, center = _estilos_excel()
    ft_t = Font(bold=True, color="FFFFFF", size=13)
    ft_h = Font(bold=True, color="FFFFFF", size=11)
    ft_d = Font(size=10)


    def p(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")


    fill_titulo  = p("0D47A1")   # azul escuro - titulo
    fill_header  = p("1565C0")   # azul medio  - cabeçalho
    fill_por_class = {
        "TOP":      p("C8E6C9"),   # verde claro
        "MEIO":     p("FFF9C4"),   # amarelo claro
        "BOTTOM":   p("FFCDD2"),   # vermelho claro
        "INVALIDO": p("FFE0B2"),   # laranja claro
    }
    fill_default = p("ECEFF1")   # cinza claro - sem classificacao


    cols_q = ["Data", "Embaixador",
              "Associado", "Processo", "Subprocess",
              "Total Defects", "Opportunities", "DPMO",
              "Percentual Embaixador", "Classificacao"]
    cols_p = ["Data", "Embaixador",
              "Associado", "Turno",
              "Processo", "UPH PPR", "Horas de processo",
              "Rate LC", "Rate Real", "Classificacao"]


    n_qual = df_qual_all["Embaixador"].nunique() if len(df_qual_all) > 0 else 0
    n_prod = df_prod_all["Embaixador"].nunique() if len(df_prod_all) > 0 else 0


    def escrever_aba(ws, df_all, cols, titulo):
        num_cols = len(cols)
        # Linha 1: titulo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = ft_t; c.fill = fill_titulo
        c.alignment = center; c.border = borda
        for k in range(2, num_cols + 1):
            ws.cell(row=1, column=k).fill   = fill_titulo
            ws.cell(row=1, column=k).border = borda
        # Linha 2: cabeçalhos
        for j, col in enumerate(cols, 1):
            c = ws.cell(row=2, column=j, value=col)
            c.font = ft_h; c.fill = fill_header
            c.alignment = center; c.border = borda
        # Linhas 3+: dados com cor por classificacao
        if len(df_all) > 0:
            for r_idx, (_, row) in enumerate(df_all.iterrows(), start=3):
                classif = str(row.get("Classificacao", "")).strip()
                fill_r  = fill_por_class.get(classif, fill_default)
                for j, col in enumerate(cols, 1):
                    v = data_str if col == "Data" else (row[col] if col in row.index else "")
                    c = ws.cell(row=r_idx, column=j, value=v)
                    c.font = ft_d; c.fill = fill_r
                    c.alignment = center; c.border = borda
        else:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
            ws.cell(row=3, column=1, value="Nenhum registro encontrado")
        _ajustar_largura(ws, [df_all], cols)
        ws.freeze_panes = "A3"


    legenda = "   |   Verde=TOP  Amarelo=MEIO  Vermelho=BOTTOM  Laranja=INVALIDO"
    for tentativa in [str(caminho), str(caminho).replace(".xlsx", "_2.xlsx")]:
        try:
            wb   = Workbook()
            ws_q = wb.active
            ws_q.title = "Qualidade"
            escrever_aba(ws_q, df_qual_all, cols_q,
                         "QUALIDADE — TODOS OS EMBAIXADORES — "
                         + str(n_qual) + " embaixadores" + legenda)
            ws_p = wb.create_sheet(title="Produtividade")
            escrever_aba(ws_p, df_prod_all, cols_p,
                         "PRODUTIVIDADE — TODOS OS EMBAIXADORES — "
                         + str(n_prod) + " embaixadores" + legenda)
            wb.save(tentativa)
            print("\n   ARQUIVO 1 salvo: " + os.path.abspath(tentativa))
            # Contagem por classificacao (nivel embaixador)
            for nome_df, label in [(df_qual_all, "Qualidade"), (df_prod_all, "Produtiv.")]:
                if len(nome_df) > 0 and "Classificacao" in nome_df.columns:
                    cnt = nome_df.drop_duplicates("Embaixador")["Classificacao"].value_counts()
                    print("      " + label.ljust(10) + ": TOP=" + str(cnt.get("TOP", 0))
                          + " | MEIO=" + str(cnt.get("MEIO", 0))
                          + " | BOTTOM=" + str(cnt.get("BOTTOM", 0))
                          + " | INVALIDO=" + str(cnt.get("INVALIDO", 0)))
            return
        except PermissionError:
            print("   '" + tentativa + "' aberto! Tentando outro nome...")
    print("   Feche o arquivo e rode novamente.")


# ============================================
# HELPER: RECONSTRÓI ABAS "TOP Performer" e "BOTTOM Performer"
# Lê TODOS os TOP / BOTTOM de Qualidade e Produtividade
# do historico (+ semente do resultado_completo) e
# reconstrói as duas abas do zero, em ordem cronológica.
# Sem cor, negrito ou formatação — mala direta.
# Colunas (ambas as abas):
#   A: Data ONB
#   B: Login Embaixador
#   C: Nome Embaixador
#   D: Classificacao (Qualidade / Produtividade)
#   E: Performace ("TOP" ou "BOTTOM")
# ============================================
def _atualizar_abas_performer(wb_hist, tops_from_src=None, bottoms_from_src=None):


    def _construir_aba(nome_aba, performace_label, sementes, classif_hist):
        # Semente garantida vinda do resultado_completo (100% confiavel)
        por_chave = {}
        if sementes:
            for chave, abas in sementes.items():
                por_chave[chave] = set(abas)
            print("   " + nome_aba + ": " + str(len(por_chave))
                  + " registros injetados do resultado_completo")


        # Tambem le o historico (Qualidade/Produtividade) para rodadas anteriores
        for sheet_name in ["Qualidade", "Produtividade"]:
            if sheet_name not in wb_hist.sheetnames:
                continue
            ws_sheet = wb_hist[sheet_name]
            emb_idx = data_idx = class_idx = None
            for row in ws_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                for j, v in enumerate(row, 1):
                    col_nome = str(v).strip().lower() if v is not None else ""
                    if col_nome == "embaixador":
                        emb_idx = j
                    elif col_nome == "data":
                        data_idx = j
                    elif col_nome in ("classificacao", "classificação",
                                       "qualificacao", "qualificação"):
                        class_idx = j
                break
            if not (emb_idx and class_idx):
                continue
            n_found = 0
            for row in ws_sheet.iter_rows(min_row=2, values_only=True):
                values  = list(row)
                classif = ""
                if class_idx <= len(values) and values[class_idx - 1] is not None:
                    classif = str(values[class_idx - 1]).strip().upper()
                if classif != classif_hist:
                    continue
                emb = ""
                if emb_idx <= len(values) and values[emb_idx - 1] is not None:
                    emb = str(values[emb_idx - 1]).strip()
                data = ""
                if data_idx and data_idx <= len(values) and values[data_idx - 1] is not None:
                    data = str(values[data_idx - 1]).strip()
                if not emb or emb.lower() in ("", "nan", "none"):
                    continue
                chave = (emb, data)
                if chave not in por_chave:
                    por_chave[chave] = set()
                por_chave[chave].add(sheet_name)
                n_found += 1
            print("   " + nome_aba + " [" + sheet_name + "]: "
                  + str(n_found) + " registros no historico")


        # Preserva os Nomes (Col C) ja existentes na aba antiga
        nomes_existentes = {}
        if nome_aba in wb_hist.sheetnames:
            ws_old = wb_hist[nome_aba]
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                vals = list(row)
                lg = str(vals[1]).strip() if len(vals) > 1 and vals[1] is not None else ""
                nm = str(vals[2]).strip() if len(vals) > 2 and vals[2] is not None else ""
                dt = str(vals[0]).strip() if len(vals) > 0 and vals[0] is not None else ""
                if lg and nm and nm.lower() not in ("", "nan", "none"):
                    nomes_existentes[(lg, dt)] = nm


        # Ordena por data cronologica (mes, dia) e depois por login
        def sort_key(item):
            emb, data = item[0]
            primeira = str(data).split("|")[0].strip()
            try:
                p = primeira.split(".")
                return (int(p[1]), int(p[0]), emb.lower())
            except Exception:
                return (99, 99, emb.lower())


        ordenados = sorted(por_chave.items(), key=sort_key)


        # Reconstrói a aba do zero
        if nome_aba in wb_hist.sheetnames:
            del wb_hist[nome_aba]
        ws_new = wb_hist.create_sheet(title=nome_aba)


        ws_new.cell(row=1, column=1, value="Data ONB")
        ws_new.cell(row=1, column=2, value="Login Embaixador")
        ws_new.cell(row=1, column=3, value="Nome Embaixador")
        ws_new.cell(row=1, column=4, value="Classificacao")
        ws_new.cell(row=1, column=5, value="Performace")
        # Coluna F: RECORRENCIA (contagem de quantas vezes o login aparece na aba)
        cab_rec = "RECORRENCIA" if performace_label == "TOP" else "RECORRÊNCIA"
        ws_new.cell(row=1, column=6, value=cab_rec)


        # Regra: Qualidade tem prioridade absoluta sobre Produtividade
        for r_idx, ((emb, data), abas) in enumerate(ordenados, start=2):
            classif_str = "Qualidade" if "Qualidade" in abas else "Produtividade"
            nome_prev   = nomes_existentes.get((emb, data), "")
            ws_new.cell(row=r_idx, column=1, value=data)
            ws_new.cell(row=r_idx, column=2, value=emb)
            ws_new.cell(row=r_idx, column=3, value=nome_prev)
            ws_new.cell(row=r_idx, column=4, value=classif_str)
            ws_new.cell(row=r_idx, column=5, value=performace_label)
            # Arrasta a formula de recorrencia (por login + classificacao,
            # acumulada ate a propria linha) para a nova linha.
            r = str(r_idx)
            ws_new.cell(row=r_idx, column=6,
                        value="=COUNTIFS($B$2:$B" + r + ",B" + r
                              + ",$D$2:$D" + r + ",D" + r + ")")


        print("   " + nome_aba + ": " + str(len(ordenados))
              + " registros (reconstruido — ordem cronologica)")
        return len(ordenados)


    n_top = _construir_aba("TOP Performer", "TOP", tops_from_src, "TOP")
    n_bot = _construir_aba("BOTTOM Performer", "BOTTOM", bottoms_from_src, "BOTTOM")
    return n_top, n_bot


# ============================================
# REGISTRAR ONB SEM DESTAQUE NO HISTORICO
# Quando uma data de ONB nao gera nenhum TOP nem BOTTOM,
# documenta a data numa aba propria ("ONBs sem destaque")
# para deixar registrado que aquele dia foi analisado.
# ============================================
def registrar_onb_sem_destaque(data_onb):
    from openpyxl import load_workbook, Workbook
    from datetime import datetime as _dt

    historico_path = HISTORICO_DIR / PLANILHA_HISTORICO
    data_onb = str(data_onb).strip() if data_onb else ""
    if not data_onb:
        data_onb = _dt.now().strftime("%d.%m")

    if historico_path.exists():
        garantir_arquivo_fechado(historico_path, "HistoricoAmbassadors.xlsx")
        try:
            wb = load_workbook(str(historico_path))
        except Exception as e:
            # PROTECAO: o arquivo existe mas nao abriu — NUNCA criar em branco,
            # pois isso apagaria todo o historico ao salvar por cima. Aborta.
            print("   ERRO: nao foi possivel abrir o historico para registrar"
                  " o ONB: " + str(e))
            print("   Registro ABORTADO para NAO apagar o historico.")
            return
    else:
        wb = Workbook()
        wb.remove(wb.active)

    nome_aba = "ONBs sem destaque"
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
    else:
        ws = wb.create_sheet(title=nome_aba)
        ws.cell(row=1, column=1, value="Data ONB")
        ws.cell(row=1, column=2, value="Observacao")
        ws.cell(row=1, column=3, value="Registrado em")

    # Evita duplicar a mesma data
    ja_registrada = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None and str(row[0]).strip() == data_onb:
            ja_registrada = True
            break

    if not ja_registrada:
        prox = ws.max_row + 1
        # se a primeira linha de dados estiver vazia, usa a linha 2
        if ws.max_row < 2:
            prox = 2
        ws.cell(row=prox, column=1, value=data_onb)
        ws.cell(row=prox, column=2,
                value="Nenhum embaixador (TOP ou BOTTOM) neste ONB")
        ws.cell(row=prox, column=3, value=_dt.now().strftime("%d/%m/%Y %H:%M"))

    try:
        wb.save(str(historico_path))
        if ja_registrada:
            print("   ONB " + data_onb
                  + " ja constava em 'ONBs sem destaque'.")
        else:
            print("   ONB " + data_onb
                  + " registrado em 'ONBs sem destaque'.")
    except Exception as e:
        print("   ERRO ao registrar ONB sem destaque: " + str(e))


# ============================================
# COPIAR RESULTADO COMPLETO PARA HISTORICO
# Append de dados em HistoricoAmbassadors.xlsx
# Cada rodada acumula novas linhas (nunca apaga)
# ============================================
def copiar_para_historico(caminho_resultado):
    from openpyxl import load_workbook, Workbook
    historico_path = HISTORICO_DIR / PLANILHA_HISTORICO
    print("\n   Copiando dados para o historico...")
    # O historico precisa estar fechado para ser gravado
    if historico_path.exists():
        garantir_arquivo_fechado(historico_path, "HistoricoAmbassadors.xlsx")
    try:
        wb_src = load_workbook(str(caminho_resultado), data_only=True)
    except Exception as e:
        print("   ERRO ao ler resultado_completo: " + str(e))
        return
    # Carrega ou cria o arquivo de historico
    if historico_path.exists():
        try:
            wb_hist = load_workbook(str(historico_path))
            print("   Historico existente carregado.")
        except Exception as e:
            # PROTECAO CRITICA: o arquivo existe mas nao abriu (aberto em
            # outro PC, travado pelo OneDrive/SharePoint, permissao, etc.).
            # NUNCA criar um workbook em branco aqui — isso apagaria todo o
            # historico ao salvar por cima. Aborta para preservar os dados.
            try:
                erro("Nao foi possivel ABRIR o historico existente: " + str(e))
                aviso("Operacao ABORTADA para NAO apagar o historico.")
                dica("Feche o HistoricoAmbassadors.xlsx em TODOS os computadores"
                     " (inclusive OneDrive/SharePoint online) e rode novamente.")
            except Exception:
                print("   ERRO: nao foi possivel abrir o historico: " + str(e))
                print("   Operacao ABORTADA para NAO apagar o historico.")
            return
    else:
        wb_hist = Workbook()
        wb_hist.remove(wb_hist.active)
        print("   Criando novo arquivo de historico...")
    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        # Obtem ou cria a aba no historico
        if sheet_name in wb_hist.sheetnames:
            ws_hist = wb_hist[sheet_name]
        else:
            ws_hist = wb_hist.create_sheet(title=sheet_name)
            # Escreve cabecalho (linha 2 do resultado = cabecalho real)
            for row in ws_src.iter_rows(min_row=2, max_row=2, values_only=True):
                for j, v in enumerate(row, 1):
                    ws_hist.cell(row=1, column=j, value=v)
                break


        # Detecta indices das colunas com formatacao especial
        # DPMO          -> numero inteiro (sem decimais)
        # Percentual    -> "35.50%"  (0.2f)
        dpmo_idx       = None
        percentual_idx = None
        for cell in ws_hist[1]:
            col_nome = str(cell.value).strip() if cell.value else ""
            if col_nome == "DPMO":
                dpmo_idx = cell.column
            elif col_nome == "Percentual Embaixador":
                percentual_idx = cell.column


        def formatar_valor(j, v):
            if v is None or str(v).strip() in ("", "None", "nan"):
                return v
            # DPMO -> inteiro
            if j == dpmo_idx:
                try:
                    return int(round(float(str(v).replace(",", "."))))
                except (ValueError, TypeError):
                    return v
            # Percentual Embaixador -> "35.50%"
            if j == percentual_idx:
                try:
                    num = float(str(v).replace(",", ".").replace("%", ""))
                    return "{:.2f}%".format(num)
                except (ValueError, TypeError):
                    return v
            return v


        # Proxima linha vazia no historico (busca real, ignora linhas em branco)
        # Percorre de baixo para cima para encontrar a ultima celula com conteudo
        next_row = 1
        for row_cells in ws_hist.iter_rows(min_row=1, max_row=ws_hist.max_row, values_only=True):
            if any(v is not None and str(v).strip() not in ("", "None", "nan")
                   for v in row_cells):
                next_row += 1
            else:
                # Linha vazia — nao incrementa, mas continua buscando
                pass
        # Recalcula: percorre de tras pra frente para achar a ultima linha real
        ultima_com_dado = 0
        for r in range(ws_hist.max_row, 0, -1):
            row_vals = [ws_hist.cell(row=r, column=c).value
                        for c in range(1, ws_hist.max_column + 1)]
            if any(v is not None and str(v).strip() not in ("", "None", "nan")
                   for v in row_vals):
                ultima_com_dado = r
                break
        next_row = ultima_com_dado + 1
        if next_row < 2:
            next_row = 2
        # Copia linhas de dados (pula titulo linha 1 e cabecalho linha 2)
        rows_added = 0
        for row in ws_src.iter_rows(min_row=3, values_only=True):
            values = list(row)
            if any(v is not None and str(v).strip() not in ("", "None", "nan")
                   for v in values):
                for j, v in enumerate(values, 1):
                    ws_hist.cell(row=next_row, column=j, value=formatar_valor(j, v))
                next_row   += 1
                rows_added += 1
        print("      " + sheet_name + ": " + str(rows_added) + " linhas adicionadas")


    # Coleta TOPs direto do resultado_completo (fonte garantida)
    # Header na linha 2, dados na linha 3+ — case-insensitive
    tops_from_src = {}
    bottoms_from_src = {}
    for sheet_name in wb_src.sheetnames:
        ws_src_sheet = wb_src[sheet_name]
        emb_idx   = None
        data_idx  = None
        class_idx = None
        for row in ws_src_sheet.iter_rows(min_row=2, max_row=2, values_only=True):
            for j, v in enumerate(row, 1):
                cn = str(v).strip().lower() if v is not None else ""
                if cn == "embaixador":
                    emb_idx = j
                elif cn == "data":
                    data_idx = j
                elif cn in ("classificacao", "classificação",
                            "qualificacao", "qualificação"):
                    class_idx = j
            break
        if not (emb_idx and class_idx):
            continue
        for row in ws_src_sheet.iter_rows(min_row=3, values_only=True):
            values  = list(row)
            classif = ""
            if class_idx <= len(values) and values[class_idx - 1] is not None:
                classif = str(values[class_idx - 1]).strip().upper()
            if classif not in ("TOP", "BOTTOM"):
                continue
            emb = ""
            if emb_idx <= len(values) and values[emb_idx - 1] is not None:
                emb = str(values[emb_idx - 1]).strip()
            data = ""
            if data_idx and data_idx <= len(values) and values[data_idx - 1] is not None:
                data = str(values[data_idx - 1]).strip()
            if not emb or emb.lower() in ("", "nan", "none"):
                continue
            chave = (emb, data)
            destino = tops_from_src if classif == "TOP" else bottoms_from_src
            if chave not in destino:
                destino[chave] = set()
            destino[chave].add(sheet_name)
    print("   TOPs extraidos: " + str(len(tops_from_src))
          + " | BOTTOMs extraidos: " + str(len(bottoms_from_src)))


    # Reconstrói as abas TOP Performer e BOTTOM Performer
    _atualizar_abas_performer(wb_hist, tops_from_src, bottoms_from_src)


    # PROTECAO: nunca salvar um historico vazio por cima de um existente.
    if not wb_hist.sheetnames:
        erro("Historico ficou sem abas — save ABORTADO para nao apagar dados.")
        return


    try:
        wb_hist.save(str(historico_path))
        print("   Historico salvo: " + str(historico_path))
    except Exception as e:
        erro("ERRO ao salvar historico: " + str(e))


# ============================================
# ARQUIVO 2: rc_cret.xlsx
# ============================================
def salvar_rc_cret(df_rc_top, df_rc_bot, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill


    borda, center = _estilos_excel()
    ft = Font(bold=True, color="FFFFFF", size=13)
    fh = Font(bold=True, color="FFFFFF", size=11)
    fd = Font(size=10)


    def p(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")


    lj_t = p("4A148C"); lj_h = p("6A1B9A"); lj_c = p("F3E5F5"); lj_a = p("E1BEE7")
    vm_t = p("B71C1C"); vm_h = p("C62828"); vm_c = p("FFEBEE"); vm_a = p("FFCDD2")


    cols = ["Embaixador", "LC LEVEL", "Associado", "Nome", "Turno",
            "Processo", "UPH PPR", "Horas de processo", "Rate LC", "Rate Real"]
    nrt  = df_rc_top["Embaixador"].nunique() if len(df_rc_top) > 0 else 0
    nrb  = df_rc_bot["Embaixador"].nunique() if len(df_rc_bot) > 0 else 0


    for tentativa in [str(caminho), str(caminho).replace(".xlsx", "_2.xlsx")]:
        try:
            wb    = Workbook()
            ws    = wb.active
            ws.title = "RC SORT"
            linha = 1
            linha = _escrever_secao(ws, df_rc_top,
                                    "TOP RC SORT  -  " + str(nrt) + " embaixadores",
                                    lj_t, lj_h, lj_c, lj_a, cols,
                                    linha, borda, center, ft, fh, fd)
            linha += 2
            _escrever_secao(ws, df_rc_bot,
                            "BOTTOM RC SORT  -  " + str(nrb) + " embaixadores",
                            vm_t, vm_h, vm_c, vm_a, cols,
                            linha, borda, center, ft, fh, fd)
            _ajustar_largura(ws, [df_rc_top, df_rc_bot], cols)
            ws.freeze_panes = "A3"
            wb.save(tentativa)
            print("\n   ARQUIVO 2 salvo: " + os.path.abspath(tentativa))
            print("      TOP: " + str(nrt) + " | BOTTOM: " + str(nrb))
            return
        except PermissionError:
            print("   '" + tentativa + "' aberto! Tentando outro nome...")
    print("   Feche o arquivo rc_cret.xlsx e rode novamente.")


# ============================================
# ENRIQUECER RANKING COM DADOS DOS ASSOCIADOS
# Para cada embaixador no ranking, busca os dados
# dos AAs nas planilhas de prod e qualidade.
# Retorna 1 linha por associado.
# ============================================
def enriquecer_ranking(ranking, df_prod_top, df_qual_top, df_prod_bot, df_qual_bot,
                       df_associacoes=None, df_prod_all=None, dpmo_map=None,
                       dpmo_assoc_map=None):
    # dpmo_map: {embaixador_lower: dpmo_int} — fallback para TOPs de produtividade
    dpmo_map = dpmo_map or {}
    # dpmo_assoc_map: {associado_lower: dpmo_int} — DPMO de cada associado
    dpmo_assoc_map = dpmo_assoc_map or {}
    # Turno: lookup pelo login do associado em df_associacoes (Shift Pattern - col F)
    turno_map = {}
    if df_associacoes is not None and "Turno" in df_associacoes.columns:
        for _, r in df_associacoes.iterrows():
            login = str(r.get("Associado", "")).strip().lower()
            turno = str(r.get("Turno", "")).strip()
            if login and turno and turno.lower() not in ("nan", ""):
                turno_map[login] = turno


    prod_all = df_prod_all if df_prod_all is not None and len(df_prod_all) > 0 else pd.DataFrame()


    linhas = []
    for entry in ranking:
        emb     = entry["Embaixador"]
        classif = entry["Classificacao"]


        df_prod = df_prod_top if classif == "TOP" else df_prod_bot
        df_qual = df_qual_top if classif == "TOP" else df_qual_bot


        prod_emb = df_prod[df_prod["Embaixador"] == emb] if len(df_prod) > 0 else pd.DataFrame()
        qual_emb = df_qual[df_qual["Embaixador"] == emb] if len(df_qual) > 0 else pd.DataFrame()


        # Coleta todos os associados unicos (prod primeiro, depois qual)
        assocs = []
        seen   = set()
        for df_ in [prod_emb, qual_emb]:
            if len(df_) > 0 and "Associado" in df_.columns:
                for a in df_["Associado"].tolist():
                    if str(a).strip() and str(a).strip() not in seen:
                        assocs.append(str(a).strip())
                        seen.add(str(a).strip())


        if not assocs:
            linhas.append({
                "Embaixador": emb, "Classificacao": classif, "Fonte": "",
                "Associado": "", "Turno": "", "Processo": "",
                "UPH PPR": "", "DPMO": "",
                "Rate LC": "", "Total Defects": "", "Opportunities": "",
                "DPMO Emb": "", "Rate Emb": "",
            })
            continue


        for assoc in assocs:
            turno = turno_map.get(assoc.lower(), "")


            # Busca dados de qualidade
            qual_row = None
            if len(qual_emb) > 0:
                q = qual_emb[qual_emb["Associado"].astype(str).str.strip() == assoc]
                if len(q) > 0:
                    qual_row = q.iloc[0]


            # Busca dados de produtividade (prod_emb primeiro, fallback prod_all)
            prod_row = None
            if len(prod_emb) > 0:
                p = prod_emb[prod_emb["Associado"].astype(str).str.strip() == assoc]
                if len(p) > 0:
                    prod_row = p.iloc[0]
            if prod_row is None and len(prod_all) > 0:
                p = prod_all[prod_all["Associado"].astype(str).str.strip() == assoc]
                if len(p) > 0:
                    prod_row = p.iloc[0]


            # QUALIDADE TEM PRIORIDADE ABSOLUTA
            if qual_row is not None:
                proc_q = str(qual_row.get("Processo", "")).strip()
                sub_q  = str(qual_row.get("Subprocess", "")).strip()
                # Col I: DPMO Embaixador (ambassador-level, inteiro)
                dpmo_emb_raw = qual_row.get("DPMO Embaixador", "")
                try:
                    dpmo_emb_str = (str(int(round(float(str(dpmo_emb_raw).replace(",", ".")))))
                                    if str(dpmo_emb_raw).strip() not in ("", "nan") else "")
                except (ValueError, TypeError):
                    dpmo_emb_str = ""
                # Col J: Percentual Embaixador (Quality Rate %, 2 casas)
                rate_emb_raw = qual_row.get("Percentual Embaixador", "")
                try:
                    rate_emb_str = ("{:.2f}%".format(float(str(rate_emb_raw).replace(",", ".").replace("%", "")))
                                    if str(rate_emb_raw).strip() not in ("", "nan") else "")
                except (ValueError, TypeError):
                    rate_emb_str = ""
                linhas.append({
                    "Embaixador":    emb,
                    "Classificacao": classif,
                    "Fonte":         "Qualidade",
                    "Associado":     assoc,
                    "Turno":         turno,
                    "Processo":      normalizar_processo(proc_q, sub_q),
                    "UPH PPR":       "",
                    "DPMO":          str(qual_row.get("DPMO", "")).strip(),
                    "Rate LC":       "",
                    "Total Defects": str(qual_row.get("Total Defects", "")).strip(),
                    "Opportunities": str(qual_row.get("Opportunities", "")).strip(),
                    "DPMO Emb":      dpmo_emb_str,
                    "Rate Emb":      rate_emb_str,
                })
            elif prod_row is not None:
                rate_raw = str(prod_row.get("Rate LC", "")).strip()
                rate_num = converter_rate_para_numero(rate_raw)
                rate_str = (str(int(round(rate_num))) + "%" if rate_num is not None else "")
                # Col J: Rate LC Medio (%) do embaixador — inteiro %
                rate_medio_raw = prod_row.get("Rate LC Medio (%)", "")
                try:
                    rate_emb_str = (str(int(round(float(str(rate_medio_raw).replace(",", "."))))) + "%"
                                    if str(rate_medio_raw).strip() not in ("", "nan") else "")
                except (ValueError, TypeError):
                    rate_emb_str = ""
                # DPMO Emb: fallback pelo mapa global (embaixador pode ter
                # DPMO mesmo sem ter sido classificado TOP/BOTTOM em qualidade)
                dpmo_fb = dpmo_map.get(str(emb).strip().lower(), "")
                dpmo_emb_str = str(dpmo_fb) if dpmo_fb not in ("", None) else ""
                # UPH PPR do associado (vem do dados_funcionarios via prod_row)
                uph_str = str(prod_row.get("UPH PPR", "")).strip()
                if uph_str.lower() in ("nan", "none"):
                    uph_str = ""
                # DPMO do associado (lookup no mapa por associado)
                dpmo_assoc = dpmo_assoc_map.get(str(assoc).strip().lower(), "")
                dpmo_assoc_str = str(dpmo_assoc) if dpmo_assoc not in ("", None) else ""
                linhas.append({
                    "Embaixador":    emb,
                    "Classificacao": classif,
                    "Fonte":         "Produtividade",
                    "Associado":     assoc,
                    "Turno":         turno,
                    "Processo":      normalizar_processo(str(prod_row.get("Processo", "")).strip()),
                    "UPH PPR":       uph_str,
                    "DPMO":          dpmo_assoc_str,
                    "Rate LC":       rate_str,
                    "Total Defects": "",
                    "Opportunities": "",
                    "DPMO Emb":      dpmo_emb_str,
                    "Rate Emb":      rate_emb_str,
                })
            else:
                linhas.append({
                    "Embaixador":    emb,
                    "Classificacao": classif,
                    "Fonte":         "",
                    "Associado":     assoc,
                    "Turno":         turno,
                    "Processo":      "",
                    "UPH PPR":       "",
                    "DPMO":          "",
                    "Rate LC":       "",
                    "Total Defects": "",
                    "Opportunities": "",
                    "DPMO Emb":      "",
                    "Rate Emb":      "",
                })


    cols = ["Embaixador", "Fonte", "Associado", "Turno",
            "Processo", "UPH PPR", "DPMO", "Rate LC", "Total Defects", "Opportunities",
            "DPMO Emb", "Rate Emb"]
    return pd.DataFrame(linhas) if linhas else pd.DataFrame(columns=cols)


# ============================================
# EXTRAIR TARGETS DE PRODUTIVIDADE POR PROCESSO
# Lê TARGET_PPR de df_prod_completo e agrupa por processo.
# Retorna dict: {processo_normalizado: valor_str}
# ============================================
def extrair_targets_prod(df_prod_completo):
    targets = {}
    if df_prod_completo is None or len(df_prod_completo) == 0:
        return targets
    if "TARGET_PPR" not in df_prod_completo.columns:
        return targets
    for _, row in df_prod_completo.iterrows():
        proc = str(row.get("Processo", "")).strip()
        tgt  = str(row.get("TARGET_PPR", "")).strip()
        if proc and tgt and tgt.lower() not in ("", "nan", "-", "nao encontrado"):
            if proc not in targets:
                try:
                    tgt_num = float(tgt.replace(",", "."))
                    targets[proc] = str(int(round(tgt_num)))
                except (ValueError, TypeError):
                    targets[proc] = tgt
    if targets:
        print("   Targets prod extraidos: "
              + " | ".join([p + "=" + v for p, v in sorted(targets.items())]))
    return targets


# ============================================
# ARQUIVO 3: ranking_final.xlsx
# ============================================
def salvar_ranking_final(ranking, caminho,
                         df_prod_top=None, df_qual_top=None,
                         df_prod_bot=None, df_qual_bot=None,
                         df_associacoes=None, df_prod_all=None,
                         data_backup=None,
                         targets_prod=None, totais_processo=None,
                         dpmo_map=None, dpmo_assoc_map=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill


    if not ranking:
        print("\n   Ranking vazio. Nada para salvar.")
        return


    borda, center = _estilos_excel()
    ft = Font(bold=True, color="FFFFFF", size=13)
    fh = Font(bold=True, color="FFFFFF", size=11)
    fd = Font(size=10)


    def p(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")


    vd_t = p("1B5E20"); vd_h = p("2E7D32"); vd_c = p("E8F5E9"); vd_a = p("C8E6C9")
    vm_t = p("B71C1C"); vm_h = p("C62828"); vm_c = p("FFEBEE"); vm_a = p("FFCDD2")


    # Normalizar DataFrames opcionais
    df_prod_top = df_prod_top if df_prod_top is not None else pd.DataFrame()
    df_qual_top = df_qual_top if df_qual_top is not None else pd.DataFrame()
    df_prod_bot = df_prod_bot if df_prod_bot is not None else pd.DataFrame()
    df_qual_bot = df_qual_bot if df_qual_bot is not None else pd.DataFrame()


    # Enriquecer ranking com dados dos associados
    df_enriquecido = enriquecer_ranking(ranking, df_prod_top, df_qual_top, df_prod_bot, df_qual_bot,
                                        df_associacoes=df_associacoes, df_prod_all=df_prod_all,
                                        dpmo_map=dpmo_map, dpmo_assoc_map=dpmo_assoc_map)


    df_top = df_enriquecido[df_enriquecido["Classificacao"] == "TOP"].reset_index(drop=True)
    df_bot = df_enriquecido[df_enriquecido["Classificacao"] == "BOTTOM"].reset_index(drop=True)


    n_top_emb = df_top["Embaixador"].nunique() if len(df_top) > 0 else 0
    n_bot_emb = df_bot["Embaixador"].nunique() if len(df_bot) > 0 else 0


    cols = ["Embaixador", "Fonte", "Associado", "Turno",
            "Processo", "UPH PPR", "DPMO", "Rate LC", "Total Defects", "Opportunities",
            "DPMO Emb", "Rate Emb"]


    for tentativa in [str(caminho), str(caminho).replace(".xlsx", "_2.xlsx")]:
        try:
            wb    = Workbook()
            ws    = wb.active
            ws.title = "Ranking Final"
            linha = 1
            linha = _escrever_secao(ws, df_top,
                                    "TOP AMBASSADORS  (100/100)  -  " + str(n_top_emb) + " embaixadores",
                                    vd_t, vd_h, vd_c, vd_a, cols,
                                    linha, borda, center, ft, fh, fd)
            linha += 2
            _escrever_secao(ws, df_bot,
                            "BOTTOM AMBASSADORS  (cap ceil/2)  -  " + str(n_bot_emb) + " embaixadores",
                            vm_t, vm_h, vm_c, vm_a, cols,
                            linha, borda, center, ft, fh, fd)
            _ajustar_largura(ws, [df_top, df_bot], cols)
            ws.freeze_panes = "A3"


            # MINI-TABELA DE TARGETS — coluna K (col 11) em diante
            # Mostra somente os processos presentes no ranking
            if (targets_prod or totais_processo) and len(df_enriquecido) > 0:
                from openpyxl.styles import PatternFill as PF
                processos_presentes = sorted(set(
                    str(v).strip() for v in df_enriquecido["Processo"].dropna()
                    if str(v).strip() and str(v).strip().lower() not in ("", "nan")
                ))
                col_k = 15   # Coluna O — buffer (M, N vazios) entre tabela e targets


                # Estilos da mini-tabela
                p_az_t = PF(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
                p_az_h = PF(start_color="1565C0", end_color="1565C0", fill_type="solid")
                p_az_d = PF(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                ft_wt  = Font(bold=True, color="FFFFFF", size=11)
                ft_wh  = Font(bold=True, color="FFFFFF", size=10)
                ft_dd  = Font(size=10)


                # Linha 1: Titulo
                t_cell = ws.cell(row=1, column=col_k, value="TARGETS POR PROCESSO")
                t_cell.fill = p_az_t; t_cell.font = ft_wt; t_cell.alignment = center
                ws.merge_cells(start_row=1, start_column=col_k,
                               end_row=1,   end_column=col_k + 2)


                # Linha 2: Cabeçalhos
                for j, titulo in enumerate(["Processo", "Target Prod (UPH)", "Target Qual (DPMO)"], col_k):
                    c = ws.cell(row=2, column=j, value=titulo)
                    c.fill = p_az_h; c.font = ft_wh
                    c.alignment = center; c.border = borda


                # Linhas 3+: Dados por processo
                for r_idx, proc in enumerate(processos_presentes, start=3):
                    # Target produtividade
                    prod_tgt = (targets_prod or {}).get(proc, "")


                    # Target qualidade: busca pela chave do totais_processo
                    qual_tgt = ""
                    if totais_processo:
                        chave = PROCESSO_TO_TOTAIS_KEY.get(proc, proc.upper())
                        val   = totais_processo.get(chave)
                        if val is None:
                            val = totais_processo.get(proc.upper())
                        if val is not None:
                            try:
                                qual_tgt = str(int(round(float(val))))
                            except (ValueError, TypeError):
                                qual_tgt = str(val)


                    for j, valor in enumerate([proc, prod_tgt, qual_tgt], col_k):
                        c = ws.cell(row=r_idx, column=j, value=valor)
                        c.fill = p_az_d; c.font = ft_dd
                        c.alignment = center; c.border = borda


                # Ajusta largura das 3 colunas de targets
                for j, larg in zip(range(col_k, col_k + 3), [18, 20, 22]):
                    letra = chr(64 + j) if j <= 26 else chr(64 + j // 26) + chr(65 + (j - 1) % 26)
                    ws.column_dimensions[letra].width = larg
                print("      Targets adicionados na coluna K: "
                      + str(len(processos_presentes)) + " processos")
            wb.save(tentativa)
            print("\n   ARQUIVO 3 salvo: " + os.path.abspath(tentativa))
            print("      TOP:    " + str(n_top_emb) + " embaixadores | " + str(len(df_top)) + " associados")
            print("      BOTTOM: " + str(n_bot_emb) + " embaixadores | " + str(len(df_bot)) + " associados")
            return
        except PermissionError:
            print("   '" + tentativa + "' aberto! Tentando outro nome...")
    print("   Feche o arquivo ranking_final.xlsx e rode novamente.")


# ============================================
# ANIMACAO DE CARREGAMENTO (spinner)
# Mostra uma animacao no terminal enquanto
# uma operacao demorada esta em andamento.
# Uso: spinner = iniciar_spinner("Mensagem")
#      ... operacao demorada ...
#      parar_spinner(spinner)
# ============================================
def iniciar_spinner(mensagem="Carregando"):
    stop_event = threading.Event()
    def _animar():
        frames = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        i = 0
        while not stop_event.is_set():
            frame = frames[i % len(frames)]
            print("\r  " + C.ORANGE + frame + C.RESET + " "
                  + C.SKY + mensagem + C.RESET
                  + C.GRAY + " ..." + C.RESET, end="", flush=True)
            i += 1
            time.sleep(0.08)
        # Limpa a linha do spinner
        print("\r" + " " * (len(mensagem) + 14) + "\r", end="", flush=True)
    t = threading.Thread(target=_animar, daemon=True)
    t.start()
    return stop_event


def parar_spinner(stop_event, concluido=None):
    stop_event.set()
    time.sleep(0.12)
    if concluido:
        ok(concluido)


# ============================================
# EMPLOYEE ROSTER (Site of Alefcas) — URL e XPath
# ============================================
URL_ROSTER = "local:roster.html"
URL_ROSTER_PARAMS = (
    "?reportFormat=HTML&warehouseId=FClocale"
    "&employeeStatusActive=on&_employeeStatusActive=on"
    "&employeeStatusLeaveOfAbsence=on&_employeeStatusLeaveOfAbsence=on"
    "&employeeStatusExempt=on&_employeeStatusExempt=on"
    "&employeeTypeAmzn=on&_employeeTypeAmzn=on"
    "&employeeTypeTemp=on&_employeeTypeTemp=on"
    "&employeeType3Pty=on&_employeeType3Pty=on"
    "&Employee+ID=Employee+ID&User+ID=User+ID"
    "&Employee+Name=Employee+Name"
    "&Badge+Barcode+ID=Badge+Barcode+ID"
    "&Department+ID=Department+ID"
    "&Employment+Start+Date=Employment+Start+Date"
    "&Employment+Type=Employment+Type"
    "&Employee+Status=Employee+Status"
    "&Manager+Name=Manager+Name"
    "&Temp+Agency+Code=Temp+Agency+Code"
    "&Job+Title=Job+Title"
    "&Management+Area+ID=Management+Area+ID"
    "&Shift+Pattern=Shift+Pattern"
    "&Badge+RFID=Badge+RFID"
    "&Exempt=Exempt&hideColumns=Photo&submit=true"
)
XPATH_ROSTER_CSV = "/html/body/div[3]/div[1]/span/form/table/tbody/tr[5]/td/a[2]"


# ============================================
# BAIXAR PLANILHA VIA NAVEGADOR
# Abre a URL no Firefox, espera o usuario baixar o arquivo,
# localiza o download, move para a pasta do projeto e renomeia.
# Usada para 'associados.xlsx' e 'dados_funcionarios.xlsx'.
# ============================================
def baixar_planilha_navegador(url, destino_nome, rotulo):
    titulo("Download: " + rotulo)
    resp = perguntar("   Baixar '" + destino_nome + "' agora? (ENTER = sim / N = pular): ").strip().lower()
    if resp == "n":
        dica("Download pulado. Usando o '" + destino_nome + "' que ja estiver na pasta.")
        return False

    driver = None
    try:
        driver = setup_firefox(usar_perfil=True)
        driver.get(url)
        handle_login(driver, url)
        esperar_pagina(driver, timeout=15)
        print()
        info(rotulo + " aberto no navegador.")
        dica("Faca o download do arquivo (Excel).")
        dica("O programa vai localizar o download automaticamente.")
        print()

        before = snapshot_downloads()
        perguntar("   ENTER apos baixar o arquivo... ")

        print("   Aguardando download...", end="", flush=True)
        downloaded = wait_new_file(before, timeout=DOWNLOAD_TIMEOUT)

        if not downloaded:
            print()
            manual = perguntar("   Cole o caminho do arquivo baixado (ou ENTER para pular): ").strip().strip('"')
            if manual and os.path.isfile(manual):
                downloaded = manual
            else:
                aviso("Download nao localizado. Usando arquivo existente, se houver.")
                return False
        print()
        ok("Arquivo baixado: " + os.path.basename(downloaded))

    except Exception as e:
        erro("Erro ao baixar " + rotulo + ": " + str(e))
        return False
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    # Move e renomeia para a pasta do projeto (4. TOP AMBASSADORS)
    destino = SCRIPT_DIR / destino_nome
    try:
        # Garante que o destino nao esteja aberto/travado
        if destino.exists():
            garantir_arquivo_fechado(destino, destino_nome)
            try:
                destino.unlink()
            except Exception:
                pass
        shutil.move(str(downloaded), str(destino))
        ok("Salvo como: " + destino_nome + "  (pasta do projeto)")
        return True
    except Exception as e:
        erro("Nao foi possivel mover o arquivo: " + str(e))
        dica("Mova manualmente para a pasta do projeto e renomeie para '" + destino_nome + "'.")
        return False


# ============================================
# BAIXAR EMPLOYEE ROSTER (Site of Alefcas)
# Abre Firefox, acessa o site do roster, clica no link CSV
# e retorna (roster_dict, ids_dict)
#   roster_dict = {login.lower(): nome_completo}
#   ids_dict    = {login.lower(): employee_id}
# ============================================
def _procurar_roster_local():
    """Procura um CSV de roster ja baixado em locais comuns.
    Retorna o caminho do mais recente, ou None."""
    locais = [
        RECURSOS_DIR,
        Path(DOWNLOAD_DIR),
        SCRIPT_DIR,
    ]
    candidatos = []
    palavras = ["roster", "employee"]
    for pasta in locais:
        try:
            for f in pasta.glob("*.csv"):
                nome = f.name.lower()
                if any(p in nome for p in palavras):
                    candidatos.append(f)
        except Exception:
            continue
    if not candidatos:
        return None
    # Mais recente por data de modificacao
    return str(max(candidatos, key=lambda p: p.stat().st_mtime))


def baixar_roster():
    print()
    print("=" * 55)
    print("   EMPLOYEE ROSTER (Site of Alefcas) — Buscando nomes dos embaixadores")
    print("=" * 55)

    # 1. Procura um CSV de roster ja existente localmente
    roster_local = _procurar_roster_local()
    if roster_local:
        print("   CSV de roster encontrado localmente:")
        print("      " + os.path.basename(roster_local))
        usar = perguntar("   Usar este arquivo? (ENTER = sim / N = baixar novo): ").strip().lower()
        if usar != "n":
            return ler_roster_csv(roster_local, retornar_ids=True)

    driver = None
    try:
        driver = setup_firefox(usar_perfil=True)   # ← perfil default-esr (Midway ativo)

        # Navega direto para a pagina do Roster
        driver.get(URL_ROSTER)
        handle_login(driver, URL_ROSTER)
        esperar_pagina(driver, timeout=15)
        print()
        info("Employee Roster (Site of Alefcas) aberto no navegador.")
        dica("Selecione os filtros necessarios e clique em baixar o CSV.")
        dica("O programa vai detectar o download automaticamente.")
        print()

        before = snapshot_downloads()
        perguntar("   ENTER apos baixar o CSV do Employee Roster... ")


        print("   Aguardando download...", end="", flush=True)
        downloaded = wait_new_file(before, timeout=30)


        if not downloaded:
            print()
            downloaded_manual = perguntar(
                "   Cole o caminho do CSV (ou ENTER para pular): "
            ).strip().strip('"')
            if downloaded_manual and os.path.isfile(downloaded_manual):
                downloaded = downloaded_manual
            else:
                print("   Roster ignorado. Coluna Nome das abas Performer ficara vazia.")
                return {}, {}

        print()
        print("   Roster baixado: " + os.path.basename(downloaded))
        return ler_roster_csv(downloaded, retornar_ids=True)


    except Exception as e:
        print("\n   ERRO ao baixar Roster: " + str(e))
        return {}, {}
    finally:
        if driver:
            try:
                driver.quit()
                print("   Firefox fechado.")
            except Exception:
                pass


# ============================================
# LER CSV DO EMPLOYEE ROSTER
# Col A (idx 0) = Employee ID (numero)
# Col B (idx 1) = User ID (login)
# Col C (idx 2) = Employee Name (nome completo)
# Retorna dict {login.lower(): nome}
# Se retornar_ids=True, retorna (roster, ids_dict)
#   onde ids_dict = {login.lower(): employee_id}
# ============================================
def ler_roster_csv(caminho, retornar_ids=False):
    roster = {}
    ids_dict = {}
    try:
        df = None
        for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            try:
                df = pd.read_csv(caminho, dtype=str, encoding=enc)
                break
            except Exception:
                continue


        if df is None:
            print("   AVISO: Nao foi possivel ler o Roster CSV.")
            return (roster, ids_dict) if retornar_ids else roster


        if len(df.columns) < 3:
            print("   AVISO: CSV tem menos de 3 colunas.")
            return (roster, ids_dict) if retornar_ids else roster


        col_id    = df.columns[0]   # Col A = Employee ID
        col_login = df.columns[1]   # Col B = User ID
        col_nome  = df.columns[2]   # Col C = Employee Name
        print("   Colunas: ID='" + col_id + "' | Login='" + col_login + "' | Nome='" + col_nome + "'")


        skip = {"", "nan", "none", "user id", "user+id", "userid",
                "employee name", "employee+name", "employee id", "employee+id"}


        for _, row in df.iterrows():
            emp_id = str(row.iloc[0]).strip()
            login  = str(row.iloc[1]).strip()
            nome   = str(row.iloc[2]).strip()
            if login.lower() not in skip and nome.lower() not in skip:
                roster[login.lower()] = nome
                if emp_id and emp_id.lower() not in ("", "nan", "none"):
                    ids_dict[login.lower()] = emp_id


        print("   Roster carregado: " + str(len(roster)) + " funcionarios")
        if retornar_ids:
            print("   IDs mapeados: " + str(len(ids_dict)))


    except Exception as e:
        print("   ERRO ao ler Roster CSV: " + str(e))


    return (roster, ids_dict) if retornar_ids else roster


# ============================================
# PREENCHER NOMES NAS ABAS TOP/BOTTOM PERFORMER
# VLOOKUP: Col B (login) -> roster_dict -> Col C (nome)
# Sem formatacao — somente valor da celula
# ============================================
def formatar_nome_embaixador(nome_completo):
    """Reduz o nome completo do roster para 'Primeiro Sobrenome'.
    Formato do roster: 'SOBRENOME,Primeiro Nome ...' (virgula separa).
      - Antes da virgula: sobrenome (usa a 1a palavra)
      - Depois da virgula: primeiro nome (usa a 1a palavra)
    Ex: 'Gomes da Costa,Renata Gomes Da' -> 'Renata Gomes'
        'Lima,Marcio Lima Dos'           -> 'Marcio Lima'
        'Caiobrit,Caiobrit'              -> 'Caiobrit'  (iguais = 1x)
        'Veloso,Suelane SUELANE VELOSO'  -> 'Suelane Veloso'
        'Alef Castro' (sem virgula)      -> 'Alef Castro'
    """
    if not nome_completo:
        return ""
    s = str(nome_completo).strip()
    if "," in s:
        antes, depois = s.split(",", 1)
        palavras_antes  = antes.strip().split()
        palavras_depois = depois.strip().split()
        sobrenome = palavras_antes[0]  if palavras_antes  else ""
        primeiro  = palavras_depois[0] if palavras_depois else ""
        if primeiro and sobrenome and primeiro.lower() == sobrenome.lower():
            resultado = primeiro
        elif primeiro and sobrenome:
            resultado = primeiro + " " + sobrenome
        else:
            resultado = primeiro or sobrenome
    else:
        resultado = s
    return resultado.title()


def preencher_nomes_reconhecimento(historico_path, roster_dict):
    if not roster_dict:
        print("   Roster vazio. Coluna Nome das abas Performer nao preenchida.")
        return


    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(historico_path))


        abas_alvo = [a for a in ("TOP Performer", "BOTTOM Performer")
                     if a in wb.sheetnames]
        if not abas_alvo:
            print("   AVISO: Abas TOP/BOTTOM Performer nao encontradas.")
            wb.close()
            return


        n_ok  = 0
        n_nao = []


        for nome_aba in abas_alvo:
            ws = wb[nome_aba]
            # Garante cabecalho da Col C
            ws.cell(row=1, column=3, value="Nome Embaixador")
            # Dados a partir da linha 2: Col B = login, Col C = nome
            for row in ws.iter_rows(min_row=2):
                if len(row) < 3:
                    continue
                login = str(row[1].value).strip().lower() if row[1].value else ""
                if not login or login in ("", "nan", "none"):
                    continue
                nome = roster_dict.get(login)
                if nome:
                    row[2].value = formatar_nome_embaixador(nome)   # Col C = nome formatado
                    n_ok += 1
                else:
                    n_nao.append(login)


        wb.save(str(historico_path))
        wb.close()


        print("   Nomes preenchidos nas abas Performer: " + str(n_ok))
        if n_nao:
            print("   Nao encontrados (" + str(len(n_nao)) + "): "
                  + ", ".join(n_nao[:5]) + ("..." if len(n_nao) > 5 else ""))


    except Exception as e:
        print("   ERRO ao preencher nomes: " + str(e))


# ============================================
# CARTAS DE RECONHECIMENTO (imagem + PDF)
# Le a aba TOP Performer do HistoricoAmbassadors.xlsx,
# pega os embaixadores do ONB selecionado e desenha o
# nome, login e data sobre o template de imagem.
# Template: TemplateReconhecimento.png em 1. HISTORICO
# Saida: PDF unico em 1. HISTORICO\2. RECONHECIMENTO
# ============================================
# Posicoes (x, y) dos campos sobre o template — ajustaveis.
# x = None  -> centralizado horizontalmente
# Coordenadas baseadas no template enviado (proporcionais a largura/altura)
CERT_POS = {
    "nome":  {"x": None, "y_rel": 0.450, "tam_rel": 0.045, "cor": (255, 255, 255), "negrito": True},
    "login": {"x": None, "y_rel": 0.535, "tam_rel": 0.022, "cor": (230, 126, 0),  "negrito": True},
    "data":  {"x": None, "y_rel": 0.895, "tam_rel": 0.020, "cor": (90, 90, 90),    "negrito": False},
}


def _carregar_fonte_cert(tamanho, negrito=False):
    """Tenta carregar uma fonte TrueType; cai no default se nao achar."""
    from PIL import ImageFont
    candidatos = []
    if negrito:
        candidatos = ["arialbd.ttf", "Arial_Bold.ttf",
                      "C:\\Windows\\Fonts\\arialbd.ttf",
                      "DejaVuSans-Bold.ttf"]
    else:
        candidatos = ["arial.ttf", "Arial.ttf",
                      "C:\\Windows\\Fonts\\arial.ttf",
                      "DejaVuSans.ttf"]
    for c in candidatos:
        try:
            return ImageFont.truetype(c, tamanho)
        except Exception:
            continue
    try:
        return ImageFont.load_default()
    except Exception:
        return None


def _desenhar_centralizado(draw, texto, y, fonte, cor, largura_img, x_fixo=None):
    try:
        bbox = draw.textbbox((0, 0), texto, font=fonte)
        w = bbox[2] - bbox[0]
    except Exception:
        w = len(texto) * 10
    x = x_fixo if x_fixo is not None else (largura_img - w) // 2
    draw.text((x, y), texto, font=fonte, fill=cor)


def gerar_mala_direta_reconhecimento(data_onb_alvo=None):
    titulo("Cartas de Reconhecimento")

    historico_path = HISTORICO_DIR / PLANILHA_HISTORICO

    # Procura o template de imagem (png/jpg)
    modelo_path = None
    for nome_modelo in ["TemplateReconhecimento.png", "TemplateReconhecimento.jpg",
                        "TemplateReconhecimento.jpeg"]:
        for base in [HISTORICO_DIR, SCRIPT_DIR, RECURSOS_DIR]:
            cand = base / nome_modelo
            if cand.exists():
                modelo_path = cand
                break
        if modelo_path:
            break

    if not historico_path.exists():
        erro("HistoricoAmbassadors.xlsx nao encontrado em 1. HISTORICO.")
        return
    if not modelo_path:
        erro("TemplateReconhecimento.png nao encontrado.")
        dica("Coloque a imagem do template (PNG) em 1. HISTORICO.")
        return

    try:
        from openpyxl import load_workbook
    except Exception:
        erro("openpyxl nao disponivel.")
        return
    try:
        from PIL import Image, ImageDraw
    except Exception:
        info("Instalando Pillow...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow", "--quiet"],
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            from PIL import Image, ImageDraw
        except Exception as e:
            erro("Nao foi possivel instalar Pillow: " + str(e))
            return

    # 1. Le a aba TOP Performer (Data=A, Login=B, Nome=C)
    try:
        wb = load_workbook(str(historico_path), data_only=True)
        if "TOP Performer" not in wb.sheetnames:
            erro("Aba TOP Performer nao encontrada no historico.")
            wb.close()
            return
        ws = wb["TOP Performer"]
        registros = []   # (login, nome, data)
        # A coluna F (RECORRENCIA) e ignorada: toda linha TOP gera carta.
        for row in ws.iter_rows(min_row=2, values_only=True):
            data    = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            login   = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            nome    = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            if "," in nome:
                nome = formatar_nome_embaixador(nome)
            if login and login.lower() not in ("", "nan", "none", "login embaixador"):
                registros.append((login, nome, data))
        wb.close()
    except Exception as e:
        erro("Erro ao ler TOP Performer: " + str(e))
        return

    if not registros:
        aviso("Nenhum registro na aba TOP Performer.")
        return

    # 2. Define a data-alvo (ONB). Lista as datas com contagem de embaixadores
    from collections import OrderedDict
    contagem = OrderedDict()
    for _, _, d in registros:
        if d:
            contagem[d] = contagem.get(d, 0) + 1

    def _ordem_data(d):
        primeira = str(d).split("|")[0].strip()
        try:
            p = primeira.split(".")
            return (int(p[1]), int(p[0]))
        except Exception:
            return (99, 99)

    datas_disp = sorted(contagem.keys(), key=_ordem_data)
    if not datas_disp:
        aviso("Nenhuma data de ONB encontrada no historico.")
        return

    if data_onb_alvo is None:
        print()
        info("Escolha o dia de ONB para gerar as cartas:")
        for i, d in enumerate(datas_disp, 1):
            passo("[" + str(i) + "] ONB " + d + "  (" + str(contagem[d]) + " embaixadores)")
        print()
        escolha = perguntar("   Numero do ONB (ENTER = mais recente): ").strip()
        if escolha and escolha.isdigit() and 1 <= int(escolha) <= len(datas_disp):
            data_onb_alvo = datas_disp[int(escolha) - 1]
        else:
            data_onb_alvo = datas_disp[-1] if datas_disp else ""

    alvo = [(lg, nm, dt) for (lg, nm, dt) in registros if dt == data_onb_alvo]
    if not alvo:
        aviso("Nenhum embaixador encontrado para o ONB " + str(data_onb_alvo))
        return

    ok(str(len(alvo)) + " carta(s) a gerar para o ONB " + str(data_onb_alvo))

    # 3. Desenha cada carta sobre o template de imagem
    try:
        base_img = Image.open(str(modelo_path)).convert("RGB")
    except Exception as e:
        erro("Nao foi possivel abrir o template: " + str(e))
        return

    larg, alt = base_img.size
    paginas = []

    for (login, nome, data) in alvo:
        nome_final = nome if nome else login
        cert = base_img.copy()
        draw = ImageDraw.Draw(cert)

        # NOME
        cfg = CERT_POS["nome"]
        f = _carregar_fonte_cert(int(alt * cfg["tam_rel"]), cfg["negrito"])
        _desenhar_centralizado(draw, nome_final.upper(), int(alt * cfg["y_rel"]),
                               f, cfg["cor"], larg, cfg["x"])

        # LOGIN
        cfg = CERT_POS["login"]
        f = _carregar_fonte_cert(int(alt * cfg["tam_rel"]), cfg["negrito"])
        _desenhar_centralizado(draw, login, int(alt * cfg["y_rel"]),
                               f, cfg["cor"], larg, cfg["x"])

        # DATA
        cfg = CERT_POS["data"]
        f = _carregar_fonte_cert(int(alt * cfg["tam_rel"]), cfg["negrito"])
        _desenhar_centralizado(draw, str(data), int(alt * cfg["y_rel"]),
                               f, cfg["cor"], larg, cfg["x"])

        paginas.append(cert)

    # 4. Salva tudo como um unico PDF
    data_tag = str(data_onb_alvo).split("|")[0].strip().replace("/", ".").replace(" ", "")
    pasta_rec = HISTORICO_DIR / "2. RECONHECIMENTO"
    try:
        pasta_rec.mkdir(parents=True, exist_ok=True)
    except Exception:
        pasta_rec = HISTORICO_DIR
    saida = pasta_rec / ("Reconhecimento_ONB_" + data_tag + ".pdf")

    try:
        if paginas:
            paginas[0].save(str(saida), "PDF", resolution=150.0,
                            save_all=True, append_images=paginas[1:])
            ok("Cartas geradas: " + saida.name)
            ok(str(len(paginas)) + " carta(s) criada(s)")
            try:
                os.startfile(str(saida))
            except Exception:
                pass
    except Exception as e:
        erro("Erro ao salvar PDF: " + str(e))


# ============================================
# SALVAR / CARREGAR TOTAIS_PROCESSO (targets qualidade)
# Persistido em JSON para reuso ao pular o ATLAS
# ============================================
def salvar_totais_processo(totais_processo, start_date="", end_date=""):
    try:
        caminho = RECURSOS_DIR / "totais_processo.json"
        dados = {
            "start_date": start_date,
            "end_date": end_date,
            "totais": totais_processo,
        }
        with open(str(caminho), "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        print("   Targets de qualidade salvos: " + str(caminho))
    except Exception as e:
        print("   AVISO: Nao foi possivel salvar targets: " + str(e))


def carregar_totais_processo():
    try:
        caminho = RECURSOS_DIR / "totais_processo.json"
        if not caminho.exists():
            return {}
        with open(str(caminho), "r", encoding="utf-8") as f:
            dados = json.load(f)
        return dados.get("totais", {})
    except Exception as e:
        print("   AVISO: Nao foi possivel carregar targets: " + str(e))
        return {}


# ============================================
# FLUXO PRINCIPAL COMPLETO
# etapa_inicial:
#   1 = completo (baixa tudo do ATLAS)
#   2 = pula ATLAS (usa ExtracaoATLAS.xlsx ja baixado)
# ============================================
def rodar_completo(df_associacoes, abas_nomes=None, etapa_inicial=1,
                   roster_dict=None, ids_dict=None):


    # Fonte de verdade imutavel: APENAS embaixadores e associados de associados.xlsx
    embaixadores_validos = set(df_associacoes["Embaixador"].str.strip().unique())
    associados_validos   = set(df_associacoes["Associado"].str.strip().unique())
    print("   Fonte de verdade carregada de associados.xlsx:")
    print("      " + str(len(embaixadores_validos)) + " embaixadores validos")
    print("      " + str(len(associados_validos)) + " associados validos")


    roster_dict = roster_dict or {}
    ids_dict    = ids_dict or {}


    # ETAPA 1: DATAS
    titulo("1/8  Configurando datas")
    if etapa_inicial >= 2:
        start_date, end_date = "", ""
        passo("(Pulado - usando dados ja baixados)")
    else:
        start_date, end_date = calcular_datas_atlas(abas_nomes if abas_nomes else [])


    # ETAPA 2: QUALIDADE (ATLAS)
    titulo("2/8  Qualidade (ATLAS)")
    totais_processo  = {}
    df_qual_completo = pd.DataFrame()
    driver           = None
    all_downloaded   = []


    if etapa_inicial >= 2:
        # PULA DOWNLOAD — reaproveita ExtracaoATLAS.xlsx e targets salvos
        print("   Pulando download do ATLAS. Usando arquivos existentes...")
        totais_processo = carregar_totais_processo()
        print("   Targets carregados: " + str(len(totais_processo)) + " processos")
        atlas_path = str(RECURSOS_DIR / PLANILHA_ATLAS)
        if os.path.exists(atlas_path):
            df_atlas = ler_extracao_atlas(atlas_path)
            if len(df_atlas) > 0:
                df_qual_completo, _ = cruzar_atlas_associados(
                    df_atlas, df_associacoes, totais_processo=totais_processo)
                df_qual_completo    = filtrar_por_opportunities(df_qual_completo)
                if len(df_qual_completo) > 0:
                    antes_q          = df_qual_completo["Embaixador"].nunique()
                    df_qual_completo = df_qual_completo[
                        df_qual_completo["Embaixador"].isin(embaixadores_validos)
                    ].reset_index(drop=True)
                    depois_q = df_qual_completo["Embaixador"].nunique()
                    if antes_q > depois_q:
                        print("   FILTRO SEGURANCA [Qualidade]: " + str(antes_q - depois_q)
                              + " embaixador(es) removido(s) - nao estao em associados.xlsx")
        else:
            print("   AVISO: " + atlas_path + " nao encontrado. Qualidade sera ignorada.")
        return _rodar_etapas_finais(
            df_associacoes, abas_nomes, embaixadores_validos,
            totais_processo, df_qual_completo,
            roster_dict=roster_dict, ids_dict=ids_dict)


    try:
        driver = setup_firefox()


        print("\n--- ETAPA 2.1: Totals Report (sem PACK) ---")
        totals_path = baixar_totals_report(driver, start_date, end_date, is_first_open=True)
        if totals_path:
            totais_processo = ler_totals_report(totals_path)
        else:
            print("   Totals Report ignorado. Quality Rate nao sera calculado.")


        print("\n--- ETAPA 2.2: Pack Single e Multi (Totals) ---")
        caminhos_pack = baixar_pack_subprocessos(driver, start_date, end_date)
        if caminhos_pack:
            pack_totals = ler_pack_totals(caminhos_pack)
            totais_processo.update(pack_totals)
            print("   Pack totals adicionados: " + str(list(pack_totals.keys())))
        else:
            print("   Pack subprocessos ignorados.")


        print("\n--- ETAPA 2.3: Raw Reports ---")
        all_downloaded = process_group(
            driver=driver,
            processes=COMBINED_PROCESSES,
            group_name="Combined",
            start_date=start_date,
            end_date=end_date,
            target_process="combined",
            is_first_group=False,
            counter_start=1,
            total=len(COMBINED_PROCESSES),
        )
    except KeyboardInterrupt:
        print("\nInterrompido pelo usuario.")
    except Exception as e:
        print("\nErro no Atlas: " + str(e))
    finally:
        if driver:
            try:
                driver.quit()
                print("\n   Firefox fechado.")
            except Exception:
                pass


    if all_downloaded:
        atlas_path = consolidate_to_excel(all_downloaded)
        if atlas_path and os.path.exists(atlas_path):
            df_atlas = ler_extracao_atlas(atlas_path)
            if len(df_atlas) > 0:
                df_qual_completo, _ = cruzar_atlas_associados(
                    df_atlas, df_associacoes, totais_processo=totais_processo)
                df_qual_completo    = filtrar_por_opportunities(df_qual_completo)


                # Filtro de seguranca: apenas embaixadores de associados.xlsx
                if len(df_qual_completo) > 0:
                    antes_q          = df_qual_completo["Embaixador"].nunique()
                    df_qual_completo = df_qual_completo[
                        df_qual_completo["Embaixador"].isin(embaixadores_validos)
                    ].reset_index(drop=True)
                    depois_q = df_qual_completo["Embaixador"].nunique()
                    if antes_q > depois_q:
                        print("   FILTRO SEGURANCA [Qualidade]: " + str(antes_q - depois_q)
                              + " embaixador(es) removido(s) - nao estao em associados.xlsx")
    else:
        print("\n   Nenhum dado do Atlas. Qualidade sera ignorada.")

    # Salva totais_processo (targets de qualidade) para reuso futuro
    salvar_totais_processo(totais_processo, start_date, end_date)
    return _rodar_etapas_finais(
        df_associacoes, abas_nomes, embaixadores_validos,
        totais_processo, df_qual_completo,
        roster_dict=roster_dict, ids_dict=ids_dict)


# ============================================
# ETAPAS FINAIS (3 a 8 + historico + email)
# Separado para permitir reuso ao pular o ATLAS
# ============================================
def _rodar_etapas_finais(df_associacoes, abas_nomes, embaixadores_validos,
                         totais_processo, df_qual_completo,
                         roster_dict=None, ids_dict=None):

    roster_dict = roster_dict or {}
    ids_dict    = ids_dict or {}

    # ETAPA 3: PRODUTIVIDADE
    titulo("3/8  Produtividade (dados_funcionarios)")
    df_prod_completo = pd.DataFrame()

    if os.path.exists(PLANILHA_DADOS):
        garantir_arquivo_fechado(PLANILHA_DADOS, "dados_funcionarios.xlsx")
        df_prod_raw, _ = processar_todas_abas_dados(PLANILHA_DADOS, df_associacoes)
        if len(df_prod_raw) > 0:
                mask_hifen       = df_prod_raw.apply(
                    lambda r: r.astype(str).str.strip().eq('-').any(), axis=1)
                df_prod_completo = filtro_horas_minimas(
                    df_prod_raw[~mask_hifen].reset_index(drop=True))


                # Filtro de seguranca: apenas embaixadores de associados.xlsx
                if len(df_prod_completo) > 0:
                    antes_p          = df_prod_completo["Embaixador"].nunique()
                    df_prod_completo = df_prod_completo[
                        df_prod_completo["Embaixador"].isin(embaixadores_validos)
                    ].reset_index(drop=True)
                    depois_p = df_prod_completo["Embaixador"].nunique()
                    if antes_p > depois_p:
                        print("   FILTRO SEGURANCA [Produtividade]: " + str(antes_p - depois_p)
                              + " embaixador(es) removido(s) - nao estao em associados.xlsx")
    else:
        print("   '" + PLANILHA_DADOS + "' nao encontrado. Produtividade ignorada.")


    # ETAPA 4: SEPARAR RC SORT
    titulo("4/8  Separando RC SORT")
    df_qual_normal = pd.DataFrame()
    df_prod_normal = pd.DataFrame()
    df_rc_cret     = pd.DataFrame()


    if len(df_qual_completo) > 0:
        mask_rc_q      = df_qual_completo["Processo"].apply(eh_rc_cret)
        df_qual_normal = df_qual_completo[~mask_rc_q].reset_index(drop=True)
        print("   Qualidade normal: " + str(len(df_qual_normal))
              + " | RC SORT excluidos: " + str(mask_rc_q.sum()))


    if len(df_prod_completo) > 0:
        mask_rc_p      = df_prod_completo["Processo"].apply(eh_rc_cret)
        df_prod_normal = df_prod_completo[~mask_rc_p].reset_index(drop=True)
        df_rc_cret     = df_prod_completo[mask_rc_p].reset_index(drop=True)
        print("   Produtividade normal: " + str(len(df_prod_normal))
              + " | RC SORT: " + str(len(df_rc_cret)))


    # ETAPA 5: FILTRO MIN-MAX ASSOCIADOS
    titulo("5/8  Filtro " + str(MIN_ASSOC_POR_EMBAIXADOR) + "-"
           + str(MAX_ASSOC_POR_EMBAIXADOR) + " associados por embaixador")
    if len(df_qual_normal) > 0:
        passo("[Qualidade]")
        df_qual_normal = filtro_embaixadores_min_max(df_qual_normal)
    if len(df_prod_normal) > 0:
        passo("[Produtividade]")
        df_prod_normal = filtro_embaixadores_min_max(df_prod_normal)
    if len(df_rc_cret) > 0:
        passo("[RC SORT]")
        df_rc_cret = filtro_embaixadores_min_max(df_rc_cret)


    # ETAPA 6: CLASSIFICAR TOP/BOTTOM
    titulo("6/8  Classificando TOP / BOTTOM")
    emb_qual_top = pd.DataFrame(columns=["Embaixador", "DPMO Embaixador", "Percentual Embaixador"])
    emb_qual_bot = pd.DataFrame(columns=["Embaixador", "DPMO Embaixador", "Percentual Embaixador"])
    if len(df_qual_normal) > 0:
        emb_qual_top, emb_qual_bot = classificar_qualidade_embaixadores(df_qual_normal, totais_processo)


    emb_prod_top = pd.DataFrame(columns=["Embaixador", "Rate LC Medio (%)"])
    emb_prod_bot = pd.DataFrame(columns=["Embaixador", "Rate LC Medio (%)"])
    if len(df_prod_normal) > 0:
        emb_prod_top, emb_prod_bot = classificar_produtividade_embaixadores(df_prod_normal)


    emb_rc_top = pd.DataFrame(columns=["Embaixador", "Rate LC Medio (%)"])
    emb_rc_bot = pd.DataFrame(columns=["Embaixador", "Rate LC Medio (%)"])
    if len(df_rc_cret) > 0:
        emb_rc_top, emb_rc_bot = classificar_produtividade_embaixadores(df_rc_cret)


    def filtrar_por_emb(df, emb_set, col="Embaixador"):
        if len(df) == 0 or len(emb_set) == 0:
            return pd.DataFrame(columns=df.columns if len(df) > 0 else [])
        return df[df[col].isin(emb_set)].reset_index(drop=True)


    embs_qt = set(emb_qual_top["Embaixador"]) if len(emb_qual_top) > 0 else set()
    embs_qb = set(emb_qual_bot["Embaixador"]) if len(emb_qual_bot) > 0 else set()
    embs_pt = set(emb_prod_top["Embaixador"]) if len(emb_prod_top) > 0 else set()
    embs_pb = set(emb_prod_bot["Embaixador"]) if len(emb_prod_bot) > 0 else set()
    embs_rt = set(emb_rc_top["Embaixador"])   if len(emb_rc_top)   > 0 else set()
    embs_rb = set(emb_rc_bot["Embaixador"])   if len(emb_rc_bot)   > 0 else set()


    # INVALIDO: UNICO caso = TOP em produtividade + BOTTOM em qualidade
    #   -> sem reconhecimento (rate alto nao compensa qualidade ruim)
    # Qualquer outro caso segue o fluxo normal:
    #   - BOTTOM qual + BOTTOM prod  -> BOTTOM
    #   - MEIO qual   + BOTTOM prod  -> BOTTOM
    #   - TOP qual    + BOTTOM prod  -> VALIDO (qualidade vence)
    # Garantia explicita: quem e BOTTOM de produtividade NUNCA e INVALIDO.
    embs_invalidos = (embs_pt & embs_qb) - embs_pb
    if embs_invalidos:
        print("\n   !!! INVALIDOS detectados: TOP prod + BOTTOM qual !!!")
        for e in sorted(embs_invalidos):
            print("      -> " + e + "  (SEM reconhecimento: removido do TOP prod e do BOTTOM qual)")
        # INVALIDO nao recebe reconhecimento: sai do TOP de produtividade
        # E TAMBEM do BOTTOM de qualidade, para nao aparecer no email/ranking.
        # (No resultado_completo.xlsx ele continua documentado como INVALIDO.)
        embs_pt      = embs_pt - embs_invalidos
        emb_prod_top = emb_prod_top[
            ~emb_prod_top["Embaixador"].isin(embs_invalidos)
        ].reset_index(drop=True)
        embs_qb      = embs_qb - embs_invalidos
        emb_qual_bot = emb_qual_bot[
            ~emb_qual_bot["Embaixador"].isin(embs_invalidos)
        ].reset_index(drop=True)
    else:
        embs_invalidos = set()


    df_qual_top = filtrar_por_emb(df_qual_normal, embs_qt)
    df_qual_bot = filtrar_por_emb(df_qual_normal, embs_qb)
    df_prod_top = filtrar_por_emb(df_prod_normal, embs_pt)
    df_prod_bot = filtrar_por_emb(df_prod_normal, embs_pb)
    df_rc_top   = filtrar_por_emb(df_rc_cret,     embs_rt)
    df_rc_bot   = filtrar_por_emb(df_rc_cret,     embs_rb)


    # Mescla metricas do embaixador (DPMO + Quality Rate) nas linhas dos associados
    cols_metrica = ["Embaixador", "DPMO Embaixador", "Percentual Embaixador"]
    if len(emb_qual_top) > 0 and len(df_qual_top) > 0:
        if "Percentual Embaixador" in df_qual_top.columns:
            df_qual_top = df_qual_top.drop(columns=["Percentual Embaixador"])
        df_qual_top = df_qual_top.merge(emb_qual_top[cols_metrica], on="Embaixador", how="left")
    if len(emb_qual_bot) > 0 and len(df_qual_bot) > 0:
        if "Percentual Embaixador" in df_qual_bot.columns:
            df_qual_bot = df_qual_bot.drop(columns=["Percentual Embaixador"])
        df_qual_bot = df_qual_bot.merge(emb_qual_bot[cols_metrica], on="Embaixador", how="left")
    # Mescla Rate LC Medio (%) nas linhas dos associados de produtividade
    # Usado em enriquecer_ranking() para preencher col J (Rate Emb) de linhas Produtividade
    cols_prod_metrica = ["Embaixador", "Rate LC Medio (%)"]
    if len(emb_prod_top) > 0 and len(df_prod_top) > 0:
        df_prod_top = df_prod_top.merge(emb_prod_top[cols_prod_metrica], on="Embaixador", how="left")
    if len(emb_prod_bot) > 0 and len(df_prod_bot) > 0:
        df_prod_bot = df_prod_bot.merge(emb_prod_bot[cols_prod_metrica], on="Embaixador", how="left")


    # ETAPA 7: SELECAO 50/50
    titulo("7/8  Selecao 50/50 para Ranking Final")
    ranking = selecionar_50_50(emb_prod_top, emb_qual_top, emb_prod_bot, emb_qual_bot)


    # ETAPA 8: SALVAR 3 ARQUIVOS
    titulo("8/8  Salvando arquivos")


    # Constroi string de datas para coluna A do resultado_completo
    data_str_res = construir_data_str(abas_nomes if abas_nomes else [])
    print("   Periodo analisado: " + data_str_res)


    # Adiciona coluna Classificacao em df_qual_normal e df_prod_normal
    df_qual_all = pd.DataFrame()
    if len(df_qual_normal) > 0:
        df_qual_all = df_qual_normal.copy()
        # INVALIDO na aba Qualidade = BOTTOM (qualidade genuinamente ruim)
        # INVALIDO na aba Produtividade = INVALIDO (TOP prod invalidado pela qual ruim)
        df_qual_all["Classificacao"] = df_qual_all["Embaixador"].apply(
            lambda e: "BOTTOM" if e in embs_invalidos
                      else ("TOP" if e in embs_qt else ("BOTTOM" if e in embs_qb else "MEIO")))


    df_prod_all = pd.DataFrame()
    if len(df_prod_normal) > 0:
        df_prod_all = df_prod_normal.copy()
        df_prod_all["Classificacao"] = df_prod_all["Embaixador"].apply(
            lambda e: "INVALIDO" if e in embs_invalidos
                      else ("TOP" if e in embs_pt else ("BOTTOM" if e in embs_pb else "MEIO")))


    print("\n   [Arquivo 1] resultado_completo.xlsx")
    caminho_resultado = str(RECURSOS_DIR / PLANILHA_SAIDA_COMPLETO)
    salvar_resultado_completo(
        df_qual_all, df_prod_all,
        caminho_resultado,
        data_str=data_str_res,
    )


    print("\n   [Arquivo 2] rc_cret.xlsx")
    if len(df_rc_top) > 0 or len(df_rc_bot) > 0:
        salvar_rc_cret(df_rc_top, df_rc_bot, str(RECURSOS_DIR / PLANILHA_SAIDA_RC_CRET))
    else:
        print("   Nenhum dado de RC SORT encontrado.")


    print("\n   [Arquivo 3] ranking_final.xlsx")
    data_backup  = extrair_data_backup(abas_nomes if abas_nomes else [])
    targets_prod = extrair_targets_prod(df_prod_completo)
    print("   Data do backup: " + data_backup
          + (" (da aba mais antiga)" if abas_nomes else " (fallback: hoje)"))

    # Mapa DPMO por embaixador (de TODOS com dados de qualidade) — fallback
    # para mostrar DPMO de embaixadores que entraram no TOP por Produtividade
    dpmo_map = {}
    if len(df_qual_normal) > 0:
        for emb, grupo in df_qual_normal.groupby("Embaixador"):
            try:
                soma_def = grupo["Total Defects"].apply(
                    lambda x: float(str(x).replace(",", ".").strip())
                    if str(x).strip() not in ("", "nan", "-") else 0.0).sum()
                soma_opp = grupo["Opportunities"].apply(
                    lambda x: float(str(x).replace(",", ".").strip())
                    if str(x).strip() not in ("", "nan", "-") else 0.0).sum()
                if soma_opp > 0:
                    dpmo_map[str(emb).strip().lower()] = int(round((soma_def / soma_opp) * 1000000))
            except Exception:
                continue

    # Mapa DPMO por ASSOCIADO (login -> DPMO do associado, inteiro)
    # Usado para mostrar o DPMO de cada associado nas linhas de Produtividade
    dpmo_assoc_map = {}
    if len(df_qual_normal) > 0 and "Associado" in df_qual_normal.columns:
        for assoc, grupo in df_qual_normal.groupby("Associado"):
            try:
                soma_def = grupo["Total Defects"].apply(
                    lambda x: float(str(x).replace(",", ".").strip())
                    if str(x).strip() not in ("", "nan", "-") else 0.0).sum()
                soma_opp = grupo["Opportunities"].apply(
                    lambda x: float(str(x).replace(",", ".").strip())
                    if str(x).strip() not in ("", "nan", "-") else 0.0).sum()
                if soma_opp > 0:
                    dpmo_assoc_map[str(assoc).strip().lower()] = int(round((soma_def / soma_opp) * 1000000))
            except Exception:
                continue

    salvar_ranking_final(
        ranking,
        str(RECURSOS_DIR / PLANILHA_SAIDA_FINAL),
        df_prod_top=df_prod_top,
        df_qual_top=df_qual_top,
        df_prod_bot=df_prod_bot,
        df_qual_bot=df_qual_bot,
        df_associacoes=df_associacoes,
        df_prod_all=df_prod_completo,
        data_backup=data_backup,
        targets_prod=targets_prod,
        totais_processo=totais_processo,
        dpmo_map=dpmo_map,
        dpmo_assoc_map=dpmo_assoc_map,
    )


    # Abre resultado_completo para revisao e pergunta sobre historico
    print("\n" + "=" * 55)
    print("   Abrindo resultado_completo.xlsx para revisao...")
    dica("Ao responder, a planilha sera fechada automaticamente.")
    try:
        os.startfile(caminho_resultado)
    except Exception:
        pass

    # Verifica se houve algum TOP no ranking final
    tops_ranking = [x for x in ranking if x.get("Classificacao") == "TOP"]
    bottoms_ranking = [x for x in ranking if x.get("Classificacao") == "BOTTOM"]

    # ============================================
    # CASO 1: NENHUM TOP encontrado
    # Pergunta se quer salvar no historico, faz a tarefa
    # independente da resposta e FINALIZA (sem email/mala direta).
    # ============================================
    if not tops_ranking:
        print()
        aviso("Nenhum embaixador TOP encontrado neste ciclo.")
        perguntar("   Salvar dados no historico mesmo assim? (ENTER para continuar): ")
        fechar_excel_arquivo(caminho_resultado)
        copiar_para_historico(caminho_resultado)
        if roster_dict:
            preencher_nomes_reconhecimento(
                HISTORICO_DIR / PLANILHA_HISTORICO, roster_dict)
        # Se tambem nao houve BOTTOM, documenta a data do ONB
        # numa aba propria para deixar registrado que o dia foi analisado.
        if not bottoms_ranking:
            registrar_onb_sem_destaque(data_str_res)
        print()
        print(C.GREEN + C.BOLD + "  ╔══════════════════════════════════════════════╗" + C.RESET)
        print(C.GREEN + C.BOLD + "  ║              PROCESSO FINALIZADO              ║" + C.RESET)
        print(C.GREEN + C.BOLD + "  ╚══════════════════════════════════════════════╝" + C.RESET)
        ok("Dados salvos no historico.")
        dica("Sem TOPs neste ciclo - email e mala direta nao sao necessarios.")
        return
    print()
    print("   Verifique os dados do resultado_completo.xlsx.")
    print("   Quando estiver pronto, responda abaixo:")
    print()
    resposta = perguntar("   Adicionar estes dados ao historico? (ENTER = sim / N = nao): ").strip().lower()
    fechar_excel_arquivo(caminho_resultado)
    if resposta != "n":
        copiar_para_historico(caminho_resultado)
        # Preenche Coluna Nome das abas Performer com nomes via Employee Roster
        # (usa o roster ja baixado no inicio; baixa agora so se nao tiver)
        if not roster_dict:
            print("\n   Buscando nomes dos embaixadores no Employee Roster...")
            roster_dict, ids_dict = baixar_roster()
        if roster_dict:
            preencher_nomes_reconhecimento(
                HISTORICO_DIR / PLANILHA_HISTORICO,
                roster_dict
            )
    else:
        print("   Copia para historico ignorada.")


    # RESUMO FINAL
    tops    = [x for x in ranking if x["Classificacao"] == "TOP"]
    bottoms = [x for x in ranking if x["Classificacao"] == "BOTTOM"]

    print()
    print(C.GREEN + C.BOLD + "  ╔══════════════════════════════════════════════╗" + C.RESET)
    print(C.GREEN + C.BOLD + "  ║              PROCESSO FINALIZADO              ║" + C.RESET)
    print(C.GREEN + C.BOLD + "  ╚══════════════════════════════════════════════╝" + C.RESET)
    print()
    print(C.GOLD + "  RANKING FINAL" + C.RESET)
    ok("TOP    : " + str(len(tops)) + " embaixadores")
    ok("BOTTOM : " + str(len(bottoms)) + " embaixadores")
    print()
    print(C.GRAY + "  Detalhe (resultado_completo.xlsx):" + C.RESET)
    passo("Qualidade   TOP " + str(len(embs_qt)) + "  ·  BOTTOM " + str(len(embs_qb)))
    passo("Produtiv.   TOP " + str(len(embs_pt)) + "  ·  BOTTOM " + str(len(embs_pb)))
    passo("RC SORT     TOP " + str(len(embs_rt)) + "  ·  BOTTOM " + str(len(embs_rb)))


    for arq in [PLANILHA_SAIDA_COMPLETO, PLANILHA_SAIDA_FINAL]:
        try:
            os.startfile(str(RECURSOS_DIR / arq))
        except Exception:
            pass


    # Preview do email — chama email_ambassador.py
    print("\n" + "=" * 55)
    dica("Ao responder, as planilhas abertas (ranking final) serao fechadas.")
    resposta_email = perguntar("   Deseja gerar o email (preview no terminal)? (ENTER = sim / N = nao): ").strip().lower()
    # Fecha as planilhas abertas para revisao (qualquer resposta)
    fechar_excel_arquivo(str(RECURSOS_DIR / PLANILHA_SAIDA_COMPLETO))
    fechar_excel_arquivo(str(RECURSOS_DIR / PLANILHA_SAIDA_FINAL))
    if resposta_email != "n":
        try:
            import importlib.util
            # Busca email_ambassador.py na pasta 2. RECURSOS
            email_script = RECURSOS_DIR / "email_ambassador.py"
            if not email_script.exists():
                # Fallback: mesma pasta do script principal
                email_script = SCRIPT_DIR / "email_ambassador.py"
            if email_script.exists():
                spec   = importlib.util.spec_from_file_location("email_ambassador", str(email_script))
                modulo = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(modulo)
                modulo.perguntar_e_enviar(
                    ranking_path=str(RECURSOS_DIR / PLANILHA_SAIDA_FINAL),
                    data_str=data_str_res,
                )
            else:
                print("   email_ambassador.py nao encontrado em '2. RECURSOS' nem na pasta do script.")
        except Exception as e_email:
            print("   AVISO email: " + str(e_email))
    else:
        print("   Envio de email ignorado.")


    # Mala direta de reconhecimento (apos o email)
    print("\n" + "=" * 55)
    resp_mala = perguntar("   Deseja gerar a mala direta de reconhecimento? (ENTER = sim / N = nao): ").strip().lower()
    if resp_mala != "n":
        gerar_mala_direta_reconhecimento()
    else:
        print("   Mala direta ignorada.")


# ============================================
# MAIN
# ============================================
def main(modo_forcado=None):
    print()
    print(C.ORANGE + C.BOLD + "  ╔══════════════════════════════════════════════╗" + C.RESET)
    print(C.ORANGE + C.BOLD + "  ║       TOP AMBASSADORS  ·  FC locale           ║" + C.RESET)
    print(C.ORANGE + C.BOLD + "  ╚══════════════════════════════════════════════╝" + C.RESET)


    # Cria pastas Recursos/ e Resultados/ se nao existirem
    criar_estrutura_pastas()


    # Carrega config.ini (sobrescreve os valores padrao)
    carregar_config()

    # Modo mala direta: nao precisa de associados.xlsx — vai direto
    if modo_forcado == "mala_direta":
        gerar_mala_direta_reconhecimento()
        return
    print()
    print(C.GRAY + "  Regras: " + str(MIN_ASSOC_POR_EMBAIXADOR) + "-" + str(MAX_ASSOC_POR_EMBAIXADOR)
          + " assoc/emb  ·  Qual TOP \u2264" + str(QUALITY_TOP_MAX) + "%  ·  Prod TOP \u2265"
          + str(PROD_TOP_MINIMO) + "%" + C.RESET)


    # ============================================
    # MODO
    # Sem menu inicial: o programa sempre inicia o PROCESSO COMPLETO
    # desde o comeco. Os modos alternativos (pular_atlas / email /
    # mala_direta) ainda podem ser acionados via --modo, mas nao ha
    # mais a tela de escolha ao abrir.
    # ============================================
    if modo_forcado in ("completo", "pular_atlas", "email", "mala_direta"):
        opcao = {"completo": "1", "pular_atlas": "2",
                 "email": "3", "mala_direta": "4"}[modo_forcado]
    else:
        opcao = "1"   # processo completo, direto

    # Modos que nao precisam de associados / preparacao
    if opcao == "4":
        gerar_mala_direta_reconhecimento()
        return
    elif opcao == "3":
        rodar_somente_email(None)
        return

    # ============================================
    # ETAPA INICIAL - EMPLOYEE ROSTER (Site of Alefcas)
    # 1o: pergunta se quer baixar o roster e direciona o usuario
    # ============================================
    roster_dict = {}
    ids_dict    = {}
    titulo("Employee Roster (Site of Alefcas)")
    resp_roster = perguntar("   Baixar o Employee Roster agora? (ENTER = sim / N = nao): ").strip().lower()
    if resp_roster != "n":
        roster_dict, ids_dict = baixar_roster()
    else:
        dica("Roster sera solicitado mais tarde se necessario.")


    # ============================================
    # 2o: download da planilha 'associados.xlsx' (opcional)
    # ============================================
    baixar_planilha_navegador(URL_ASSOCIADOS, "associados.xlsx", "Planilha de Associados (SharePoint)")


    # ============================================
    # 3o: prepara o associados.xlsx e da ENTER para seguir
    # ============================================
    titulo("Prepare o associados.xlsx")
    print(C.SKY + "  Antes de continuar, verifique no arquivo associados.xlsx:" + C.RESET)
    print(C.GRAY + "    1." + C.RESET + " O download dos associados foi feito e esta atualizado")
    print(C.GRAY + "    2." + C.RESET + " As abas que NAO serao usadas estao OCULTAS")
    print(C.GRAY + "    3." + C.RESET + " As datas (nomes das abas) estao corretas")
    print(C.GRAY + "    4." + C.RESET + " O arquivo esta SALVO e FECHADO")
    print()
    perguntar("   Tudo pronto? Pressione ENTER para seguir... ")


    # ============================================
    # 4o: download da planilha 'dados_funcionarios.xlsx' (opcional)
    # ============================================
    baixar_planilha_navegador(URL_DADOS, "dados_funcionarios.xlsx",
                              "Planilha de Produtividade (QuickSight)")


    titulo("Carregando associados")
    if not os.path.exists(PLANILHA_ASSOCIADOS):
        erro("Arquivo '" + str(PLANILHA_ASSOCIADOS) + "' nao encontrado.")
        passo("Baixe a planilha ou coloque-a na pasta do projeto.")
        return
    # Garante que o arquivo esteja fechado (Excel pode travar a leitura)
    garantir_arquivo_fechado(PLANILHA_ASSOCIADOS, "associados.xlsx")
    df_associacoes = carregar_associacoes(PLANILHA_ASSOCIADOS)


    if len(df_associacoes) == 0:
        erro("Nenhuma associacao encontrada!")
        return


    # Captura nomes das abas VISIVEIS para calculo automatico de datas
    # Abas ocultas sao ignoradas (controle interno, modelos, etc.)
    try:
        abas_nomes = obter_abas_visiveis(PLANILHA_ASSOCIADOS)
    except Exception:
        abas_nomes = []


    # ============================================
    # DISPATCH para o fluxo completo ou pular ATLAS
    # ============================================
    if opcao == "2":
        atlas_existe = os.path.exists(str(RECURSOS_DIR / PLANILHA_ATLAS))
        if not atlas_existe:
            print("\n   AVISO: ExtracaoATLAS.xlsx nao encontrado.")
            print("   Nao e possivel pular o ATLAS. Rodando completo...")
            rodar_completo(df_associacoes, abas_nomes, etapa_inicial=1,
                           roster_dict=roster_dict, ids_dict=ids_dict)
        else:
            rodar_completo(df_associacoes, abas_nomes, etapa_inicial=2,
                           roster_dict=roster_dict, ids_dict=ids_dict)
    else:
        rodar_completo(df_associacoes, abas_nomes, etapa_inicial=1,
                       roster_dict=roster_dict, ids_dict=ids_dict)


# ============================================
# RODAR SOMENTE O EMAIL
# Usa ranking_final.xlsx e targets ja salvos
# ============================================
def rodar_somente_email(abas_nomes=None):
    print("\n" + "=" * 55)
    print("   MODO: SOMENTE EMAIL")
    print("=" * 55)

    ranking_path = str(RECURSOS_DIR / PLANILHA_SAIDA_FINAL)
    if not os.path.exists(ranking_path):
        print("   ERRO: ranking_final.xlsx nao encontrado.")
        print("   Rode o fluxo completo primeiro.")
        return

    # ============================================
    # DATA DO ONB — pergunta ao usuario qual data usar
    # Sugere as datas das abas de associados.xlsx (se existir)
    # ============================================
    datas_sugeridas = []
    try:
        if os.path.exists(PLANILHA_ASSOCIADOS):
            abas = obter_abas_visiveis(PLANILHA_ASSOCIADOS)
            for nome in (abas or []):
                d, m = extrair_data_de_aba(nome)
                if d and m:
                    tag = str(d).zfill(2) + "." + str(m).zfill(2)
                    if tag not in datas_sugeridas:
                        datas_sugeridas.append(tag)
    except Exception:
        pass

    titulo("Data do ONB para o email")
    if datas_sugeridas:
        info("Datas encontradas no associados.xlsx:")
        for i, d in enumerate(datas_sugeridas, 1):
            passo("[" + str(i) + "] " + d)
        print()
        escolha = perguntar("   Numero da data, ou digite manualmente (ex: 07.05): ").strip()
        if escolha.isdigit() and 1 <= int(escolha) <= len(datas_sugeridas):
            data_str_res = datas_sugeridas[int(escolha) - 1]
        elif escolha:
            data_str_res = escolha
        else:
            data_str_res = datas_sugeridas[-1]
    else:
        data_str_res = perguntar("   Digite a data do ONB (ex: 07.05): ").strip()
        if not data_str_res:
            data_str_res = datetime.now().strftime("%d.%m")

    ok("ONB selecionado: " + data_str_res)

    # DEMO / PORTFOLIO: sem fotos e sem Outlook. Gera direto o preview.
    print("\n" + "=" * 55)
    resposta_email = perguntar("   Deseja gerar o email (preview no terminal)? (ENTER = sim / N = nao): ").strip().lower()
    if resposta_email != "n":
        try:
            import importlib.util
            email_script = RECURSOS_DIR / "email_ambassador.py"
            if not email_script.exists():
                email_script = SCRIPT_DIR / "email_ambassador.py"
            if email_script.exists():
                spec   = importlib.util.spec_from_file_location("email_ambassador", str(email_script))
                modulo = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(modulo)
                modulo.perguntar_e_enviar(
                    ranking_path=ranking_path,
                    data_str=data_str_res,
                )
            else:
                print("   email_ambassador.py nao encontrado.")
        except Exception as e_email:
            print("   AVISO email: " + str(e_email))
    else:
        print("   Email ignorado.")


if __name__ == "__main__":
    # Sem tela inicial: o programa inicia direto o PROCESSO COMPLETO.
    # Modo ainda pode ser passado explicitamente (compatibilidade):
    #   python AMBASSADORS.py --modo completo|pular_atlas|email|mala_direta
    modo = "completo"
    if "--modo" in sys.argv:
        try:
            modo = sys.argv[sys.argv.index("--modo") + 1]
        except IndexError:
            modo = "completo"

    try:
        main(modo_forcado=modo)
    except KeyboardInterrupt:
        print("\n   Interrompido pelo usuario.")
    except Exception as e:
        import traceback
        print("\n" + "=" * 55)
        print("   ERRO inesperado:")
        print("   " + str(e))
        traceback.print_exc()
        print("=" * 55)
    finally:
        try:
            input("\n   Pressione ENTER para fechar a janela... ")
        except Exception:
            pass

