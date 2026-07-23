# -*- coding: utf-8 -*-
# email_ambassador.py
# Módulo de preview de email Top/Bottom Ambassador
# Chamado automaticamente ao final do AMBASSADORS.py


import os
import base64
import webbrowser
import subprocess
import sys
import pandas as pd
from datetime import datetime
import time


# ============================================================
# CORES NO TERMINAL (ANSI) — consistente com AMBASSADORS.py
# ============================================================
if os.name == "nt":
    try:
        import ctypes
        _k = ctypes.windll.kernel32
        _k.SetConsoleMode(_k.GetStdHandle(-11), 7)
    except Exception:
        pass

class C:
    RESET="\033[0m"; BOLD="\033[1m"; ITALIC="\033[3m"
    ORANGE="\033[38;5;208m"; BLUE="\033[38;5;39m"; SKY="\033[38;5;117m"
    GREEN="\033[38;5;42m"; RED="\033[38;5;203m"; GOLD="\033[38;5;220m"; GRAY="\033[38;5;245m"

def titulo(t):
    print(); print(C.ORANGE + C.BOLD + "  " + t + C.RESET)
    print(C.GRAY + "  " + ("\u2500" * 48) + C.RESET)
def ok(t):    print(C.GREEN + "  \u2714 " + C.RESET + t)
def info(t):  print(C.BLUE + "  \u2022 " + C.RESET + t)
def dica(t):  print(C.SKY + C.ITALIC + "  \u203a " + t + C.RESET)
def aviso(t): print(C.GOLD + "  ! " + C.RESET + t)
def erro(t):  print(C.RED + "  \u2717 " + C.RESET + t)

def perguntar(texto):
    """Input com prompt em laranja (toda interacao com o usuario)."""
    try:
        return input(C.ORANGE + C.BOLD + "  " + texto + C.RESET)
    except EOFError:
        return ""


try:
    from selenium import webdriver
    from selenium.webdriver.firefox.options import Options
    from selenium.webdriver.firefox.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.firefox import GeckoDriverManager
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False


# Cache do geckodriver (evita checagem de rede repetida)
_GECKO_PATH_CACHE = None

def _obter_gecko_path():
    global _GECKO_PATH_CACHE
    if _GECKO_PATH_CACHE and os.path.isfile(_GECKO_PATH_CACHE):
        return _GECKO_PATH_CACHE
    _GECKO_PATH_CACHE = GeckoDriverManager().install()
    return _GECKO_PATH_CACHE


def _esperar_pagina(driver, timeout=15):
    """Espera document.readyState=complete. Retorna ao ficar pronta."""
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            if driver.execute_script("return document.readyState") == "complete":
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


# ============================================================
# CONFIGURAÇÕES
# ============================================================
FOTO_DIR = os.path.join(os.path.expanduser("~"), "Documents", "fotos_embaixadores")
URL_BADGE_PHOTO = "https://badgephotos.corp.amazon.com/?employeeid="


# Le preferencias do config.ini ([EMAIL])
#   metodo   = pwa | outlook_web   (pwa = so clipboard, mais rapido)
#   gerar_png = true | false       (screenshot do email para o historico)
def _ler_config_email():
    metodo, gerar_png = "pwa", True
    try:
        import configparser as _cp
        cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")
        if os.path.exists(cfg_path):
            cfg = _cp.ConfigParser()
            cfg.read(cfg_path, encoding="utf-8")
            if cfg.has_section("EMAIL"):
                metodo = cfg.get("EMAIL", "metodo", fallback="pwa").strip().lower()
                gerar_png = cfg.getboolean("EMAIL", "gerar_png", fallback=True)
    except Exception:
        pass
    return metodo, gerar_png


EMAIL_METODO, EMAIL_GERAR_PNG = _ler_config_email()


# ============================================================
# (LEGADO / NAO USADO NA DEMO) Download de fotos dos embaixadores.
# A versao de portfolio NAO baixa fotos. Mantido apenas por referencia.
# O ID viria do Employee Roster (Col A = Employee ID).
# ============================================================
def baixar_fotos_embaixadores(logins, ids_dict=None, foto_dir=FOTO_DIR):
    if not SELENIUM_OK:
        print("   Selenium nao disponivel. Fotos nao serao baixadas.")
        return

    if not ids_dict:
        print("   Nenhum mapeamento login->ID disponivel. Fotos nao serao baixadas.")
        print("   (O Employee Roster precisa ser baixado antes para obter os IDs)")
        return

    os.makedirs(foto_dir, exist_ok=True)

    # Verifica quais fotos ja existem
    exts = [".jpg", ".jpeg", ".png"]
    faltantes = []
    for login in logins:
        encontrado = False
        login_lower = login.lower()
        try:
            for arquivo in os.listdir(foto_dir):
                nome, ext_arq = os.path.splitext(arquivo)
                if ext_arq.lower() in exts and nome.lower() == login_lower:
                    encontrado = True
                    break
        except Exception:
            pass
        if not encontrado:
            if login_lower in ids_dict:
                faltantes.append(login)
            else:
                print("   " + login + ": ID nao encontrado no Roster. Foto ignorada.")

    if not faltantes:
        print("   Todas as fotos ja existem localmente. Nenhum download necessario.")
        return

    print("   " + str(len(faltantes)) + " foto(s) para baixar: " + ", ".join(faltantes))
    print("   Abrindo Firefox para acessar Badge Photos...")

    import configparser as _cp

    # Detecta perfil Firefox (mesma logica do AMBASSADORS.py)
    def _obter_perfil():
        candidates = [
            os.path.join(os.environ.get("APPDATA", ""), "Mozilla", "Firefox", "profiles.ini"),
            os.path.expanduser("~/Library/Application Support/Firefox/profiles.ini"),
            os.path.expanduser("~/.mozilla/firefox/profiles.ini"),
        ]
        for profiles_ini in candidates:
            if not os.path.exists(profiles_ini):
                continue
            cfg = _cp.ConfigParser()
            try:
                cfg.read(profiles_ini, encoding="utf-8")
            except Exception:
                continue
            base_dir = os.path.dirname(profiles_ini)
            for section in cfg.sections():
                if section.lower().startswith("install"):
                    default = cfg.get(section, "Default", fallback=None)
                    if default:
                        p = os.path.join(base_dir, default.replace("/", os.sep))
                        if os.path.isdir(p):
                            return p
            for section in cfg.sections():
                if section.lower().startswith("profile"):
                    if cfg.get(section, "Default", fallback="") == "1":
                        is_rel = cfg.getint(section, "IsRelative", fallback=1)
                        path = cfg.get(section, "Path", fallback=None)
                        if path:
                            p = (os.path.join(base_dir, path.replace("/", os.sep))
                                 if is_rel else path)
                            if os.path.isdir(p):
                                return p
        return None

    driver = None
    try:
        # Fecha Firefox para liberar o perfil (-profile e rapido, sem copiar)
        _kill_firefox()
        options = Options()
        perfil_path = _obter_perfil()
        if perfil_path:
            options.add_argument("-profile")
            options.add_argument(perfil_path)
            print("   Perfil Firefox: " + perfil_path)
        else:
            print("   AVISO: Perfil nao encontrado. Login pode ser necessario.")

        service = Service(_obter_gecko_path())
        driver = webdriver.Firefox(service=service, options=options)
        driver.set_page_load_timeout(30)

        # Testa autenticacao com a primeira foto
        primeiro_id = ids_dict[faltantes[0].lower()]
        primeiro_url = URL_BADGE_PHOTO + str(primeiro_id)
        driver.get(primeiro_url)
        _esperar_pagina(driver, timeout=8)

        # Verifica se caiu em pagina de login
        current = driver.current_url.lower()
        if any(k in current for k in ["midway", "signin", "login", "sso", "auth"]):
            print()
            print("   LOGIN NECESSARIO")
            print("   Faca login com sua YubiKey no navegador e pressione ENTER.")
            perguntar("   ENTER apos fazer login... ")
            driver.get(primeiro_url)
            _esperar_pagina(driver, timeout=8)

        from selenium.webdriver.common.by import By

        baixados = 0
        erros = []

        for i, login in enumerate(faltantes):
            emp_id = ids_dict[login.lower()]
            url_foto = URL_BADGE_PHOTO + str(emp_id)
            print("   [" + str(i + 1) + "/" + str(len(faltantes)) + "] "
                  + login + " (ID: " + str(emp_id) + ")...", end="", flush=True)

            try:
                driver.get(url_foto)
                # Espera ate a imagem carregar (max 4s) em vez de 2s fixos
                fim = time.time() + 4
                imgs = []
                while time.time() < fim:
                    try:
                        if driver.execute_script("return document.readyState") == "complete":
                            imgs = driver.find_elements(By.TAG_NAME, "img")
                            if imgs:
                                # confirma que ao menos 1 img tem dimensao
                                pronta = driver.execute_script(
                                    "return Array.from(document.images).some(i => (i.naturalWidth||0) >= 50)")
                                if pronta:
                                    break
                    except Exception:
                        pass
                    time.sleep(0.2)
                if not imgs:
                    imgs = driver.find_elements(By.TAG_NAME, "img")

                foto_salva = False
                for img in imgs:
                    try:
                        b64_data = driver.execute_script("""
                            var img = arguments[0];
                            var w = img.naturalWidth || img.width;
                            var h = img.naturalHeight || img.height;
                            if (w < 50 || h < 50) return '';
                            var canvas = document.createElement('canvas');
                            canvas.width = w;
                            canvas.height = h;
                            var ctx = canvas.getContext('2d');
                            ctx.drawImage(img, 0, 0);
                            return canvas.toDataURL('image/png').split(',')[1];
                        """, img)

                        if b64_data and len(b64_data) > 500:
                            caminho_foto = os.path.join(foto_dir, login.lower() + ".png")
                            with open(caminho_foto, "wb") as f:
                                f.write(base64.b64decode(b64_data))
                            print(" OK")
                            baixados += 1
                            foto_salva = True
                            break
                    except Exception:
                        continue

                if not foto_salva:
                    print(" NAO ENCONTRADA")
                    erros.append(login)

            except Exception as e:
                print(" ERRO: " + str(e))
                erros.append(login)

        print()
        print("   Download concluido: " + str(baixados) + " foto(s) baixada(s)")
        if erros:
            print("   Nao encontradas: " + ", ".join(erros))

    except Exception as e:
        print("   ERRO ao baixar fotos: " + str(e))
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# ============================================================
# LER RANKING FINAL
# Lê ranking_final.xlsx e separa dados por Fonte e Classificação
# Retorna (df_top, df_bot, df_targets)
# ============================================================
def ler_ranking_final(caminho):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
    except Exception as e:
        print("   ERRO ao ler ranking_final.xlsx: " + str(e))
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


    col_names = ["Embaixador", "Fonte", "Associado", "Turno", "Processo",
                 "UPH PPR", "DPMO", "Rate LC", "Total Defects", "Opportunities",
                 "DPMO Emb", "Rate Emb"]


    top_data    = []
    bot_data    = []
    target_data = []
    current_sec = None


    # Detecta coluna inicial da mini-tabela procurando "TARGETS POR PROCESSO"
    mini_col_start = 15   # fallback: coluna O (1-based)
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for ci, cell_val in enumerate(row, 1):
            if cell_val and "TARGETS" in str(cell_val).upper():
                mini_col_start = ci
                break


    for row in ws.iter_rows(values_only=True):
        values  = list(row)
        val_a   = str(values[0]).strip() if values[0] is not None else ""
        val_nan = all(v is None or str(v).strip() in ("", "None", "nan")
                      for v in values[:10])


        if val_nan:
            continue


        # Detecta secao TOP ou BOTTOM
        if "TOP AMBASSADORS" in val_a.upper():
            current_sec = "TOP"
            continue
        if "BOTTOM AMBASSADORS" in val_a.upper():
            current_sec = "BOTTOM"
            continue


        # Pula linha de cabecalho
        if val_a == "Embaixador":
            continue


        # Linha de dados
        if current_sec in ("TOP", "BOTTOM") and val_a and val_a.lower() not in ("nan", "none"):
            row_dict = {}
            for i, col in enumerate(col_names):
                row_dict[col] = (str(values[i]).strip()
                                 if i < len(values) and values[i] is not None else "")
            if current_sec == "TOP":
                top_data.append(row_dict)
            else:
                bot_data.append(row_dict)


        # Mini-tabela de targets (colunas M em diante)
        if len(values) >= mini_col_start:
            idx = mini_col_start - 1   # 0-based
            proc_val = str(values[idx]).strip() if values[idx] is not None else ""
            if (proc_val and proc_val.lower() not in
                    ("", "nan", "none", "targets por processo", "processo")):
                target_data.append({
                    "Processo":           proc_val,
                    "Target Prod (UPH)":  (str(values[idx + 1]).strip()
                                           if len(values) > idx + 1 and values[idx + 1] is not None else ""),
                    "Target Qual (DPMO)": (str(values[idx + 2]).strip()
                                           if len(values) > idx + 2 and values[idx + 2] is not None else ""),
                })


    wb.close()
    df_top = pd.DataFrame(top_data)    if top_data else pd.DataFrame(columns=col_names)
    df_bot = pd.DataFrame(bot_data)    if bot_data else pd.DataFrame(columns=col_names)
    df_targets = (pd.DataFrame(target_data) if target_data
                  else pd.DataFrame(columns=["Processo", "Target Prod (UPH)", "Target Qual (DPMO)"]))


    print("   ranking_final lido:")
    print("      TOP:    " + str(df_top["Embaixador"].nunique() if len(df_top) > 0 else 0) + " embaixadores")
    print("      BOTTOM: " + str(df_bot["Embaixador"].nunique() if len(df_bot) > 0 else 0) + " embaixadores")
    print("      Targets: " + str(len(df_targets)) + " processos")
    return df_top, df_bot, df_targets


# ============================================================
# BUSCAR FOTO
# Tenta: login exato, login em maiusculo, login em minusculo,
# e busca case-insensitive em todos os arquivos da pasta.
# ============================================================
def buscar_foto_base64(login, foto_dir=FOTO_DIR):
    os.makedirs(foto_dir, exist_ok=True)
    exts = [".jpg", ".jpeg", ".png"]

    # 1. Tentativas diretas: exato, upper, lower
    for variante in [login, login.upper(), login.lower()]:
        for ext in exts:
            caminho = os.path.join(foto_dir, variante + ext)
            if os.path.exists(caminho):
                mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
                with open(caminho, "rb") as f:
                    encoded = base64.b64encode(f.read()).decode("utf-8")
                return "data:" + mime + ";base64," + encoded, True

    # 2. Busca case-insensitive em todos os arquivos da pasta
    login_lower = login.lower()
    try:
        for arquivo in os.listdir(foto_dir):
            nome, ext = os.path.splitext(arquivo)
            if ext.lower() in exts and nome.lower() == login_lower:
                caminho = os.path.join(foto_dir, arquivo)
                mime = "image/jpeg" if ext.lower() in (".jpg", ".jpeg") else "image/png"
                with open(caminho, "rb") as f:
                    encoded = base64.b64encode(f.read()).decode("utf-8")
                return "data:" + mime + ";base64," + encoded, True
    except Exception:
        pass

    return None, False


# ============================================================
# VERIFICAR EMBAIXADOR NO FC
# Returns: (manter: bool, foto_uri: str|None)
# DEMO / PORTFOLIO: nao usamos fotos. O embaixador e SEMPRE mantido
# no email; quando nao ha foto, o card usa as iniciais como fallback.
# (Sem pausas nem perguntas — o email sai com todos os embaixadores.)
# ============================================================
def verificar_embaixador_no_fc(login, foto_dir=FOTO_DIR):
    foto_uri, encontrado = buscar_foto_base64(login, foto_dir)
    if encontrado:
        return True, foto_uri
    # Sem foto: mantem no email com iniciais.
    return True, None


# ============================================================
# UTILITÁRIOS
# ============================================================
def num_para_float(valor):
    try:
        return float(str(valor).strip().replace(",", "."))
    except Exception:
        return 0.0


def _td_style():
    return ("padding:8px 10px;font-size:12px;text-align:center;"
            "border:1px solid #e0e0e0;")


def _th_style(bg):
    return ("padding:9px 10px;font-size:12px;font-weight:bold;"
            "text-align:center;color:#fff;background:" + bg + ";"
            "border:1px solid rgba(255,255,255,0.2);")


# ============================================================
# SUB-TABELA QUALIDADE
# Colunas: Associado | Turno | Processo | Total Defects | Opportunities
# (Rate LC ignorado)
# ============================================================
def _tabela_qualidade(df_emb, bg_h, bg_par, bg_impar):
    if df_emb is None or len(df_emb) == 0:
        return ""
    colunas    = ["Associado", "Turno", "Processo", "Total Defects", "Opportunities"]
    cabecalhos = "".join('<th style="' + _th_style(bg_h) + '">' + c + "</th>"
                         for c in colunas)
    linhas_html = ""
    for i, (_, row) in enumerate(df_emb.iterrows()):
        bg = bg_par if i % 2 == 0 else bg_impar
        linhas_html += (
            '<tr style="background:' + bg + ';">'
            + '<td style="' + _td_style() + '">' + str(row.get("Associado",   "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Turno",       "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Processo",    "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Total Defects","-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Opportunities","-")) + "</td>"
            + "</tr>"
        )
    return ('<table width="100%" cellpadding="0" cellspacing="0" '
            'style="border-collapse:collapse;">'
            "<tr>" + cabecalhos + "</tr>"
            + linhas_html + "</table>")


# ============================================================
# SUB-TABELA PRODUTIVIDADE
# Colunas: Associado | Turno | Processo | UPH PPR | DPMO | Rate LC
# ============================================================
def _tabela_produtividade(df_emb, bg_h, bg_par, bg_impar):
    if df_emb is None or len(df_emb) == 0:
        return ""
    colunas    = ["Associado", "Turno", "Processo", "UPH PPR", "DPMO", "Rate LC"]
    cabecalhos = "".join('<th style="' + _th_style(bg_h) + '">' + c + "</th>"
                         for c in colunas)
    linhas_html = ""
    for i, (_, row) in enumerate(df_emb.iterrows()):
        bg = bg_par if i % 2 == 0 else bg_impar
        linhas_html += (
            '<tr style="background:' + bg + ';">'
            + '<td style="' + _td_style() + '">' + str(row.get("Associado", "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Turno",     "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("Processo",  "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("UPH PPR",   "-")) + "</td>"
            + '<td style="' + _td_style() + '">' + str(row.get("DPMO",      "-")) + "</td>"
            + '<td style="' + _td_style() + ';font-weight:bold;">'
            + str(row.get("Rate LC", "-")) + "</td>"
            + "</tr>"
        )
    return ('<table width="100%" cellpadding="0" cellspacing="0" '
            'style="border-collapse:collapse;">'
            "<tr>" + cabecalhos + "</tr>"
            + linhas_html + "</table>")


# ============================================================
# CARD TOP — foto + DPMO/Rate Emb + qual + prod
# ============================================================
def _card_top_embaixador(emb, df_qual, df_prod, dpmo_emb, rate_emb, foto_uri):
    if foto_uri:
        foto_html = ('<img src="' + foto_uri + '" '
                     'style="width:90px;height:90px;border-radius:50%;'
                     'border:3px solid #FFD700;object-fit:cover;'
                     'display:block;margin:0 auto;" />')
    else:
        iniciais  = emb[:2].upper()
        foto_html = ('<div style="width:90px;height:90px;border-radius:50%;'
                     'background:#2E7D32;border:3px solid #FFD700;'
                     'display:flex;align-items:center;justify-content:center;'
                     'margin:0 auto;font-size:28px;color:#fff;font-weight:bold;'
                     'line-height:90px;text-align:center;">'
                     + iniciais + "</div>")


    dpmo_str  = str(dpmo_emb).strip() if str(dpmo_emb).strip() not in ("", "nan") else "—"
    rate_str  = str(rate_emb).strip() if str(rate_emb).strip() not in ("", "nan") else "—"

    tem_qual = df_qual is not None and len(df_qual) > 0
    tem_prod = df_prod is not None and len(df_prod) > 0

    # Linha de metrica conforme o destaque:
    # QUALIDADE -> mostra so DPMO Performace
    # PRODUTIVIDADE -> mostra so RATE Performace
    if tem_qual:
        metrica_html = "DPMO Performace: " + dpmo_str
    elif tem_prod:
        metrica_html = "RATE Performace: " + rate_str
    else:
        metrica_html = "DPMO Performace: " + dpmo_str + " &nbsp;|&nbsp; RATE Performace: " + rate_str

    qual_section = ""
    if tem_qual:
        qual_section = (
            '<tr style="background:#1B5E20;">'
            '<td colspan="2" style="padding:10px 16px;text-align:center;">'
            '<span style="color:#FFD700;font-weight:bold;font-size:14px;letter-spacing:1px;">QUALIDADE</span>'
            "</td></tr>"
            "<tr><td colspan=\"2\" style=\"padding:0;\">"
            + _tabela_qualidade(df_qual, "#1B5E20", "#E8F5E9", "#C8E6C9")
            + "</td></tr>"
        )

    prod_section = ""
    if tem_prod:
        prod_section = (
            '<tr style="background:#1B5E20;">'
            '<td colspan="2" style="padding:10px 16px;text-align:center;">'
            '<span style="color:#FFD700;font-weight:bold;font-size:14px;letter-spacing:1px;">PRODUTIVIDADE</span>'
            "</td></tr>"
            "<tr><td colspan=\"2\" style=\"padding:0;\">"
            + _tabela_produtividade(df_prod, "#2E7D32", "#E8F5E9", "#C8E6C9")
            + "</td></tr>"
        )


    return (
        '<table width="100%" cellpadding="0" cellspacing="0" '
        'style="margin-bottom:28px;border-radius:8px;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(0,0,0,0.1);">'
        # Cabecalho centralizado: foto + info agrupados no meio do card
        '<tr style="background:#2E7D32;">'
        '<td style="padding:22px 20px;">'
        '<table align="center" cellpadding="0" cellspacing="0" style="margin:0 auto;">'
        '<tr>'
        # Coluna da foto
        '<td style="text-align:center;vertical-align:middle;padding-right:22px;">'
        + foto_html
        + "</td>"
        # Coluna do texto (alinhado a esquerda dentro do bloco centralizado)
        '<td style="text-align:left;vertical-align:middle;">'
        '<h2 style="color:#FFD700;margin:0;font-size:20px;">' + emb.upper() + "</h2>"
        '<p style="color:#A5D6A7;margin:4px 0 0;font-size:13px;">Embaixador Top Performer</p>'
        '<p style="color:#FFD700;margin:6px 0 0;font-size:13px;font-weight:bold;">'
        + metrica_html
        + "</p>"
        "</td>"
        "</tr></table>"
        "</td></tr>"
        + qual_section
        + prod_section
        + "</table>"
    )


# ============================================================
# CARD BOTTOM — com DPMO/Rate + qual + prod
# ============================================================
def _card_bottom_embaixador(emb, df_qual, df_prod, dpmo_emb="", rate_emb=""):
    dpmo_str = str(dpmo_emb).strip() if str(dpmo_emb).strip() not in ("", "nan") else "—"
    rate_str = str(rate_emb).strip() if str(rate_emb).strip() not in ("", "nan") else "—"

    tem_qual = df_qual is not None and len(df_qual) > 0
    tem_prod = df_prod is not None and len(df_prod) > 0

    # Linha de metrica conforme o destaque (so DPMO ou so RATE)
    if tem_qual:
        metrica_html = "DPMO Performace: " + dpmo_str
    elif tem_prod:
        metrica_html = "RATE Performace: " + rate_str
    else:
        metrica_html = "DPMO Performace: " + dpmo_str + " &nbsp;|&nbsp; RATE Performace: " + rate_str

    qual_section = ""
    if tem_qual:
        qual_section = (
            '<tr style="background:#8E0000;">'
            '<td style="padding:10px 16px;text-align:center;">'
            '<span style="color:#FFFFFF;font-weight:bold;font-size:14px;letter-spacing:1px;">QUALIDADE</span>'
            "</td></tr>"
            "<tr><td style=\"padding:0;\">"
            + _tabela_qualidade(df_qual, "#B71C1C", "#FFEBEE", "#FFCDD2")
            + "</td></tr>"
        )

    prod_section = ""
    if tem_prod:
        prod_section = (
            '<tr style="background:#8E0000;">'
            '<td style="padding:10px 16px;text-align:center;">'
            '<span style="color:#FFFFFF;font-weight:bold;font-size:14px;letter-spacing:1px;">PRODUTIVIDADE</span>'
            "</td></tr>"
            "<tr><td style=\"padding:0;\">"
            + _tabela_produtividade(df_prod, "#BF360C", "#FFF3E0", "#FFE0B2")
            + "</td></tr>"
        )


    return (
        '<table width="100%" cellpadding="0" cellspacing="0" '
        'style="margin-bottom:28px;border-radius:8px;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(0,0,0,0.1);">'
        '<tr style="background:#C62828;">'
        '<td style="padding:18px 20px;text-align:center;">'
        '<h3 style="color:#FFCDD2;margin:0;font-size:17px;">'
        "Embaixador: <strong style=\"color:#fff;\">" + emb.upper() + "</strong></h3>"
        '<p style="color:#FFCDD2;margin:4px 0 0;font-size:12px;">'
        "Associados que necessitam de suporte adicional"
        "</p>"
        '<p style="color:#FFCDD2;margin:6px 0 0;font-size:13px;font-weight:bold;">'
        + metrica_html
        + "</p></td></tr>"
        + qual_section
        + prod_section
        + "</table>"
    )


# ============================================================
# SEÇÃO TOP — verifica foto com pausa, filtra ausentes do FC
# Returns: (html, ambassadors_sem_foto)
# ============================================================
def gerar_secao_top(df_top, foto_dir=FOTO_DIR):
    if df_top is None or len(df_top) == 0:
        return "<p style='color:#777;'>Nenhum embaixador TOP neste período.</p>", set()


    ambassadors_sem_foto = set()
    cards  = ""
    embs_ok = []


    for emb in sorted(df_top["Embaixador"].unique()):
        manter, foto_uri = verificar_embaixador_no_fc(emb, foto_dir)
        if not manter:
            ambassadors_sem_foto.add(emb)
            continue
        embs_ok.append(emb)


        df_emb      = df_top[df_top["Embaixador"] == emb].copy()
        df_qual_emb = df_emb[df_emb["Fonte"] == "Qualidade"].reset_index(drop=True)
        df_prod_emb = df_emb[df_emb["Fonte"] == "Produtividade"].reset_index(drop=True)

        # Mostra APENAS a fonte pela qual o embaixador foi destaque.
        # Qualidade tem prioridade: se tem dados de qualidade, e destaque
        # de QUALIDADE -> mostra so qualidade. Senao, so produtividade.
        if len(df_qual_emb) > 0:
            df_prod_emb = df_prod_emb.iloc[0:0]   # zera produtividade
        else:
            df_qual_emb = df_qual_emb.iloc[0:0]   # garante qualidade vazia


        # Pega DPMO Emb e Rate Emb (mesmo valor em todas as linhas do embaixador)
        dpmo_emb = ""
        rate_emb = ""
        for _, row in df_emb.iterrows():
            dv = str(row.get("DPMO Emb", "")).strip()
            rv = str(row.get("Rate Emb", "")).strip()
            if dv and dv.lower() not in ("", "nan", "none"):
                dpmo_emb = dv
            if rv and rv.lower() not in ("", "nan", "none"):
                rate_emb = rv
            if dpmo_emb and rate_emb:
                break


        cards += _card_top_embaixador(emb, df_qual_emb, df_prod_emb, dpmo_emb, rate_emb, foto_uri)


    n_ok      = len(embs_ok)
    n_removed = len(ambassadors_sem_foto)

    html = (
        '<tr style="background:#1B5E20;">'
        '<td style="padding:18px 30px;">'
        '<h2 style="color:#FFD700;margin:0;font-size:18px;">'
        "🏆 TOP PERFORMERS"
        '<span style="font-size:13px;color:#A5D6A7;"> (' + str(n_ok) + " embaixadores)</span>"
        "</h2></td></tr>"
        "<tr><td style=\"padding:20px 30px;\">" + cards + "</td></tr>"
    )
    return html, ambassadors_sem_foto


# ============================================================
# SEÇÃO BOTTOM — sem foto, remove ausentes do FC
# ============================================================
def gerar_secao_bottom(df_bot, ambassadors_sem_foto=None):
    if df_bot is None or len(df_bot) == 0:
        return ""
    if ambassadors_sem_foto is None:
        ambassadors_sem_foto = set()


    cards     = ""
    embs_validos = []


    for emb in sorted(df_bot["Embaixador"].unique()):
        if emb in ambassadors_sem_foto:
            print("   BOTTOM: " + emb + " removido (saiu do FC)")
            continue
        embs_validos.append(emb)
        df_emb      = df_bot[df_bot["Embaixador"] == emb].copy()
        df_qual_emb = df_emb[df_emb["Fonte"] == "Qualidade"].reset_index(drop=True)
        df_prod_emb = df_emb[df_emb["Fonte"] == "Produtividade"].reset_index(drop=True)

        # Mostra APENAS a fonte pela qual o embaixador foi classificado.
        # Qualidade tem prioridade: se tem dados de qualidade, mostra so
        # qualidade; senao, mostra so produtividade.
        if len(df_qual_emb) > 0:
            df_prod_emb = df_prod_emb.iloc[0:0]
        else:
            df_qual_emb = df_qual_emb.iloc[0:0]

        # Extrair DPMO Emb e Rate Emb do embaixador
        dpmo_emb = ""
        rate_emb = ""
        for _, row in df_emb.iterrows():
            dv = str(row.get("DPMO Emb", "")).strip()
            rv = str(row.get("Rate Emb", "")).strip()
            if dv and dv.lower() not in ("", "nan", "none"):
                dpmo_emb = dv
            if rv and rv.lower() not in ("", "nan", "none"):
                rate_emb = rv
            if dpmo_emb and rate_emb:
                break

        cards += _card_bottom_embaixador(emb, df_qual_emb, df_prod_emb, dpmo_emb, rate_emb)


    if not cards:
        return ""


    return (
        "<tr><td style=\"padding:10px 30px 0;\">"
        "<hr style=\"border:none;border-top:2px solid #e5e7eb;margin:0;\" /></td></tr>"
        '<tr style="background:#B71C1C;">'
        '<td style="padding:18px 30px;">'
        '<h2 style="color:#FFCDD2;margin:0;font-size:18px;">'
        "⚠️ BOTTOM PERFORMERS"
        '<span style="font-size:13px;"> (' + str(len(embs_validos)) + " embaixadores)</span>"
        "</h2>"
        '<p style="color:#FFCDD2;margin:6px 0 0;font-size:13px;">'
        "Atencao! Identificamos Associados LC1 e LC2 que necessitam de suporte adicional."
        " Os tickets correspondentes ja foram abertos para acompanhamento."
        "</p></td></tr>"
        "<tr><td style=\"padding:20px 30px;\">" + cards
        + '<p style="font-size:13px;color:#555;margin-top:10px;">'
        'Em caso de duvidas, consulte a <a href="#" '
        'style="color:#007185;">Wiki – Guia de mecanismos (FC locale)</a> ou entre em contato.'
        "</p></td></tr>"
    )


# ============================================================
# SEÇÃO TARGETS POR PROCESSO
# Dados da mini-tabela do ranking_final
# ============================================================
def gerar_secao_targets(df_targets, data_str=""):
    if df_targets is None or len(df_targets) == 0:
        return ""


    cabecalhos = "".join(
        '<th style="' + _th_style("#0D47A1") + '">' + c + "</th>"
        for c in ["Processo", "Rate Target (UPH)", "DPMO Target"]
    )
    linhas_html = ""
    for i, (_, row) in enumerate(df_targets.iterrows()):
        proc = str(row.get("Processo",           "")).strip()
        rate = str(row.get("Target Prod (UPH)",  "")).strip()
        dpmo = str(row.get("Target Qual (DPMO)", "")).strip()
        if not proc or proc.lower() in ("", "nan", "none"):
            continue
        bg = "#E3F2FD" if i % 2 == 0 else "#BBDEFB"
        linhas_html += (
            '<tr style="background:' + bg + ';">'
            '<td style="' + _td_style() + ';font-weight:bold;">' + proc + "</td>"
            '<td style="' + _td_style() + '">' + (rate if rate else "—") + "</td>"
            '<td style="' + _td_style() + '">' + (dpmo if dpmo else "—") + "</td>"
            "</tr>"
        )


    if not linhas_html:
        return ""


    titulo_data = ("ONB DO DIA " + data_str) if data_str else "ESTE PERIODO"


    return (
        "<tr><td style=\"padding:10px 30px 0;\">"
        "<hr style=\"border:none;border-top:2px solid #e5e7eb;margin:0;\" /></td></tr>"
        '<tr style="background:#0D47A1;">'
        '<td style="padding:18px 30px;">'
        '<h2 style="color:#FFD700;margin:0;font-size:18px;">'
        "📊 TARGETS DO PERIODO &mdash; " + titulo_data
        + "</h2>"
        '<p style="color:#90CAF9;margin:6px 0 0;font-size:13px;">'
        "Estes foram os parametros de referencia utilizados para classificar os embaixadores "
        "neste ciclo de ONB. DPMO acima do target indica risco de qualidade; "
        "Rate abaixo do target indica baixa produtividade."
        "</p></td></tr>"
        '<tr><td style="padding:16px 30px 24px;">'
        '<table width="100%" cellpadding="0" cellspacing="0" '
        'style="border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);">'
        "<tr>" + cabecalhos + "</tr>"
        + linhas_html + "</table>"
        '<p style="font-size:12px;color:#777;margin-top:10px;font-style:italic;">'
        "Rate Target (UPH) = unidades por hora esperadas &nbsp;&middot;&nbsp; "
        "DPMO Target = defeitos por milhao de oportunidades aceitaveis"
        "</p>"

        # Explicacao clara das metricas de classificacao
        '<div style="margin-top:16px;padding:14px 18px;background:#E3F2FD;'
        'border-left:4px solid #0D47A1;border-radius:4px;">'
        '<p style="font-size:13px;color:#0D47A1;font-weight:bold;margin:0 0 8px;">'
        "Como funciona a classificacao?</p>"
        '<p style="font-size:12px;color:#333;margin:0 0 6px;line-height:1.6;">'
        "<strong>TOP (Qualidade):</strong> DPMO do embaixador &le; 30% do Target DPMO do processo.<br />"
        "<strong>TOP (Produtividade):</strong> Rate LC medio dos associados &ge; 90% do Target UPH.<br />"
        "<strong>BOTTOM (Qualidade):</strong> DPMO do embaixador &gt; 100% do Target DPMO do processo.<br />"
        "<strong>BOTTOM (Produtividade):</strong> Rate LC medio dos associados &le; 60% do Target UPH."
        "</p>"
        '<p style="font-size:12px;color:#555;margin:8px 0 0;line-height:1.5;">'
        "O DPMO do embaixador e calculado somando os defeitos e oportunidades de todos os seus associados. "
        "O Rate LC e a media da produtividade real dividida pelo target do processo."
        "</p></div>"

        "</td></tr>"
    )


# ============================================================
# MONTAR HTML COMPLETO DO EMAIL
# ============================================================
def montar_html_email(df_top, df_bot, df_targets, data_str, foto_dir=FOTO_DIR):
    secao_top, ambassadors_sem_foto = gerar_secao_top(df_top, foto_dir)
    secao_bottom  = gerar_secao_bottom(df_bot, ambassadors_sem_foto)
    secao_targets = gerar_secao_targets(df_targets, data_str)

    n_top = max(0, (df_top["Embaixador"].nunique() if len(df_top) > 0 else 0)
                - len(ambassadors_sem_foto))
    n_bot = len([e for e in (df_bot["Embaixador"].unique() if len(df_bot) > 0 else [])
                 if e not in ambassadors_sem_foto])

    # Pega o login do usuario logado no sistema operacional
    login_usuario = os.getenv("USERNAME") or os.getenv("USER") or os.getenv("LOGNAME") or "ambassador"
    login_usuario = login_usuario.lower()

    # Titulo: "TOP AMBASSADOR\nONBOARDING: XX/XX"
    # data_str pode ser "21.04 | 28.04" — converte pontos para barras
    data_titulo = data_str.replace(".", "/") if data_str else ""
    titulo_linha1 = "TOP AMBASSADOR"
    titulo_linha2 = "ONBOARDING: " + data_titulo if data_titulo else "ONBOARDING"


    html = (
        "<html><head><meta charset='utf-8' /></head>"
        "<body style='margin:0;padding:0;font-family:Arial,sans-serif;"
        "background:#f3f4f6;color:#333;'>"
        "<table width='100%' cellpadding='0' cellspacing='0' bgcolor='#f3f4f6'>"
        "<tr><td align='center' style='padding:20px 10px;'>"
        "<table width='720' cellpadding='0' cellspacing='0' bgcolor='#ffffff'"
        " style='border-radius:8px;overflow:hidden;box-shadow:0 0 12px rgba(0,0,0,0.08);'>"


        # HEADER
        "<tr style='background:#131A22;'>"
        "<td style='padding:28px 30px;text-align:center;'>"
        "<h1 style='color:#FFD700;font-size:24px;margin:0;line-height:1.3;'>"
        + titulo_linha1 + "<br />"
        "<span style='font-size:18px;'>" + titulo_linha2 + "</span></h1>"
        "<p style='color:#aaa;font-size:13px;margin:8px 0 0;'>"
        + str(n_top) + " embaixador(es) TOP &nbsp;&middot;&nbsp; "
        + str(n_bot) + " embaixador(es) BOTTOM</p>"
        "</td></tr>"


        # INTRO
        "<tr><td style='padding:20px 30px 10px;'>"
        "<p style='font-size:14px;line-height:1.6;margin:0;'>"
        "Ol&aacute;,<br /><br />"
        "Segue abaixo o relat&oacute;rio de performance dos "
        "<strong>Embaixadores (FC locale)</strong>. Os dados foram extra&iacute;dos "
        "do Employee Roster e do ATLAS, refletindo a performance dos associados treinados "
        "no per&iacute;odo <strong>ONB " + data_str + "</strong>."
        "</p></td></tr>"


        + secao_top
        + secao_bottom
        + secao_targets


        # ASSINATURA
        + "<tr><td style='padding:20px 30px;'>"
        "<p style='font-size:14px;margin:0;'>Atenciosamente,</p>"
        "<p style='font-size:14px;font-weight:bold;margin:4px 0 0;'>" + login_usuario + "</p>"
        "<p style='font-size:13px;margin:2px 0;'>" + login_usuario + " &middot; Amazon Brasil</p>"
        "<p style='font-size:13px;margin:2px 0;'>"
        "<a href='mailto:" + login_usuario + "@amazon.com' style='color:#007185;'>"
        + login_usuario + "@amazon.com</a></p>"
        "<p style='font-size:11px;color:#999;font-style:italic;margin-top:8px;'>"
        "Work hard. Have fun. Make history.</p>"
        "</td></tr>"


        # RODAPÉ
        "<tr style='background:#232f3e;'>"
        "<td style='padding:16px;text-align:center;'>"
        "<p style='color:#ff9900;font-size:12px;font-style:italic;margin:0;'>"
        "Top Ambassador Learning &middot; FC locale &middot; Demo</p>"
        "</td></tr>"
        "</table></td></tr></table></body></html>"
    )
    return html


# ============================================================
# FECHAR FIREFOX EXISTENTE (libera o perfil para o Selenium)
# ============================================================
def _kill_firefox():
    print("   Fechando Firefox existente...", end="", flush=True)
    killed = False
    if sys.platform == "win32":
        try:
            r = subprocess.run(["taskkill", "/F", "/IM", "firefox.exe"],
                               capture_output=True, text=True)
            killed = r.returncode == 0
        except Exception:
            pass
    else:
        try:
            subprocess.run(["pkill", "-f", "firefox"], capture_output=True)
            killed = True
        except Exception:
            pass
    print(" OK" if killed else " OK (nenhum aberto)")
    if killed:
        time.sleep(1.2)


def _safe_find(driver, selectors, timeout=15):
    for sel in selectors:
        try:
            return WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
            )
        except Exception:
            continue
    return None


# ============================================================
# ABRIR EMAIL NO OUTLOOK WEB COM HTML INJETADO
# Abre o Outlook Web, clica em "New mail" e injeta o HTML
# no corpo (preserva todas as cores).
# NAO preenche destinatarios e NAO envia — usuario faz isso.
# ============================================================
def abrir_email_outlook(html, data_str=""):
    if not SELENIUM_OK:
        print("   Selenium nao disponivel. Nao foi possivel abrir no Outlook Web.")
        return False

    import configparser as _cp

    def _obter_perfil():
        if sys.platform == "win32":
            ff_dir = os.path.join(os.environ.get("APPDATA", ""), "Mozilla", "Firefox")
        elif sys.platform == "darwin":
            ff_dir = os.path.expanduser("~/Library/Application Support/Firefox")
        else:
            ff_dir = os.path.expanduser("~/.mozilla/firefox")
        ini_path = os.path.join(ff_dir, "profiles.ini")
        if not os.path.exists(ini_path):
            return None
        c = _cp.ConfigParser()
        try:
            c.read(ini_path)
        except Exception:
            return None
        # Prioriza perfil default-esr / Default=1
        for section in c.sections():
            if section.lower().startswith("install"):
                default = c.get(section, "Default", fallback=None)
                if default:
                    p = os.path.join(ff_dir, default.replace("/", os.sep))
                    if os.path.isdir(p):
                        return p
        for section in c.sections():
            if not section.startswith("Profile"):
                continue
            path = c.get(section, "Path", fallback="")
            if not path:
                continue
            is_rel = c.get(section, "IsRelative", fallback="1") == "1"
            full = os.path.join(ff_dir, path) if is_rel else path
            if os.path.isdir(full):
                return full
        return None

    assunto = "TOP AMBASSADOR - ONBOARDING: " + (data_str.replace(".", "/") if data_str else "")

    driver = None
    try:
        # Fecha Firefox para liberar o perfil
        _kill_firefox()

        print("   Abrindo Outlook Web no Firefox...")
        options = Options()
        perfil_path = _obter_perfil()
        if perfil_path:
            options.add_argument("-profile")
            options.add_argument(perfil_path)

        service = Service(_obter_gecko_path())
        driver = webdriver.Firefox(service=service, options=options)
        driver.set_page_load_timeout(60)
        driver.implicitly_wait(3)

        driver.get("https://outlook.office.com/mail/")
        _esperar_pagina(driver, timeout=20)

        # Verifica se precisa de login
        current = driver.current_url.lower()
        if any(k in current for k in ["login", "signin", "sso", "auth", "microsoftonline"]):
            print("   Login necessario. Faca login no navegador.")
            perguntar("   ENTER apos fazer login... ")
            driver.get("https://outlook.office.com/mail/")
            _esperar_pagina(driver, timeout=20)

        # Clica em "New mail"
        new_btn = _safe_find(driver, [
            'button[aria-label="New mail"]',
            'button[aria-label="Nova mensagem"]',
            'button[title="New mail"]',
            'button[title="Nova mensagem"]',
            '[data-icon-name="ComposeRegular"]',
        ], timeout=20)
        if not new_btn:
            print("   Botao 'New mail' nao encontrado. Abra um novo email manualmente.")
            perguntar("   ENTER apos abrir o novo email... ")
        else:
            new_btn.click()
            time.sleep(3)

        # Preenche o assunto (opcional)
        subj_field = _safe_find(driver, [
            'input[aria-label="Add a subject"]',
            'input[aria-label="Adicionar um assunto"]',
            'input[aria-label="Subject"]',
        ], timeout=8)
        if subj_field:
            try:
                subj_field.click()
                time.sleep(0.3)
                subj_field.send_keys(assunto)
                time.sleep(0.4)
            except Exception:
                pass

        # Localiza o corpo do email
        time.sleep(2)
        body_el = None
        for sel in [
            'div[aria-label="Message body, press Alt+F10 to exit"]',
            'div[aria-label="Corpo da mensagem"]',
            'div[aria-label="Message body"]',
            'div[role="textbox"][contenteditable="true"]',
        ]:
            try:
                body_el = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                )
                break
            except Exception:
                continue
        if not body_el:
            try:
                all_ed = driver.find_elements(By.CSS_SELECTOR, 'div[contenteditable="true"]')
                body_el = all_ed[-1] if all_ed else None
            except Exception:
                pass

        if not body_el:
            print("   Campo de corpo nao encontrado!")
            print("   O navegador continua aberto para acao manual.")
            perguntar("   ENTER apos enviar (o navegador sera fechado)... ")
            try:
                driver.quit()
            except Exception:
                pass
            return False

        # Injeta o HTML no corpo (mesma mecanica que preserva cores)
        try:
            driver.execute_script(
                "var el=arguments[0];el.focus();el.innerHTML=arguments[1];"
                "el.dispatchEvent(new InputEvent('input',{bubbles:true}));"
                "el.dispatchEvent(new Event('change',{bubbles:true}));",
                body_el, html,
            )
            time.sleep(1)
        except Exception as e:
            print("   Erro ao injetar HTML: " + str(e))
            perguntar("   ENTER para continuar... ")
            try:
                driver.quit()
            except Exception:
                pass
            return False

        print("   Email montado no Outlook Web com as cores preservadas!")
        print("   Adicione os destinatarios e clique em Enviar.")
        print()
        print("   *** NAO FECHE O NAVEGADOR ATE ENVIAR ***")
        perguntar("   ENTER apos enviar o email (o navegador sera fechado)... ")
        try:
            driver.quit()
        except Exception:
            pass
        return True

    except Exception as e:
        print("   ERRO ao abrir Outlook Web: " + str(e))
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        return False


# ============================================================
# COPIAR HTML PARA CLIPBOARD (formato CF_HTML)
# Preserva cores e formatacao ao colar no Outlook PWA
# Converte color inline para <font color=""> (Outlook respeita)
# ============================================================
def _converter_cores_para_outlook(html):
    """
    Outlook PWA ignora 'color:' em style inline.
    Workaround: envolve o conteudo de cada tag com <font color="">.
    Converte: <span style="color:#FFD700;...">TEXTO</span>
    Para:     <span style="color:#FFD700;..."><font color="#FFD700">TEXTO</font></span>
    Tambem converte <h1-h3>, <p>, <td>, <div> com color inline.
    """
    import re as _re

    def _add_font_tag(match):
        tag_full = match.group(0)
        # Extrai a cor do style
        cor_match = _re.search(r'color\s*:\s*(#[0-9A-Fa-f]{3,8})', tag_full)
        if not cor_match:
            return tag_full
        cor = cor_match.group(1)
        # Encontra onde a tag fecha (>)
        pos_gt = tag_full.find(">")
        if pos_gt == -1:
            return tag_full
        return tag_full[:pos_gt + 1] + '<font color="' + cor + '">'

    def _close_font_tag(match):
        return "</font>" + match.group(0)

    # Tags que podem ter color inline
    tags = r'(?:span|h[1-6]|p|td|div|a)'

    # Adiciona <font color=""> apos a tag de abertura que tem color no style
    resultado = _re.sub(
        r'<(' + tags + r')([^>]*style\s*=\s*["\'][^"\']*color\s*:\s*#[0-9A-Fa-f]{3,8}[^"\']*["\'][^>]*)>',
        _add_font_tag,
        html
    )

    # Para cada <font color=""> adicionado, precisa fechar antes do </tag>
    # Abordagem simples: adiciona </font> antes de cada </tag> correspondente
    # Isso funciona porque o Outlook e tolerante com font tags extras
    resultado = _re.sub(r'</(' + tags + r')>', _close_font_tag, resultado)

    return resultado


def copiar_html_clipboard(html):
    # Converte cores para formato que Outlook PWA respeita
    html_outlook = _converter_cores_para_outlook(html)

    try:
        import win32clipboard

        # Formato CF_HTML do Windows
        # Requer um header especifico com offsets
        def _make_cf_html(html_content):
            header = (
                "Version:0.9\r\n"
                "StartHTML:{start_html:010d}\r\n"
                "EndHTML:{end_html:010d}\r\n"
                "StartFragment:{start_frag:010d}\r\n"
                "EndFragment:{end_frag:010d}\r\n"
            )
            prefix = "<!--StartFragment-->"
            suffix = "<!--EndFragment-->"

            # Calcula placeholder para determinar tamanho do header
            dummy_header = header.format(start_html=0, end_html=0, start_frag=0, end_frag=0)
            start_html = len(dummy_header.encode("utf-8"))
            start_frag = start_html + len(prefix.encode("utf-8"))
            end_frag = start_frag + len(html_content.encode("utf-8"))
            end_html = end_frag + len(suffix.encode("utf-8"))

            final_header = header.format(
                start_html=start_html,
                end_html=end_html,
                start_frag=start_frag,
                end_frag=end_frag,
            )
            return (final_header + prefix + html_content + suffix).encode("utf-8")

        cf_html_data = _make_cf_html(html_outlook)

        CF_HTML = win32clipboard.RegisterClipboardFormat("HTML Format")
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(CF_HTML, cf_html_data)
        win32clipboard.CloseClipboard()
        return True

    except ImportError:
        # win32clipboard nao disponivel — tenta alternativa com ctypes
        try:
            import ctypes
            from ctypes import wintypes

            def _make_cf_html(html_content):
                header = (
                    "Version:0.9\r\n"
                    "StartHTML:{start_html:010d}\r\n"
                    "EndHTML:{end_html:010d}\r\n"
                    "StartFragment:{start_frag:010d}\r\n"
                    "EndFragment:{end_frag:010d}\r\n"
                )
                prefix = "<!--StartFragment-->"
                suffix = "<!--EndFragment-->"
                dummy_header = header.format(start_html=0, end_html=0, start_frag=0, end_frag=0)
                start_html = len(dummy_header.encode("utf-8"))
                start_frag = start_html + len(prefix.encode("utf-8"))
                end_frag = start_frag + len(html_content.encode("utf-8"))
                end_html = end_frag + len(suffix.encode("utf-8"))
                final_header = header.format(
                    start_html=start_html, end_html=end_html,
                    start_frag=start_frag, end_frag=end_frag,
                )
                return (final_header + prefix + html_content + suffix).encode("utf-8")

            kernel32 = ctypes.windll.kernel32
            user32 = ctypes.windll.user32

            CF_HTML = user32.RegisterClipboardFormatW("HTML Format")
            data = _make_cf_html(html_outlook)

            user32.OpenClipboard(0)
            user32.EmptyClipboard()

            GMEM_MOVEABLE = 0x0002
            h_mem = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data) + 1)
            p_mem = kernel32.GlobalLock(h_mem)
            ctypes.memmove(p_mem, data, len(data))
            ctypes.cast(p_mem, ctypes.POINTER(ctypes.c_char))[len(data)] = b'\x00'
            kernel32.GlobalUnlock(h_mem)
            user32.SetClipboardData(CF_HTML, h_mem)
            user32.CloseClipboard()
            return True

        except Exception as e2:
            print("   AVISO: Nao foi possivel copiar para clipboard: " + str(e2))
            print("   Instale pywin32: pip install pywin32")
            return False

    except Exception as e:
        print("   AVISO: Erro ao copiar para clipboard: " + str(e))
        return False


# ============================================================
# SCREENSHOT (opcional — requer Selenium)
# ============================================================
def salvar_como_imagem(caminho_html, data_str, pasta_saida):
    if not SELENIUM_OK:
        print("\n   Selenium nao encontrado — imagem nao gerada.")
        print("   Execute:  pip install selenium webdriver-manager")
        return None
    print("\n   Gerando imagem do email...")
    try:
        options = Options()
        options.add_argument("--headless")
        service = Service(_obter_gecko_path())
        driver  = webdriver.Firefox(service=service, options=options)
        try:
            driver.get("file:///" + caminho_html.replace(os.sep, "/"))
            time.sleep(1)
            largura = 800
            altura  = driver.execute_script("return document.body.scrollHeight")
            driver.set_window_size(largura, altura + 60)
            time.sleep(0.5)
            ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Nome: PNG-XX.XX.png (data do ONB)
            data_nome = data_str.split("|")[0].strip().replace("/", ".") if data_str else datetime.now().strftime("%d.%m")
            nome_img = "PNG-" + data_nome + ".png"
            caminho_img = os.path.join(pasta_saida, nome_img)
            driver.save_screenshot(caminho_img)
            print("   Imagem salva: " + caminho_img)
            return caminho_img
        finally:
            driver.quit()
    except Exception as e:
        print("   Erro ao gerar imagem: " + str(e))
        return None


# ============================================================
# PASTA DE SAIDA DOS EMAILS (HTML e PNG)
# Salva em: ..\1. HISTÓRICO\1. EMAIL
# ============================================================
def obter_pasta_email():
    # Detecta a pasta do script (email_ambassador.py esta em 2. RECURSOS)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Sobe um nivel (pasta do projeto TOP AMBASSADORS)
    projeto_dir = os.path.dirname(script_dir)
    pasta_email = os.path.join(projeto_dir, "1. HISTÓRICO", "1. EMAIL")
    os.makedirs(pasta_email, exist_ok=True)
    return pasta_email


def obter_pasta_png():
    # PNG vai para 1. HISTÓRICO/1. EMAIL/1. PNG
    script_dir = os.path.dirname(os.path.abspath(__file__))
    projeto_dir = os.path.dirname(script_dir)
    pasta_png = os.path.join(projeto_dir, "1. HISTÓRICO", "1. EMAIL", "1. PNG")
    os.makedirs(pasta_png, exist_ok=True)
    return pasta_png


# ============================================================
# SALVAR HTML E PNG (sem abrir automaticamente)
# ============================================================
def salvar_e_abrir_preview(html, data_str, pasta_saida=None):
    if pasta_saida is None:
        pasta_saida = obter_pasta_email()
    os.makedirs(pasta_saida, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Nome: email-XX.XX.html (data do ONB)
    # data_str pode ser "21.04 | 28.04" — pega a primeira data
    data_nome = data_str.split("|")[0].strip().replace("/", ".") if data_str else datetime.now().strftime("%d.%m")
    nome_html    = "email-" + data_nome + ".html"
    caminho_html = os.path.join(pasta_saida, nome_html)
    with open(caminho_html, "w", encoding="utf-8") as f:
        f.write(html)
    print("\n   HTML salvo: " + caminho_html)
    caminho_img = None
    if EMAIL_GERAR_PNG:
        pasta_png = obter_pasta_png()
        caminho_img = salvar_como_imagem(caminho_html, data_str, pasta_png)
    return caminho_html, caminho_img


# ============================================================
# INTERFACE — chamada pelo AMBASSADORS.py
# ============================================================
def perguntar_e_enviar(ranking_path, data_str="", ids_dict=None):
    titulo("EMAIL  ·  TOP / BOTTOM AMBASSADOR")


    if not os.path.exists(ranking_path):
        erro("Arquivo nao encontrado: " + ranking_path)
        return


    info("Lendo ranking_final.xlsx")
    df_top, df_bot, df_targets = ler_ranking_final(ranking_path)


    if len(df_top) == 0 and len(df_bot) == 0:
        erro("Nenhum dado encontrado. Email cancelado.")
        return


    n_top = df_top["Embaixador"].nunique() if len(df_top) > 0 else 0
    n_bot = df_bot["Embaixador"].nunique() if len(df_bot) > 0 else 0


    ok("TOP " + str(n_top) + "  ·  BOTTOM " + str(n_bot)
       + "  ·  " + str(len(df_targets)) + " targets  ·  ONB " + data_str)
    print()

    # DEMO / PORTFOLIO: NAO baixamos fotos dos embaixadores.
    # O email e gerado sem fotos (os cards usam as iniciais como fallback).
    dica("Montando email (sem fotos - versao demo)...")


    html = montar_html_email(df_top, df_bot, df_targets, data_str, FOTO_DIR)
    caminho_html, caminho_img = salvar_e_abrir_preview(html, data_str)

    # DEMO / PORTFOLIO: o email NAO e enviado pelo Outlook.
    # Apenas mostramos no terminal e oferecemos abrir o HTML de preview.
    print("\n   Concluido!")
    print("   HTML: " + caminho_html)
    if caminho_img:
        print("   PNG:  " + caminho_img)

    # Opcao final: abrir HTML ou fechar
    print()
    print("   " + "=" * 50)
    escolha = perguntar("   ENTER = abrir HTML  |  N = apenas fechar: ").strip().lower()
    if escolha != "n":
        print("   Abrindo HTML...")
        webbrowser.open("file:///" + caminho_html.replace(os.sep, "/"))
    print("   Programa finalizado.")
